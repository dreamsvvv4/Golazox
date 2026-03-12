#!/usr/bin/env python3
"""
Ultrajson.py
----------------
Interfaz gráfica para consultar instalaciones y mostrar resumen o JSON completo.
Permite buscar, filtrar y visualizar datos de instalaciones, así como procesar archivos CSV/XLS/XLSX.

Funciones principales:
- Consultar instalación (resumen o JSON completo)
- Guardar JSON opcionalmente
- Buscar y filtrar texto
- Filtrar por rutas JSON (ej: cu.serialNumber, behaviours[id=autolock].config.timeout)
- Procesar archivos con columnas ID_INSTALL y DS_COUNTRY_SHORT
- Vista en árbol opcional (nodos estructurales y hojas escalares)

Autor: victor.vega
Fecha de modernización: 2026-02-16
"""

# --- Standard Library Imports ---
import os
import sys
import csv
import json
import re
import time
import tempfile
import random
import traceback
import unicodedata
import threading
import queue
from typing import Optional, Any, List, Tuple

# --- Third-Party Imports ---

try:
    import tkinter as tk
    from tkinter import ttk, filedialog, messagebox
    import tkinter.font as tkfont
except ImportError as e:
    print(f"[ERROR] No se pudo importar tkinter: {e}. Asegúrate de tener instalado Tkinter para Python.")
    sys.exit(1)

try:
    import sv_ttk  # Modern ttk theme (Windows 11 style)
except Exception:
    sv_ttk = None

try:
    import requests
except Exception as e:
    requests = None  # type: ignore
    _requests_import_error = e

# Default base URL (keeps backward-compatible behavior: download works out of the box).
# Can be overridden by setting the DCR_BASE_URL environment variable.
DEFAULT_DCR_BASE_URL = "http://mc-dcr.gtm.securitasdirect.local:30150"


def buscar_patrones_en_texto(text_widget: tk.Text, patrones: List[str]) -> List[Tuple[str, str]]:
    """
    Busca y resalta patrones en un widget Text de Tkinter.
    Args:
        text_widget (tk.Text): Widget de texto donde buscar.
        patrones (List[str]): Lista de patrones a buscar.
    Returns:
        List[Tuple[str, str]]: Lista de tuplas (inicio, fin) para cada coincidencia encontrada.
    Si ocurre un error, muestra un mensaje en consola y continúa.
    """
    resultados: List[Tuple[str, str]] = []
    start = '1.0'
    for pat in patrones:
        pos = start
        while True:
            try:
                # Buscar el patrón en el widget de texto
                idx = text_widget.search(pat, pos, stopindex='end', nocase=True)
                if not idx:
                    break
                end = f"{idx}+{len(pat)}c"
                # Resaltar coincidencia
                text_widget.tag_add('search_match', idx, end)
                resultados.append((idx, end))
                pos = end
            except Exception as e:
                print(f"[ERROR] Fallo al buscar el patrón '{pat}': {e}")
                break
    return resultados


def filtrar_lineas_por_patrones(contenido, patrones):
    """Filters lines of text that contain any of the patterns."""
    resultados = []
    for linea in contenido:
        low = linea.lower()
        for pat in patrones:
            if pat.lower() in low:
                resultados.append(linea)
                break
    return resultados


def _buscar_json_local(inst: str, country: str):
    """Search for a local JSON file using a few common names/paths."""
    candidatos = [
        f"{inst}.json",
        f"{country}{inst}.json",
    ]

    try:
        here = os.path.dirname(os.path.abspath(__file__))
        repo_root = os.path.dirname(here)
    except Exception:
        repo_root = os.getcwd()

    download_dir = os.getenv("ULTRAJSON_DOWNLOAD_DIR")
    if not download_dir:
        try:
            download_dir = os.path.join(tempfile.gettempdir(), "Ultrajson")
        except Exception:
            download_dir = None

    search_dirs = [os.getcwd(), repo_root]
    if download_dir:
        search_dirs.append(download_dir)

    for name in list(candidatos):
        for base in search_dirs:
            p = os.path.join(base, name)
            if os.path.isfile(p):
                try:
                    with open(p, "r", encoding="utf-8") as f:
                        data = json.load(f)
                    return data, p
                except Exception:
                    continue
    return None, None


def descargar_instalacion(inst, country="ES"):
    """Download the installation JSON and return (data, filename).
    If network/DNS fails, try loading a local `<inst>.json` as an offline fallback.
    """
    inst = str(inst or "").strip()
    country = str(country or "").strip().upper()

    offline_only = os.getenv("DCR_OFFLINE", "").strip().lower() in ("1", "true", "yes")
    if offline_only:
        data_local, path_local = _buscar_json_local(inst, country)
        if data_local is not None:
            return data_local, path_local
        raise RuntimeError(
            "Offline mode is enabled (DCR_OFFLINE=1) and no local JSON was found. Use Open JSON or place <inst>.json in the download folder."
        )

    if requests is None:
        raise RuntimeError(f"requests is not available: {_requests_import_error}")

    base_url = os.getenv("DCR_BASE_URL", DEFAULT_DCR_BASE_URL)

    if not inst.isdigit():
        raise ValueError("Installation must be numeric")
    if not country.isalpha() or len(country) != 2:
        raise ValueError("Country must be a 2-letter code")

    url = f"{base_url.rstrip('/')}/device-support/device-config-repository/v2.0/installation/{country}{inst}"

    headers = {}
    auth_header = os.getenv("DCR_AUTHORIZATION", "").strip()
    if auth_header:
        headers["Authorization"] = auth_header
    else:
        bearer = os.getenv("DCR_BEARER_TOKEN", "").strip()
        if bearer:
            if not bearer.lower().startswith("bearer "):
                bearer = "Bearer " + bearer
            headers["Authorization"] = bearer

    download_dir = os.getenv("ULTRAJSON_DOWNLOAD_DIR", "").strip()
    if not download_dir:
        download_dir = os.path.join(tempfile.gettempdir(), "Ultrajson")
    os.makedirs(download_dir, exist_ok=True)

    try:
        max_retries = int(os.getenv("DCR_MAX_RETRIES", "3"))
    except Exception:
        max_retries = 3
    try:
        backoff_base = float(os.getenv("DCR_BACKOFF_BASE", "0.8"))
    except Exception:
        backoff_base = 0.8
    try:
        backoff_max = float(os.getenv("DCR_BACKOFF_MAX", "10"))
    except Exception:
        backoff_max = 10.0

    last_exc = None
    for attempt in range(max_retries + 1):
        try:
            respuesta = requests.get(url, timeout=10, headers=headers)
            if respuesta.status_code in (429, 503):
                raise requests.exceptions.HTTPError(
                    f"HTTP {respuesta.status_code}", response=respuesta
                )
            respuesta.raise_for_status()
            data = respuesta.json()
            filename = os.path.join(download_dir, f"{inst}.json")
            with open(filename, "w", encoding="utf-8") as f:
                json.dump(data, f, indent=4)
            return data, filename
        except (requests.exceptions.Timeout, requests.exceptions.ConnectionError) as e:
            last_exc = e
            if attempt >= max_retries:
                break
        except requests.exceptions.HTTPError as e:
            last_exc = e
            status = None
            try:
                status = e.response.status_code if e.response is not None else None
            except Exception:
                status = None
            retryable = status in (429, 503)
            if not retryable or attempt >= max_retries:
                break
        except requests.exceptions.RequestException as e:
            last_exc = e
            break

        sleep_s = min(backoff_max, (backoff_base * (2 ** attempt)) + random.uniform(0, 0.25))
        time.sleep(max(0.0, sleep_s))

    data_local, path_local = _buscar_json_local(inst, country)
    if data_local is not None:
        return data_local, path_local
    raise RuntimeError(f"Error downloading installation: {last_exc}")


def generar_resumen(data):
    """Return a pretty-printed JSON string for the given object."""
    return json.dumps(data, indent=2, ensure_ascii=False)
PAISES = ["ES","PT","IT","FR","DE","GB","AR","CL","MX","BR"]
COUNTRY_MAP = {
    'ESP':'ES','ES':'ES','ITA':'IT','IT':'IT','FRA':'FR','FR':'FR','DEU':'DE','DE':'DE','GBR':'GB','UK':'GB','GB':'GB',
    'ARG':'AR','AR':'AR','CHL':'CL','CL':'CL','MEX':'MX','MX':'MX','BRA':'BR','BR':'BR','PRT':'PT','POR':'PT','PT':'PT'
}


def map_country(code: str) -> Optional[str]:
    if not code:
        return None
    return COUNTRY_MAP.get(code.strip().upper())

## === JSON extraction and filtering logic (outside the class) ===
def extract_json_path(data: Any, ruta: str) -> Any:
    """Extracts the value(s) of a JSON path (see original docstring)."""
    if data is None or not ruta:
        return None
    ruta = ruta.strip()
    # Backward compatible parsing (old Spanish prefixes + newer English ones)
    low = ruta.lower()
    if low.startswith('### ruta:') or low.startswith('### path:'):
        ruta = ruta.split(':', 1)[1].strip()
    elif low.startswith('ruta:') or low.startswith('path:'):
        ruta = ruta.split(':', 1)[1].strip()
    if ' =' in ruta:
        ruta = ruta.split(' =',1)[0].strip()
    tokens = []
    buf = ''
    i = 0
    while i < len(ruta):
        c = ruta[i]
        if c == '.':
            if buf:
                tokens.append(buf); buf=''
            i += 1; continue
        if c == '[':
            if buf:
                tokens.append(buf); buf=''
            j = i+1
            while j < len(ruta) and ruta[j] != ']': j += 1
            contenido = ruta[i+1:j]
            i = j+1
            contenido = contenido.strip()
            if contenido == '' or contenido == '*':
                tokens.append('*')
                continue
            # Selectors:
            # - [field=value]     -> first match in list (backward compatible)
            # - [?field=value]    -> ALL matches in list (returns list)
            if '=' in contenido:
                campo, _, val = contenido.partition('=')
                campo = campo.strip()
                val = val.strip()
                if campo.startswith('?'):
                    tokens.append(('selector_all', campo[1:].strip(), val))
                else:
                    tokens.append(('selector', campo, val))
            else:
                try:
                    tokens.append(int(contenido))
                except ValueError:
                    tokens.append(contenido)
            continue
        buf += c; i += 1
    if buf: tokens.append(buf)
    def aplicar(nodo, restante):
        if not restante:
            return nodo
        t = restante[0]
        tail = restante[1:]
        if t == '*':
            if isinstance(nodo, list):
                resultados = []
                for el in nodo:
                    val = aplicar(el, tail)
                    if val is not None:
                        resultados.append(val)
                return resultados if resultados else None
            else:
                return None
        if isinstance(t, tuple) and t[0] == 'selector':
            campo, valor = t[1], t[2]
            if isinstance(nodo, list):
                vlow = valor.lower()
                for el in nodo:
                    if isinstance(el, dict):
                        comp = el.get(campo)
                        if isinstance(comp, (str,int,float,bool)) and str(comp).lower() == vlow:
                            return aplicar(el, tail)
                return None
            return None
        if isinstance(t, tuple) and t[0] == 'selector_all':
            campo, valor = t[1], t[2]
            if isinstance(nodo, list):
                vlow = str(valor).lower()
                matches = []
                for el in nodo:
                    if not isinstance(el, dict):
                        continue
                    comp = el.get(campo)
                    if isinstance(comp, (str, int, float, bool)) and str(comp).lower() == vlow:
                        val = aplicar(el, tail)
                        if val is not None:
                            matches.append(val)
                return matches if matches else None
            return None
        if isinstance(t, int):
            if isinstance(nodo, list) and 0 <= t < len(nodo):
                return aplicar(nodo[t], tail)
            return None
        if isinstance(nodo, dict):
            if t in nodo:
                return aplicar(nodo[t], tail)
            low = t.lower()
            for k in nodo.keys():
                if k.lower() == low:
                    return aplicar(nodo[k], tail)
            return None
        return None
    resultado = aplicar(data, tokens)
    if isinstance(resultado, list):
        if not resultado:
            return None
        flat = []
        for el in resultado:
            if isinstance(el, list):
                flat.extend(el)
            else:
                flat.append(el)
        return flat
    if resultado is not None:
        return resultado
    # Fallback: partial path search
    if '.' in ruta:
        sufijo = ruta.strip()
        sufijo_wants_paths = sufijo.startswith('.')
        coincidencias = []
        coincidencias_con_ruta: List[Tuple[str, Any]] = []

        def ruta_str(parts: List[str], keep_indices: bool) -> str:
            s = ""
            for part in parts:
                if part.startswith('['):
                    if keep_indices:
                        s += part
                    continue
                if s:
                    s += "."
                s += str(part)
            return s

        def ruta_str_indices(parts: List[str]) -> str:
            s = ""
            for part in parts:
                if part.startswith('['):
                    s += part
                else:
                    if s:
                        s += "."
                    s += str(part)
            return s

        def recorrer(nodo, path_parts):
            if isinstance(nodo, dict):
                for k,v in nodo.items():
                    recorrer(v, path_parts + [k])
            elif isinstance(nodo, list):
                for i,el in enumerate(nodo):
                    recorrer(el, path_parts + [f"[{i}]"])
            ruta_actual = ruta_str(path_parts, keep_indices=False)
            if ruta_actual.endswith(sufijo):
                coincidencias.append(nodo)
                if sufijo_wants_paths:
                    coincidencias_con_ruta.append((ruta_str_indices(path_parts), nodo))
        try:
            recorrer(data, [])
        except Exception:
            pass
        if coincidencias:
            if sufijo_wants_paths:
                return coincidencias_con_ruta
            return coincidencias if len(coincidencias) > 1 else coincidencias[0]
    return None

def extract_all_json_paths(obj: Any, ruta_actual: List[str]=None) -> List[Tuple[str, Any]]:
    """Traverses the JSON and returns a list of (path, value)."""
    if ruta_actual is None:
        ruta_actual = []
    resultados = []
    def ruta_str(ruta):
        s = ""
        for part in ruta:
            if part.startswith("["):
                s += part
            else:
                if s:
                    s += "."
                s += part
        return s
    if isinstance(obj, dict):
        for k, v in obj.items():
            resultados += extract_all_json_paths(v, ruta_actual + [str(k)])
    elif isinstance(obj, list):
        for i, v in enumerate(obj):
            resultados += extract_all_json_paths(v, ruta_actual + [f"[{i}]"])
    elif isinstance(obj, (str, int, float, bool)) or obj is None:
        resultados.append((ruta_str(ruta_actual), obj))
    return resultados

def find_key_matches(obj: Any, key_name: str, ruta_actual: List[str] = None) -> List[Tuple[str, Any]]:
    """Find all occurrences of a dict key anywhere in a JSON-like structure.

    Returns a list of (path, value) where path is dot/bracket notation.
    Matching is case-insensitive on the key name.
    """
    if ruta_actual is None:
        ruta_actual = []
    if obj is None or not key_name:
        return []
    key_low = key_name.strip().lower()
    if not key_low:
        return []

    def ruta_str(ruta: List[str]) -> str:
        s = ""
        for part in ruta:
            if part.startswith("["):
                s += part
            else:
                if s:
                    s += "."
                s += part
        return s

    resultados: List[Tuple[str, Any]] = []
    if isinstance(obj, dict):
        for k, v in obj.items():
            k_str = str(k)
            child_path = ruta_actual + [k_str]
            if k_str.lower() == key_low:
                resultados.append((ruta_str(child_path), v))
            # Keep traversing nested values
            if isinstance(v, (dict, list)):
                resultados.extend(find_key_matches(v, key_name, child_path))
        return resultados
    if isinstance(obj, list):
        for i, v in enumerate(obj):
            child_path = ruta_actual + [f"[{i}]"]
            if isinstance(v, (dict, list)):
                resultados.extend(find_key_matches(v, key_name, child_path))
        return resultados
    return []


class InstalacionGUI:
    APP_NAME = "Ultrajson"
    APP_VERSION = "1.1.0"

    def _add_experimental_delete_button(self, edit_win, old_val, exists, msg_var, exp_path, path_var, new_value_var, _refresh_current_value, _invalidate_check, _write):
        """
        Adds the 'Eliminar objeto del array' button and its logic to the given edit_win.
        All required variables must be passed in from the context where this is called.
        """
        frm_del = ttk.Frame(edit_win)
        frm_del.pack(fill="x", padx=8, pady=2)
        ttk.Label(frm_del, text="Eliminar por índice:").pack(side="left")
        idx_var = tk.StringVar()
        ent_idx = ttk.Entry(frm_del, textvariable=idx_var, width=4)
        ent_idx.pack(side="left", padx=(2, 8))
        ttk.Label(frm_del, text="o por _id:").pack(side="left")
        id_var = tk.StringVar()
        ent_id = ttk.Entry(frm_del, textvariable=id_var, width=12)
        ent_id.pack(side="left", padx=(2, 8))

        def eliminar_objeto():
            arr = []
            if exists and isinstance(old_val, list):
                arr = list(old_val)
            if not arr:
                msg_var.set("El array experimental está vacío.")
                return
            idx_txt = idx_var.get().strip()
            id_txt = id_var.get().strip()
            eliminado = False
            # Eliminar por índice
            if idx_txt.isdigit():
                idx = int(idx_txt)
                if 0 <= idx < len(arr):
                    arr.pop(idx)
                    eliminado = True
                else:
                    msg_var.set(f"Índice fuera de rango (0-{len(arr)-1})")
                    return
            # Eliminar por _id
            elif id_txt:
                n = len(arr)
                arr = [o for o in arr if not (isinstance(o, dict) and str(o.get('_id','')) == id_txt)]
                if len(arr) < n:
                    eliminado = True
                else:
                    msg_var.set(f"No se encontró ningún objeto con _id='{id_txt}'")
                    return
            else:
                msg_var.set("Introduce un índice o un _id para eliminar.")
                return
            # Actualizar experimental
            ptrs = exp_path.strip("/").split("/")
            obj = self._last_json_data
            for p in ptrs[:-1]:
                p = self._unescape_json_pointer_token(p)
                if isinstance(obj, dict):
                    obj = obj.setdefault(p, {})
                else:
                    msg_var.set("No se pudo navegar hasta experimental (estructura inesperada)")
                    return
            last = self._unescape_json_pointer_token(ptrs[-1])
            obj[last] = arr
            msg_var.set("Objeto eliminado del array experimental. Cierra la ventana y usa 'Check' y 'Send' para aplicar.")
            # Refrescar valor actual en la ventana principal
            try:
                _refresh_current_value()
            except Exception:
                pass
            try:
                if path_var.get().strip().replace("/","").lower() == "experimental":
                    new_value_var.set(json.dumps(arr, ensure_ascii=False, indent=None))
            except Exception:
                pass
            try:
                _invalidate_check("experimental delete object")
                _write("Experimental actualizado. Haz 'Check' para validar y enviar si es necesario.")
            except Exception:
                pass

        btn_del = ttk.Button(frm_del, text="Eliminar objeto del array", command=eliminar_objeto)
        btn_del.pack(side="left", padx=(8, 0))

    def _add_tooltip(self, widget, text):
        tooltip = tk.Toplevel(widget)
        tooltip.withdraw()
        tooltip.overrideredirect(True)
        label = tk.Label(tooltip, text=text, background='#ffffe0', relief='solid', borderwidth=1, font=('Consolas', 9))
        label.pack(ipadx=4, ipady=2)
        def show(event):
            x = event.x_root + 10
            y = event.y_root + 10
            tooltip.geometry(f'+{x}+{y}')
            tooltip.deiconify()
        def hide(event):
            tooltip.withdraw()
        widget.bind('<Enter>', show)
        widget.bind('<Leave>', hide)
        widget.bind('<ButtonPress>', hide)
        return tooltip

    # --- Calls to external functions for JSON logic ---
    def extraer_valor_ruta(self, data, ruta):
        return extract_json_path(data, ruta)

    def extraer_rutas_json(self, obj, ruta_actual=None):
        return extract_all_json_paths(obj, ruta_actual)

    def __init__(self, root: tk.Tk):
        self.root = root
        root.title(f"{self.APP_NAME} v{self.APP_VERSION}")
        # Apply modern theme if available
        try:
            if sv_ttk:
                # Default to Light to match typical Windows appearance.
                sv_ttk.set_theme("light")
        except Exception:
            pass

        self._apply_premium_styles()

        self.install_var = tk.StringVar()
        self.country_var = tk.StringVar(value="ES")
        self.compare_install_var = tk.StringVar()
        self.compare_country_var = tk.StringVar(value="ES")
        self.view_mode = tk.StringVar(value="resumen")
        # Default to unchecked to avoid unintended implicit actions
        self.save_var = tk.BooleanVar(value=False)
        # Auto export removed
        self.status_var = tk.StringVar(value="Ready")
        self.theme_var = tk.StringVar(value="light" if sv_ttk else "native")
        self._q = queue.Queue()
        self._last_json_data = None
        self._original_text = None
        # Batch-to-CSV streaming removed (was confusing / non-functional in practice)
        self._compare_active = False
        self.compare_only_diffs_var = tk.BooleanVar(value=False)
        self._last_compare_state = None  # (base_data, data_b, changed, added, removed, inst_a, country_a, inst_b, country_b)
        # Batch (Filter CSV) cancellation support
        self._batch_cancel_event = threading.Event()
        self._batch_running = False
        self.btn_stop = None
        # Initialize tree view widgets placeholders (real widgets created in _build_ui under a fixed container)
        self.tree_frame = None
        self.tree_sub = None
        self.tree = None
        self.tree_scroll = None
        self._tree_path_map = {}
        self._tree_original_labels = {}
        self._tree_search_results = []
        self._tree_search_pos = -1
        # Initialize search structures in text mode (avoids AttributeError if "Next" is pressed before searching)
        self._search_results = []
        self._search_pos = -1
        self._build_ui()
        # Show Quick Start only on startup (Clear should leave an empty panel).
        self._show_welcome()
        # Menubar removed by request (use buttons + shortcuts instead)
        self._update_action_states()
        # Global shortcuts
        self.root.bind('<Control-f>', lambda e: self._focus_search())
        self.root.bind('<F3>', lambda e: self.buscar_siguiente())
        self.root.bind('<Shift-F3>', lambda e: self.buscar_anterior())
        self.root.bind('<Control-o>', lambda e: self.cargar_json_local())
        self.root.bind('<Escape>', lambda e: self.limpiar_busqueda())
        self.root.after(250, self._poll_queue)

    def _show_welcome(self):
        """Write a short onboarding guide into the output panel for first-time users."""
        try:
            if self._last_json_data is not None:
                return
            current = self.text.get('1.0', 'end-1c')
            if current.strip():
                return
        except Exception:
            return

        welcome = [
            "Ultrajson — Quick Start",
            "",
            "1) Installation + Country → Query",
            "   - View: Summary (readable) or JSON (full)",
            "   - Save JSON: keep the downloaded file on disk",
            "",
            "2) Open JSON (offline): click 'Open JSON' or press Ctrl+O",
            "",
            "3) Search vs Filter",
            "   - Search: highlights matches (F3 / Shift+F3)",
            "   - Filter: reduces panel to matching lines/paths",
            "",
            "4) Tree View: browse JSON structure",
            "",
            "5) Compare",
            "   - Fill 'Compare to' → Compare",
            "   - Only diffs: show only changed/added/removed fields",
            "   - Close Compare: back to normal view",
            "",
            "Path examples (Filter / Extract Path Value):",
            "   cu.serialNumber",
            "   nodes[0].serialNumber",
            "   behaviours[id=autolock].config.timeout",
            "   behaviours[?id=autolock].config.timeout   (all matches)",
            "",
            "Tip: See README.md for more examples",
        ]
        try:
            self.text.insert('1.0', "\n".join(welcome) + "\n")
            self.text.see('1.0')
        except Exception:
            pass

    def _normalize_key(self, s: str) -> str:
        try:
            s = unicodedata.normalize('NFKD', s)
            s = ''.join(ch for ch in s if not unicodedata.combining(ch))
        except Exception:
            pass
        return (s or '').strip().casefold()

    def _pick_key(self, d: dict, *candidates: str):
        if not isinstance(d, dict):
            return None
        want = {self._normalize_key(c) for c in candidates}
        for k, v in d.items():
            if self._normalize_key(str(k)) in want:
                return v
        return None

    def _get_history_file_candidates(self) -> List[str]:
        env = os.environ.get('ULTRAJSON_HISTORY_FILE')
        base_dir = os.path.abspath(os.path.join(os.path.dirname(os.path.abspath(__file__)), os.pardir))
        candidates = []
        if env:
            candidates.append(env)
        candidates.extend([
            os.path.join(os.getcwd(), 'historico_instalaciones.json'),
            os.path.join(base_dir, 'historico_instalaciones.json'),
        ])
        # De-dup while keeping order
        out = []
        seen = set()
        for p in candidates:
            p2 = os.path.abspath(os.path.expanduser(str(p)))
            if p2 not in seen:
                out.append(p2)
                seen.add(p2)
        return out

    def _load_installation_history_rows(self) -> Tuple[List[Tuple[str, str, str]], Optional[str]]:
        """Return [(installation, serial, description)], plus the path used (or None)."""
        path_used = None
        for p in self._get_history_file_candidates():
            if os.path.isfile(p):
                path_used = p
                break
        if not path_used:
            return ([], None)

        try:
            with open(path_used, 'r', encoding='utf-8') as f:
                data = json.load(f)
        except Exception:
            return ([], path_used)

        # Supported shapes:
        # - {"resumen": [ ... ]}
        # - {"instalaciones": [ ... ]}
        # - [ ... ]
        items = None
        if isinstance(data, dict):
            if isinstance(data.get('resumen'), list):
                items = data.get('resumen')
            elif isinstance(data.get('instalaciones'), list):
                items = data.get('instalaciones')
        elif isinstance(data, list):
            items = data
        if not isinstance(items, list):
            return ([], path_used)

        rows: List[Tuple[str, str, str]] = []
        for it in items:
            if isinstance(it, (int, float)):
                inst = str(int(it))
                rows.append((inst, '', ''))
                continue
            if not isinstance(it, dict):
                continue

            inst_v = self._pick_key(
                it,
                'Numero de instalacion',
                'Número de instalación',
                'Numero de instalación',
                'Número de instalacion',
                'Installation',
                'Installation number',
                'installationNumber',
            )
            serial_v = self._pick_key(it, 'Serial Number', 'Serial', 'serialNumber')
            desc_v = self._pick_key(it, 'Descripcion', 'Descripción', 'Description', 'desc')

            if inst_v is None:
                continue
            inst = str(inst_v).strip()
            if not inst:
                continue
            serial = '' if serial_v is None else str(serial_v).strip()
            desc = '' if desc_v is None else str(desc_v).strip()
            rows.append((inst, serial, desc))

        # Stable sort by installation number (numeric if possible)
        def sort_key(r):
            try:
                return (0, int(re.sub(r'\D+', '', r[0]) or '0'))
            except Exception:
                return (1, r[0])

        rows = sorted(rows, key=sort_key)
        return (rows, path_used)

    def open_installation_history(self):
        rows, path_used = self._load_installation_history_rows()
        if not rows and not path_used:
            messagebox.showinfo(
                'Installation History',
                "Couldn't find historico_instalaciones.json.\n\n"
                "Looked in:\n- current folder\n- project root\n\n"
                "Tip: set ULTRAJSON_HISTORY_FILE to a custom path.",
            )
            return
        if not rows:
            messagebox.showwarning(
                'Installation History',
                f"Found history file but couldn't read rows:\n{path_used}",
            )
            return

        win = tk.Toplevel(self.root)
        win.title('Installation History')
        win.transient(self.root)
        try:
            win.grab_set()
        except Exception:
            pass

        header = ttk.Frame(win)
        header.pack(fill='x', padx=10, pady=(10, 6))
        ttk.Label(header, text='Search:').pack(side='left')
        q_var = tk.StringVar()
        ent_q = ttk.Entry(header, textvariable=q_var, width=40)
        ent_q.pack(side='left', padx=(6, 0), fill='x', expand=True)
        if path_used:
            ttk.Label(header, text=os.path.basename(path_used)).pack(side='right')

        body = ttk.Frame(win)
        body.pack(fill='both', expand=True, padx=10, pady=(0, 10))
        body.rowconfigure(0, weight=1)
        body.columnconfigure(0, weight=1)

        tree = ttk.Treeview(body, columns=('inst', 'serial', 'desc'), show='headings', selectmode='browse', height=16)
        tree.heading('inst', text='Installation')
        tree.heading('serial', text='Serial')
        tree.heading('desc', text='Description')
        tree.column('inst', width=110, anchor='w', stretch=False)
        tree.column('serial', width=110, anchor='w', stretch=False)
        tree.column('desc', width=420, anchor='w', stretch=True)
        y = ttk.Scrollbar(body, orient='vertical', command=tree.yview)
        tree.configure(yscrollcommand=y.set)
        tree.grid(row=0, column=0, sticky='nsew')
        y.grid(row=0, column=1, sticky='ns')

        btns = ttk.Frame(win)
        btns.pack(fill='x', padx=10, pady=(0, 10))
        btn_use = ttk.Button(btns, text='Use selection', command=lambda: choose())
        btn_use.pack(side='right')
        ttk.Button(btns, text='Close', command=win.destroy).pack(side='right', padx=(0, 6))

        def repopulate(*_):
            q = (q_var.get() or '').strip().casefold()
            for it in tree.get_children():
                tree.delete(it)
            for inst, serial, desc in rows:
                if q and (q not in inst.casefold() and q not in serial.casefold() and q not in desc.casefold()):
                    continue
                tree.insert('', 'end', values=(inst, serial, desc))

            children = tree.get_children()
            if children:
                tree.selection_set(children[0])

        def choose(event=None):
            sel = tree.selection()
            if not sel:
                return
            values = tree.item(sel[0], 'values')
            if not values:
                return
            inst = str(values[0]).strip()
            if inst:
                self.install_var.set(inst)
                try:
                    if getattr(self, '_ent_install_main', None):
                        self._ent_install_main.focus_set()
                        self._ent_install_main.icursor('end')
                except Exception:
                    pass
            win.destroy()
            # Lanzar query automáticamente al seleccionar desde el historial
            self.root.after(0, self.consultar)

        ent_q.bind('<KeyRelease>', repopulate)
        ent_q.bind('<Return>', lambda e: choose())
        tree.bind('<Double-1>', choose)
        tree.bind('<Return>', choose)

        repopulate()
        ent_q.focus_set()

    def _update_action_states(self):
        """Enable/disable actions based on current app state."""
        has_json = self._last_json_data is not None
        in_compare = bool(getattr(self, '_compare_active', False))
        has_compare_result = bool(getattr(self, '_last_compare_state', None))
        batch_running = bool(getattr(self, '_batch_running', False))

        def safe_config(widget, **kwargs):
            try:
                widget.configure(**kwargs)
            except Exception:
                pass

        if hasattr(self, 'btn_show_paths'):
            safe_config(self.btn_show_paths, state=('normal' if has_json else 'disabled'))
        if hasattr(self, 'btn_extract_path'):
            safe_config(self.btn_extract_path, state=('normal' if has_json else 'disabled'))
        if hasattr(self, 'btn_remote_edit'):
            safe_config(self.btn_remote_edit, state=('normal' if has_json else 'disabled'))

        if hasattr(self, 'btn_compare'):
            safe_config(self.btn_compare, state=('normal' if (has_json and not batch_running) else 'disabled'))
        if hasattr(self, 'btn_close_compare'):
            safe_config(self.btn_close_compare, state=('normal' if in_compare else 'disabled'))
        if hasattr(self, 'chk_only_diffs'):
            safe_config(self.chk_only_diffs, state=('normal' if (in_compare and has_compare_result) else 'disabled'))

    # === Compare scroll sync ===
    def _scroll_y(self, *args):
        """Scrollbar callback: scroll left pane and (if Compare active) right pane too."""
        try:
            self.text.yview(*args)
        except Exception:
            pass
        if getattr(self, '_compare_active', False):
            try:
                self.compare_text.yview(*args)
            except Exception:
                pass

    def _scroll_x(self, *args):
        """Scrollbar callback: scroll left pane and (if Compare active) right pane too."""
        try:
            self.text.xview(*args)
        except Exception:
            pass
        if getattr(self, '_compare_active', False):
            try:
                self.compare_text.xview(*args)
            except Exception:
                pass

    def _on_mousewheel_compare(self, event):
        """Mouse wheel: when Compare is active, scroll both panes together."""
        if not getattr(self, '_compare_active', False):
            return
        try:
            delta = int(event.delta)
        except Exception:
            return "break"
        if delta == 0:
            return "break"
        # Windows: delta is multiple of 120. Negative delta -> scroll down.
        units = int(-delta / 120)
        if units == 0:
            units = -1 if delta > 0 else 1
        try:
            self.text.yview_scroll(units, 'units')
        except Exception:
            pass
        try:
            self.compare_text.yview_scroll(units, 'units')
        except Exception:
            pass
        return "break"

    def _on_shift_mousewheel_compare(self, event):
        """Shift+Mouse wheel: horizontal scroll both panes in Compare mode."""
        if not getattr(self, '_compare_active', False):
            return
        try:
            delta = int(event.delta)
        except Exception:
            return "break"
        if delta == 0:
            return "break"
        units = int(-delta / 120)
        if units == 0:
            units = -1 if delta > 0 else 1
        try:
            self.text.xview_scroll(units, 'units')
        except Exception:
            pass
        try:
            self.compare_text.xview_scroll(units, 'units')
        except Exception:
            pass
        return "break"

    def _apply_premium_styles(self):
        """Apply subtle premium UI tweaks (fonts/padding) without hard-coding new colors."""
        try:
            default_font = tkfont.nametofont("TkDefaultFont")
            default_font.configure(family="Segoe UI", size=10)
            text_font = tkfont.nametofont("TkTextFont")
            text_font.configure(family="Segoe UI", size=10)
            fixed_font = tkfont.nametofont("TkFixedFont")
            fixed_font.configure(family="Consolas", size=10)
        except Exception:
            pass

        try:
            style = ttk.Style()
            # Make controls feel less cramped
            style.configure('TButton', padding=(10, 6))
            style.configure('TEntry', padding=(6, 4))
            style.configure('TCombobox', padding=(6, 4))
            style.configure('TCheckbutton', padding=(4, 2))
            style.configure('TRadiobutton', padding=(4, 2))

            # Slightly bolder default label feel
            style.configure('TLabel', padding=(2, 2))

            # Dark output style for Tree View (to match the dark output panel).
            # Only applied when Tree View is enabled.
            try:
                style.configure('Dark.Treeview', background='#000000', fieldbackground='#000000', foreground='#e6e6e6')
                style.map('Dark.Treeview', background=[('selected', '#333366')], foreground=[('selected', '#ffffff')])
            except Exception:
                pass

            # sv-ttk on Windows can ignore custom Treeview styles for the empty area.
            # Configure the base 'Treeview' style as well so the full widget paints dark.
            try:
                style.configure('Treeview', background='#000000', fieldbackground='#000000', foreground='#e6e6e6')
                style.map('Treeview', background=[('selected', '#333366')], foreground=[('selected', '#ffffff')])
            except Exception:
                pass

            # Fallback for themes without Accent.TButton: map it to TButton
            try:
                if 'Accent.TButton' not in style.theme_names():
                    # theme_names isn't styles; keep the fallback anyway
                    pass
            except Exception:
                pass
        except Exception:
            pass

    def _ensure_treeview_style(self, dark: bool):
        """Apply the desired style to the Treeview without affecting the rest of the UI."""
        try:
            # Use the base 'Treeview' style; we configure it to dark for consistent painting.
            self.tree.configure(style='Treeview')
        except Exception:
            pass

    def _show_tree_placeholder(self):
        """Show a placeholder entry so the Treeview paints its background even when empty."""
        try:
            for it in self.tree.get_children():
                self.tree.delete(it)
            self.tree.insert('', 'end', text='No JSON loaded. Run a query or open a JSON file.', open=False)
        except Exception:
            pass

    def _build_ui(self):
        self.tree_visible = tk.BooleanVar(value=False)
        # Top row
        top = ttk.Frame(self.root); top.pack(fill='x', padx=10, pady=6)
        lbl_install = ttk.Label(top, text="Installation:"); lbl_install.grid(row=0,column=0,sticky='w')
        install_box = ttk.Frame(top)
        install_box.grid(row=0, column=1, sticky='w')
        ent_install = ttk.Entry(install_box, textvariable=self.install_var, width=14)
        ent_install.pack(side='left')
        ent_install.bind('<Return>', lambda e: self.consultar())
        self._ent_install_main = ent_install
        btn_history = ttk.Button(install_box, text='History…', command=self.open_installation_history)
        btn_history.pack(side='left', padx=(6, 0))
        self._add_tooltip(ent_install, "Enter the installation number to query.")
        self._add_tooltip(btn_history, "Pick an installation from historico_instalaciones.json.")
        lbl_country = ttk.Label(top, text="Country:"); lbl_country.grid(row=0,column=2,padx=(12,0),sticky='w')
        self.country_combo = ttk.Combobox(top, values=PAISES, width=6, state='readonly'); self.country_combo.set(self.country_var.get())
        self.country_combo.grid(row=0,column=3,sticky='w')
        self.country_combo.bind('<<ComboboxSelected>>', lambda e: self.country_var.set(self.country_combo.get()))
        self._add_tooltip(self.country_combo, "Select the country code for the installation.")
        chk_save = ttk.Checkbutton(top, text="Save JSON", variable=self.save_var, command=self._update_download_button); chk_save.grid(row=0,column=4,padx=(12,0))
        self._add_tooltip(chk_save, "Save the downloaded JSON file locally.")
        mode = ttk.Frame(top); mode.grid(row=0,column=5,padx=(12,0))
        lbl_view = ttk.Label(mode,text="View:"); lbl_view.pack(side='left')
        rb_summary = ttk.Radiobutton(mode,text="Summary",value="resumen",variable=self.view_mode, command=self._update_download_button); rb_summary.pack(side='left')
        rb_json = ttk.Radiobutton(mode,text="JSON",value="json",variable=self.view_mode, command=self._update_download_button); rb_json.pack(side='left')
        self._add_tooltip(rb_summary, "Show a summary of the installation data.")
        self._add_tooltip(rb_json, "Show the full JSON data for the installation.")
        # Primary action (accent style if supported by theme)
        btn_query = ttk.Button(top, text="Query", command=self.consultar, style='Accent.TButton')
        btn_query.grid(row=0, column=6, padx=6)
        self._add_tooltip(btn_query, "Query the installation and display results.")
        btn_clear = ttk.Button(top,text="Clear",command=self.limpiar); btn_clear.grid(row=0,column=7)
        self._add_tooltip(btn_clear, "Clear everything (panel + loaded JSON + compare state).")
        btn_filter_csv = ttk.Button(top, text="Filter CSV", command=self.buscar_en_lista)
        btn_filter_csv.grid(row=0, column=8, padx=(6, 0))
        self._add_tooltip(
            btn_filter_csv,
            "Batch search installations from CSV/XLS/XLSX. "
            "Header columns supported: ID_INSTALL / INSTALL / ID / INSTALLATION and DS_COUNTRY_SHORT / COUNTRY / PAIS. "
            "If there is no header, the first two columns are used (country + installation)."
        )
        # Extract path value button moved below
        # Search bar
        search = ttk.Frame(self.root); search.pack(fill='x', padx=10, pady=(0,6))
        lbl_search = ttk.Label(search,text="Search:"); lbl_search.pack(side='left')
        self.search_var = tk.StringVar(); self.ent_search = ttk.Entry(search,textvariable=self.search_var,width=32); self.ent_search.pack(side='left',padx=(4,4))
        self._add_tooltip(self.ent_search, "Enter text, path, or pattern to search/filter.")
        btn_search = ttk.Button(search,text="Search",command=self.buscar); btn_search.pack(side='left')
        self._add_tooltip(btn_search, "Search for text or pattern in the results panel.")
        btn_next = ttk.Button(search,text="Next",command=self.buscar_siguiente); btn_next.pack(side='left',padx=(4,0))
        self._add_tooltip(btn_next, "Go to the next search match.")
        btn_prev = ttk.Button(search,text="Previous",command=self.buscar_anterior); btn_prev.pack(side='left',padx=(4,0))
        self._add_tooltip(btn_prev, "Go to the previous search match.")
        btn_clear_search = ttk.Button(search,text="Reset View",command=self.limpiar_busqueda); btn_clear_search.pack(side='left',padx=(4,0))
        self._add_tooltip(btn_clear_search, "Clear search highlights and undo Filter/Paths/Extract view changes.")
        btn_filter = ttk.Button(search,text="Filter",command=self.filtrar_resultados); btn_filter.pack(side='left',padx=(8,0))
        self._add_tooltip(btn_filter, "Filter the panel to only matching lines (text, path, or pattern).")
        chk_tree = ttk.Checkbutton(search,text="Tree View",variable=self.tree_visible, command=self.toggle_tree_view); chk_tree.pack(side='left',padx=(12,0))
        self._add_tooltip(chk_tree, "Toggle between text and tree view for JSON data.")
        # Center area with a fixed container to avoid layout jumps when toggling views
        self.content_frame = ttk.Frame(self.root)
        self.content_frame.pack(fill='both', expand=True, padx=10, pady=(6,6))
        self.content_frame.rowconfigure(0, weight=1)
        self.content_frame.columnconfigure(0, weight=1)
        self.content_frame.columnconfigure(1, weight=1)
        # Text panel + Tree panel are stacked in the SAME grid cell; we switch using tkraise
        self.text_frame = ttk.Frame(self.content_frame)
        self.text_frame.grid(row=0, column=0, columnspan=2, sticky='nsew')
        # Dark output panel (user preference): keep UI theme light, but output area black.
        self.text = tk.Text(
            self.text_frame,
            wrap='none',
            font=('Consolas', 10),
            background='#000',
            foreground='#e6e6e6',
            insertbackground='#ffffff',
        )
        yscroll = ttk.Scrollbar(self.text_frame, orient='vertical', command=self._scroll_y)
        xscroll = ttk.Scrollbar(self.text_frame, orient='horizontal', command=self._scroll_x)
        self.text.configure(yscrollcommand=yscroll.set, xscrollcommand=xscroll.set)
        self.text_frame.rowconfigure(0,weight=1); self.text_frame.columnconfigure(0,weight=1)
        self.text.grid(row=0,column=0,sticky='nsew'); yscroll.grid(row=0,column=1,sticky='ns'); xscroll.grid(row=1,column=0,sticky='ew')
        self._add_tooltip(self.text, "Results panel. Shows summary, JSON, or filtered data.")
        # Tree panel (created and gridded into the same cell; kept behind until enabled)
        # Use tk.Frame with explicit background because ttk.Treeview doesn't always paint
        # its empty background on Windows themes; without this the empty area can appear white.
        self.tree_frame = tk.Frame(self.content_frame, background='#000')
        self.tree_frame.grid(row=0, column=0, columnspan=2, sticky='nsew')
        self.tree_sub = tk.Frame(self.tree_frame, background='#000')
        self.tree_sub.pack(fill='both', expand=True)
        self.tree = ttk.Treeview(self.tree_sub, show='tree', selectmode='browse')
        self.tree_scroll = ttk.Scrollbar(self.tree_sub, orient='vertical', command=self.tree.yview)
        self.tree.configure(yscrollcommand=self.tree_scroll.set)
        # Keep Tree View output dark even when empty (before any JSON is loaded)
        self._ensure_treeview_style(True)
        self.tree.grid(row=0, column=0, sticky='nsew')
        self.tree_scroll.grid(row=0, column=1, sticky='ns')
        self.tree_sub.rowconfigure(0, weight=1)
        self.tree_sub.columnconfigure(0, weight=1)

        # Right panel for Compare mode (hidden by default)
        self.compare_right_frame = ttk.Frame(self.content_frame)
        self.compare_right_frame.grid(row=0, column=1, sticky='nsew')
        self.compare_text = tk.Text(
            self.compare_right_frame,
            wrap='none',
            font=('Consolas', 10),
            background='#000',
            foreground='#e6e6e6',
            insertbackground='#ffffff',
        )
        cmp_yscroll = ttk.Scrollbar(self.compare_right_frame, orient='vertical', command=self._scroll_y)
        cmp_xscroll = ttk.Scrollbar(self.compare_right_frame, orient='horizontal', command=self._scroll_x)
        self.compare_text.configure(yscrollcommand=cmp_yscroll.set, xscrollcommand=cmp_xscroll.set)
        self.compare_right_frame.rowconfigure(0, weight=1)
        self.compare_right_frame.columnconfigure(0, weight=1)
        self.compare_text.grid(row=0, column=0, sticky='nsew')
        cmp_yscroll.grid(row=0, column=1, sticky='ns')
        cmp_xscroll.grid(row=1, column=0, sticky='ew')
        try:
            self.compare_right_frame.grid_remove()
        except Exception:
            pass
        # Default view
        try:
            self.text_frame.tkraise()
        except Exception:
            pass
        # Highlight tag configuration now that the Text exists
        self.text.tag_configure('search_match', background='#264d73')  # dark blue
        self.text.tag_configure('current_match', background='#ffcc00', foreground='#000000')  # yellow for current match
        # Diff highlight tag (used in Compare mode)
        try:
            self.text.tag_configure('diff_field', foreground='red')
            self.compare_text.tag_configure('diff_field', foreground='red')
        except Exception:
            pass

        # Scroll sync bindings (only actively sync when Compare mode is enabled)
        try:
            self.text.bind('<MouseWheel>', self._on_mousewheel_compare)
            self.compare_text.bind('<MouseWheel>', self._on_mousewheel_compare)
            self.text.bind('<Shift-MouseWheel>', self._on_shift_mousewheel_compare)
            self.compare_text.bind('<Shift-MouseWheel>', self._on_shift_mousewheel_compare)
        except Exception:
            pass
        # Bottom
        bottom = ttk.Frame(self.root); bottom.pack(fill='x', padx=10, pady=(0,8))
        left_bar = ttk.Frame(bottom)
        left_bar.pack(side='left', fill='x', expand=True)
        right_bar = ttk.Frame(bottom)
        right_bar.pack(side='right')

        # File actions
        file_bar = ttk.Frame(left_bar)
        file_bar.pack(side='left')
        self.btn_download = ttk.Button(file_bar, text="Save JSON", command=self.forzar_descarga_json)
        self.btn_download.pack(side='left')
        self._add_tooltip(self.btn_download, "Save the current installation JSON to disk.")
        self.btn_open_json = ttk.Button(file_bar, text="Open JSON", command=self._open_json_folder)
        self.btn_open_json.pack(side='left', padx=(6, 0))
        self._add_tooltip(self.btn_open_json, "Open the temp folder where downloaded JSON files are saved.")

        ttk.Separator(left_bar, orient='vertical').pack(side='left', fill='y', padx=10)

        # Panel action
        panel_bar = ttk.Frame(left_bar)
        panel_bar.pack(side='left')
        btn_copy = ttk.Button(panel_bar, text="Copy", command=self.copiar_texto)
        btn_copy.pack(side='left')
        self._add_tooltip(btn_copy, "Copy the current view to clipboard (Compare copies both panes).")

        ttk.Separator(left_bar, orient='vertical').pack(side='left', fill='y', padx=10)

        # Tools (advanced)
        tools_bar = ttk.Frame(left_bar)
        tools_bar.pack(side='left')
        self.btn_show_paths = ttk.Button(tools_bar, text="Show Paths", command=self.mostrar_rutas)
        self.btn_show_paths.pack(side='left')
        self._add_tooltip(self.btn_show_paths, "Show all scalar paths and values in the loaded JSON.")

        self.btn_extract_path = ttk.Button(tools_bar, text="Extract Path Value", command=self.extraer_valor_ruta_gui)
        self.btn_extract_path.pack(side='left', padx=(6, 0))
        self._add_tooltip(self.btn_extract_path, "Extract the value of the path written in the Search box.")

        self.btn_remote_edit = ttk.Button(tools_bar, text="Remote Edit…", command=self.open_remote_edit_apim)
        self.btn_remote_edit.pack(side='left', padx=(6, 0))
        self._add_tooltip(self.btn_remote_edit, "Open a gated window to pre-check and apply a remote PATCH (APIM).")

        self.btn_remote_edit_bulk = ttk.Button(tools_bar, text="Bulk Remote Edit…", command=self.open_remote_edit_apim_bulk)
        self.btn_remote_edit_bulk.pack(side='left', padx=(6, 0))
        self._add_tooltip(self.btn_remote_edit_bulk, "Apply the same remote PATCH to multiple installations (APIM), with per-installation Check.")

        # Batch-to-CSV option removed

        # Right-side controls
        btn_exit = ttk.Button(right_bar, text="Exit", command=self._on_exit_clicked)
        btn_exit.pack(side='right')
        self._add_tooltip(btn_exit, "Exit the application.")

        self.btn_stop = ttk.Button(right_bar, text="Stop", command=self.stop_batch, state='disabled')
        self.btn_stop.pack(side='right', padx=(0, 6))
        self._add_tooltip(self.btn_stop, "Stop the current Filter CSV batch run.")

        # Compare controls (moved to the bottom)
        compare_bar = ttk.Frame(self.root)
        compare_bar.pack(fill='x', padx=10, pady=(0, 6))
        lbl_cmp = ttk.Label(compare_bar, text="Compare to:")
        lbl_cmp.pack(side='left')
        ent_cmp = ttk.Entry(compare_bar, textvariable=self.compare_install_var, width=14)
        ent_cmp.pack(side='left', padx=(6, 8))
        ent_cmp.bind('<Return>', lambda e: self.compare_installations())
        self._add_tooltip(ent_cmp, "Enter another installation number to compare against the current one.")
        lbl_cmp_country = ttk.Label(compare_bar, text="Country:")
        lbl_cmp_country.pack(side='left')
        self.cmp_country_combo = ttk.Combobox(compare_bar, values=PAISES, width=6, state='readonly')
        self.cmp_country_combo.set(self.compare_country_var.get())
        self.cmp_country_combo.pack(side='left', padx=(6, 8))
        self.cmp_country_combo.bind('<<ComboboxSelected>>', lambda e: self.compare_country_var.set(self.cmp_country_combo.get()))
        self._add_tooltip(self.cmp_country_combo, "Select the country code for the comparison installation.")
        btn_compare = ttk.Button(compare_bar, text="Compare", command=self.compare_installations)
        btn_compare.pack(side='left', padx=(6, 0))
        self._add_tooltip(btn_compare, "Compare the currently loaded installation JSON with another installation.")
        self.btn_compare = btn_compare

        btn_close_compare = ttk.Button(compare_bar, text="Close Compare", command=self.close_compare_view)
        btn_close_compare.pack(side='left', padx=(8, 0))
        self._add_tooltip(btn_close_compare, "Exit Compare mode and return to the normal single-pane view.")
        self.btn_close_compare = btn_close_compare

        chk_only_diffs = ttk.Checkbutton(
            compare_bar,
            text="Only diffs",
            variable=self.compare_only_diffs_var,
            command=self._toggle_compare_only_diffs,
        )
        chk_only_diffs.pack(side='left', padx=(10, 0))
        self._add_tooltip(chk_only_diffs, "When comparing, show only the fields that differ.")
        self.chk_only_diffs = chk_only_diffs
        # Status bar (visible feedback for self.status_var)
        status_frame = ttk.Frame(self.root); status_frame.pack(fill='x', padx=10, pady=(0,6))
        self.status_label = ttk.Label(status_frame, textvariable=self.status_var, anchor='w')
        self.status_label.pack(side='left', fill='x', expand=True)
        self.progress = ttk.Progressbar(status_frame, mode='indeterminate', length=160)
        # Hide the progress bar initially
        try:
            self.progress.pack_forget()
        except Exception:
            pass
        # Sync initial state for the Download button visibility/label
        self._update_download_button()
        self._update_action_states()

    def _on_exit_clicked(self):
        """Exit button handler: if a batch is running, confirm stopping."""
        try:
            if getattr(self, '_batch_running', False):
                resp = messagebox.askyesno(
                    "Exit",
                    "A batch run is still running. Stop it and exit now?",
                )
                if not resp:
                    return
                try:
                    self.stop_batch()
                except Exception:
                    pass
        except Exception:
            pass
        try:
            self.root.destroy()
        except Exception:
            pass

    def _open_paths_tools(self):
        """Open a small popup with advanced JSON path tools."""
        if self._last_json_data is None:
            messagebox.showinfo("Paths", "No JSON loaded. Query an installation first.")
            return

        try:
            if getattr(self, '_paths_win', None) is not None and self._paths_win.winfo_exists():
                try:
                    # Rebuild the window to pick up new buttons/options without requiring a full app restart.
                    self._paths_win.destroy()
                except Exception:
                    pass
        except Exception:
            pass

        win = tk.Toplevel(self.root)
        self._paths_win = win
        win.title("Paths")
        try:
            win.transient(self.root)
        except Exception:
            pass
        try:
            win.resizable(False, False)
        except Exception:
            pass

        frm = ttk.Frame(win, padding=10)
        frm.pack(fill='both', expand=True)
        ttk.Label(frm, text="JSON path tools", font=('Segoe UI', 10, 'bold')).pack(anchor='w')
        ttk.Label(frm, text="Tip: use the Search box as input for Extract Path Value.").pack(anchor='w', pady=(2, 10))

        btn1 = ttk.Button(frm, text="Show Paths", command=lambda: (win.destroy(), self.mostrar_rutas()))
        btn1.pack(fill='x')
        self._add_tooltip(btn1, "Show all scalar paths and values in the loaded JSON.")

        btn2 = ttk.Button(frm, text="Extract Path Value", command=lambda: (win.destroy(), self.extraer_valor_ruta_gui()))
        btn2.pack(fill='x', pady=(6, 0))
        self._add_tooltip(btn2, "Extract the value of the path written in the Search box.")

        btn3 = ttk.Button(frm, text="Remote Edit (APIM)…", command=lambda: (win.destroy(), self.open_remote_edit_apim()))
        btn3.pack(fill='x', pady=(6, 0))
        self._add_tooltip(btn3, "Open a separate, gated window to pre-check and apply remote JSON Patch operations via APIM.")

        btn4 = ttk.Button(frm, text="Bulk Remote Edit (APIM)…", command=lambda: (win.destroy(), self.open_remote_edit_apim_bulk()))
        btn4.pack(fill='x', pady=(6, 0))
        self._add_tooltip(btn4, "Apply the same remote PATCH to multiple installations (with per-installation Check before Send).")

        ttk.Button(frm, text="Close", command=win.destroy).pack(fill='x', pady=(10, 0))

    def open_remote_edit_apim_bulk(self):
        """Open a separate, gated bulk remote edit window (APIM PATCH config/changes) for multiple installations."""
        if self._last_json_data is None:
            messagebox.showinfo(
                "Bulk Remote Edit",
                "No JSON loaded. You can still use Bulk Remote Edit; per-installation GET will be used during Check.",
                parent=self.root,
            )
        if requests is None:
            messagebox.showerror("Bulk Remote Edit", f"requests is not available: {_requests_import_error}")
            return

        try:
            if getattr(self, '_remote_edit_bulk_win', None) is not None and self._remote_edit_bulk_win.winfo_exists():
                try:
                    self._remote_edit_bulk_win.lift()
                    return
                except Exception:
                    pass
        except Exception:
            pass

        win = tk.Toplevel(self.root)
        self._remote_edit_bulk_win = win
        win.title("Bulk Remote Edit (APIM)")
        try:
            win.transient(self.root)
        except Exception:
            pass

        frm = ttk.Frame(win, padding=10)
        frm.pack(fill='both', expand=True)

        ttk.Label(frm, text="Bulk Remote Edit (APIM)", font=('Segoe UI', 11, 'bold')).grid(row=0, column=0, columnspan=4, sticky='w')
        ttk.Label(frm, text="Gated window. Use Check all before Send PATCH.").grid(row=1, column=0, columnspan=4, sticky='w', pady=(2, 10))

        # Gate
        enable_var = tk.BooleanVar(value=False)
        phrase_var = tk.StringVar(value="")
        ttk.Checkbutton(frm, text="Enable editing (danger)", variable=enable_var).grid(row=2, column=0, columnspan=2, sticky='w')
        ttk.Label(frm, text="Type EDIT to unlock:").grid(row=2, column=2, sticky='e')
        ttk.Entry(frm, textvariable=phrase_var, width=10).grid(row=2, column=3, sticky='w')

        # Defaults (.env)
        dotenv = {}
        try:
            dotenv.update(self._read_dotenv_file(os.path.join(os.path.dirname(__file__), '.env')))
        except Exception:
            pass
        try:
            dotenv.update(self._read_dotenv_file(os.path.join(os.getcwd(), '.env')))
        except Exception:
            pass

        default_base = (
            os.getenv("DCR_APIM_URL", "").strip()
            or os.getenv("DCR_APIM_BASE_URL", "").strip()
            or str(dotenv.get("DCR_APIM_URL", "")).strip()
            or str(dotenv.get("DCR_APIM_BASE_URL", "")).strip()
            or os.getenv("DCR_BASE_URL", DEFAULT_DCR_BASE_URL).strip()
        )

        VALUE_MODE_FIXED = "Auto (infer)"
        PERSISTENT_FIXED = True
        REQ_BY_USER_FIXED = "Rnd"
        req_by_app_default = (
            os.getenv("DCR_REQUESTED_BY_APP", "").strip()
            or str(dotenv.get("DCR_REQUESTED_BY_APP", "")).strip()
            or "Rnd"
        ).strip()

        def _resolve_base_now() -> str:
            return (default_base or "").strip()

        def _is_unlocked() -> bool:
            return bool(enable_var.get()) and phrase_var.get().strip().upper() == 'EDIT'

        country_var = tk.StringVar(value=(self.country_var.get() or "").strip().upper())
        serial_var = tk.StringVar(value="")
        path_var = tk.StringVar(value="/config/entryType")
        new_value_var = tk.StringVar(value="1")
        allow_add_var = tk.BooleanVar(value=False)
        show_cur_var = tk.BooleanVar(value=True)
        verify_after_send_var = tk.BooleanVar(value=True)
        verify_timeout_s_var = tk.StringVar(value="8")

        r = 3
        ttk.Separator(frm, orient='horizontal').grid(row=r, column=0, columnspan=4, sticky='we', pady=(10, 10)); r += 1

        ttk.Label(frm, text="Country (default):").grid(row=r, column=0, sticky='w')
        ttk.Entry(frm, textvariable=country_var, width=6).grid(row=r, column=1, sticky='w', pady=(0, 6))
        ttk.Label(frm, text="Node serialNumber (optional):").grid(row=r, column=2, sticky='e')
        ttk.Entry(frm, textvariable=serial_var, width=24).grid(row=r, column=3, sticky='w', pady=(0, 6)); r += 1

        ttk.Label(frm, text="Installations (one per line):").grid(row=r, column=0, sticky='nw')
        inst_txt = tk.Text(frm, width=26, height=7, wrap='none', font=('Consolas', 9))
        inst_txt.grid(row=r, column=1, sticky='w', pady=(0, 6))
        ttk.Label(frm, text="Examples:\n5499266\nES5499266\nES 5499266", foreground="#666666").grid(row=r, column=2, columnspan=2, sticky='w', pady=(0, 6))
        r += 1

        def importar_csv():
            file_path = filedialog.askopenfilename(title="Select CSV file", filetypes=[("CSV files", "*.csv"), ("All files", "*")], parent=win)
            if not file_path:
                return
            try:
                ext = os.path.splitext(file_path)[1].lower()
                rows = []

                # If Excel file, try using openpyxl
                if ext in ('.xlsx', '.xlsm', '.xltx', '.xltm'):
                    try:
                        from openpyxl import load_workbook
                    except Exception:
                        messagebox.showerror("Import CSV", "Para importar archivos Excel instala 'openpyxl' (pip install openpyxl).", parent=win)
                        return

                    wb = load_workbook(filename=file_path, read_only=True, data_only=True)
                    ws = wb.active
                    it = ws.iter_rows(values_only=True)
                    try:
                        headers = [ (str(h).strip().lower() if h is not None else '') for h in next(it) ]
                    except StopIteration:
                        messagebox.showwarning("Import CSV", "El archivo Excel está vacío.", parent=win)
                        return

                    for row in it:
                        if not row:
                            continue
                        vals = [ (str(v).strip() if v is not None else '') for v in row ]
                        d = { headers[i]: vals[i] for i in range(min(len(headers), len(vals))) }
                        country = d.get('ds_country_short') or d.get('country') or d.get('pais')
                        inst = d.get('id_install') or d.get('install') or d.get('installation') or d.get('id')
                        if country and inst:
                            rows.append((str(country).strip().upper(), str(inst).strip()))

                else:
                    # Treat as CSV: try common encodings until one works
                    encodings_to_try = ["utf-8", "utf-8-sig", "cp1252", "latin-1"]
                    chosen_encoding = None
                    dialect = csv.excel
                    sample = None
                    for enc in encodings_to_try:
                        try:
                            with open(file_path, "r", newline='', encoding=enc) as csvfile:
                                sample = csvfile.read(4096)
                                csvfile.seek(0)
                                try:
                                    dialect = csv.Sniffer().sniff(sample)
                                except Exception:
                                    dialect = csv.excel
                            chosen_encoding = enc
                            break
                        except UnicodeDecodeError:
                            continue

                    if not chosen_encoding:
                        messagebox.showerror("Import CSV", "No se pudo detectar una codificación válida para el archivo CSV.", parent=win)
                        return

                    with open(file_path, "r", newline='', encoding=chosen_encoding) as csvfile:
                        csvfile.seek(0)
                        dr = csv.DictReader(csvfile, dialect=dialect)
                        for row in dr:
                            if not row:
                                continue
                            # row values may be lists if multiple columns with same name -> normalize
                            d = {}
                            for k, v in row.items():
                                key = (k or '').strip().lower()
                                # if v is a list, join; otherwise cast to str
                                if isinstance(v, list):
                                    val = ','.join([str(x) for x in v if x is not None]).strip()
                                else:
                                    val = (v or '').strip()
                                d[key] = val
                            country = d.get('ds_country_short') or d.get('country') or d.get('pais')
                            inst = d.get('id_install') or d.get('install') or d.get('installation') or d.get('id')
                            if country and inst:
                                rows.append((country.strip().upper(), str(inst).strip()))

                        # If no rows found via DictReader, fallback to simple reader
                        if not rows:
                            csvfile.seek(0)
                            rr = csv.reader(csvfile, dialect=dialect)
                            for rrow in rr:
                                if not rrow:
                                    continue
                                if len(rrow) >= 2:
                                    c = (rrow[0] or '').strip().upper()
                                    i = (rrow[1] or '').strip()
                                    if c and i:
                                        rows.append((c, i))
                                else:
                                    token = (rrow[0] or '').strip()
                                    m = re.fullmatch(r"([A-Za-z]{2})[-_ ]?([0-9]+)", token)
                                    if m:
                                        rows.append((m.group(1).upper(), m.group(2)))
                                    elif token.isdigit():
                                        rows.append(((country_var.get() or '').strip().upper(), token))

                if not rows:
                    messagebox.showwarning("Import CSV", "No valid installations found in the file.", parent=win)
                    return

                # Insert into the installations textbox (one per line)
                try:
                    inst_txt.delete('1.0', 'end')
                except Exception:
                    pass
                for ctry, inst in rows:
                    inst_txt.insert('end', f"{ctry}{inst}\n")

                try:
                    _invalidate_check('import csv')
                except Exception:
                    pass

                messagebox.showinfo("Import CSV", f"Se han importado {len(rows)} instalaciones.", parent=win)
            except Exception as e:
                messagebox.showerror("Import CSV", f"Error al importar CSV: {e}", parent=win)

        btn_import_csv = ttk.Button(frm, text="Importar CSV", command=importar_csv)
        btn_import_csv.grid(row=r, column=1, sticky='w', pady=(0, 6))
        r += 1

        ttk.Label(frm, text="Path (pointer or simple dot):").grid(row=r, column=0, sticky='w')
        ttk.Entry(frm, textvariable=path_var, width=54).grid(row=r, column=1, columnspan=3, sticky='we', pady=(0, 6)); r += 1

        ttk.Label(frm, text="New value:").grid(row=r, column=0, sticky='w')
        ttk.Entry(frm, textvariable=new_value_var, width=24).grid(row=r, column=1, sticky='w', pady=(0, 6))
        # 'Allow add' control removed from UI (kept as variable for internal logic)

        def _quick_set_user_json():
            user_json = '[{"_id": "coredumpsUpload", "enabled": true, "params": [{"_id": "endpoint", "value": ""}]}, {"_id": "logVerbose", "enabled": true, "params": [{"_id": "LOG_VERBOSE_MQTT", "value": "false"}, {"_id": "LOG_VERBOSE_EVENT_LOOP", "value": "false"}]}, {"_id": "coreEventLogging", "enabled": true}]'
            try:
                new_value_var.set(user_json)
                _invalidate_check("quick set user json")
                messagebox.showinfo("New value set", "New value updated with provided JSON.", parent=win)
            except Exception:
                try:
                    messagebox.showerror("Error", "Failed to set New value.", parent=win)
                except Exception:
                    pass

        btn_quick_set = ttk.Button(frm, text="Set example JSON", command=_quick_set_user_json)
        btn_quick_set.grid(row=r, column=3, sticky='w', padx=(6,0), pady=(0,6))
        r += 1

        # 'Show current value' is enabled by default but hidden in UI
        r += 1

        # 'Verify after SEND' is enabled by default but hidden in UI
        r += 1

        ttk.Label(frm, text="Verify timeout (sec):").grid(row=r, column=0, sticky='w')
        ttk.Entry(frm, textvariable=verify_timeout_s_var, width=8).grid(row=r, column=1, sticky='w', pady=(0, 6))
        ttk.Label(frm, text="(increase if backend applies changes async)", foreground="#666666").grid(row=r, column=2, columnspan=2, sticky='w', pady=(0, 6)); r += 1

        ttk.Label(frm, text="Preview / Progress:").grid(row=r, column=0, sticky='nw', pady=(6, 0))
        out_txt = tk.Text(frm, width=92, height=16, wrap='word', font=('Consolas', 9))
        out_txt.grid(row=r, column=1, columnspan=3, sticky='we', pady=(6, 0)); r += 1

        log_q: "queue.Queue[tuple[str, str]]" = queue.Queue()
        running = {"active": False, "mode": ""}
        stop_flag = {"stop": False}
        last_checked = {"ok": False, "items": [], "stale_reason": ""}

        def _write(s: str):
            try:
                out_txt.delete('1.0', 'end')
                out_txt.insert('1.0', s)
                out_txt.see('1.0')
            except Exception:
                pass

        def _append(s: str):
            try:
                out_txt.insert('end', s + ("\n" if not s.endswith("\n") else ""))
                out_txt.see('end')
            except Exception:
                pass

        def _pump_queue():
            try:
                while True:
                    kind, msg = log_q.get_nowait()
                    if kind == 'clear':
                        _write(msg)
                    else:
                        _append(msg)
            except Exception:
                pass
            try:
                win.after(120, _pump_queue)
            except Exception:
                pass

        win.after(120, _pump_queue)

        def _invalidate_check(reason: str = "inputs changed") -> None:
            try:
                if last_checked.get("ok") or last_checked.get("items"):
                    last_checked["ok"] = False
                    last_checked["items"] = []
                    last_checked["stale_reason"] = str(reason or "inputs changed")
                else:
                    if not last_checked.get("stale_reason"):
                        last_checked["stale_reason"] = str(reason or "inputs changed")
            except Exception:
                pass
            try:
                _refresh_buttons()
            except Exception:
                pass

        auth_var = tk.StringVar(value=self._resolve_initial_auth(dotenv))

        def _resolve_auth_now() -> str:
            cached = self._normalize_auth(auth_var.get())
            if cached:
                return cached
            return self._resolve_initial_auth(dotenv)

        def _parent_pointer(ptr: str) -> str:
            ptr = str(ptr or '').strip()
            if ptr in ('', '/'):
                return '/'
            if not ptr.startswith('/'):
                return '/'
            parts = ptr.split('/')
            if len(parts) <= 2:
                return '/'
            out = '/'.join(parts[:-1])
            return out if out else '/'

        def _last_pointer_token(ptr: str) -> str:
            ptr = str(ptr or '').strip()
            if not ptr or ptr == '/':
                return ''
            parts = ptr.split('/')
            raw = parts[-1] if parts else ''
            try:
                return self._unescape_json_pointer_token(raw)
            except Exception:
                return raw

        def _validate_patch_target(data: Any, full_ptr: str, exists: bool, old_val: Any, why: str) -> tuple[str, str]:
            if exists:
                return 'replace', 'OK: existing path (replace)'
            if not bool(allow_add_var.get()):
                raise ValueError(
                    "Path does not exist. For safety, bulk edit blocks creating new keys by default. "
                    "Enable 'Allow add (create new key)' if you really want to add it."
                )

            parent_ptr = _parent_pointer(full_ptr)
            p_exists, p_val, p_why = self._get_by_pointer(data, parent_ptr)
            if not p_exists:
                raise ValueError(f"Invalid path: parent is missing at {parent_ptr} ({p_why})")
            if not isinstance(p_val, (dict, list)):
                raise ValueError(f"Invalid path: parent {parent_ptr} is not an object/array (type={self._json_type_name(p_val)})")
            if isinstance(p_val, list):
                tok = _last_pointer_token(full_ptr)
                if tok == '-':
                    return 'add', 'OK: append to array (add /-)'
                try:
                    idx = int(tok)
                except Exception:
                    raise ValueError("Invalid path: array target must be a numeric index or '-' for append")
                if idx < 0 or idx > len(p_val):
                    raise ValueError(f"Invalid path: array index out of range for add (idx={idx}, len={len(p_val)})")
                return 'add', 'OK: array index add'

            return 'add', 'OK: missing leaf under existing object (add)'

        def _parse_new_value(raw: str, exists: bool, old_val: Any) -> tuple[Any, str, str]:
            raw = str(raw or '').strip()
            if raw == "":
                raise ValueError("New value is empty")

            def _try_json_load(s: str):
                try:
                    return True, json.loads(s)
                except Exception:
                    return False, None

            if exists:
                # Comprobación extra: mostrar advertencia clara si el tipo no coincide
                if isinstance(old_val, bool):
                    rl = raw.strip().lower()
                    if rl in ('true', 'false'):
                        return (rl == 'true'), 'auto:boolean', ''
                    if raw in ('True', 'False'):
                        return (raw == 'True'), 'auto:boolean', ''
                    raise ValueError("Expected a boolean for this path. Use true/false.")
                if isinstance(old_val, (int, float)) and not isinstance(old_val, bool):
                    ok, val = _try_json_load(raw)
                    if ok and isinstance(val, (int, float)) and not isinstance(val, bool):
                        return val, 'auto:number(json)', ''
                    try:
                        if any(c in raw for c in ('.', 'e', 'E')):
                            return float(raw), 'auto:number', ''
                        return int(raw), 'auto:number', ''
                    except Exception:
                        raise ValueError("Expected a number for this path (e.g. 10 or 3.14).")
                if isinstance(old_val, str):
                    if len(raw) >= 2 and raw[0] == raw[-1] and raw[0] in ('\"', "'"):
                        ok, val = _try_json_load(raw)
                        if ok and isinstance(val, str):
                            return val, 'auto:string(json)', ''
                    return raw, 'auto:string', ''
                if old_val is None:
                    rl = raw.strip().lower()
                    if rl in ('null', 'none'):
                        return None, 'auto:null', ''
                    raise ValueError("Expected null for this path. Use null.")
                if isinstance(old_val, (dict, list)):
                    ok, val = _try_json_load(raw)
                    if not ok:
                        raise ValueError("Expected a JSON object/array for this path. Paste valid JSON.")
                    # Nueva comprobación: si el tipo no coincide, mostrar sugerencia clara
                    if isinstance(old_val, dict) and isinstance(val, dict):
                        return val, 'auto:object(json)', ''
                    if isinstance(old_val, list) and isinstance(val, list):
                        return val, 'auto:array(json)', ''
                    # Si el tipo no coincide, mostrar advertencia y sugerir el tipo correcto
                    raise ValueError(
                        f"Type mismatch: current value is {self._json_type_name(old_val)} but new value is {self._json_type_name(val)}. "
                        f"Tip: Si el valor actual es un array, el nuevo valor debe ser una lista (ejemplo: [{{...}}]), no un objeto. "
                        f"Si quieres añadir un objeto a un array, usa el path adecuado (por ejemplo, /array/- para añadir al final)."
                    )

            ok, val = _try_json_load(raw)
            if ok:
                return val, 'json', ''
            if raw in ("True", "False", "None"):
                raise ValueError("Value is not valid JSON. Use true/false/null (lowercase) or quote it as a string.")
            return raw, 'string', ''

        def _parse_installation_lines(text: str, default_country: str) -> list[tuple[str, str]]:
            dc = (default_country or '').strip().upper()
            out: list[tuple[str, str]] = []
            seen = set()
            for line in (text or '').splitlines():
                line = (line or '').strip()
                if not line:
                    # Si la línea está vacía, la ignoramos y seguimos con la siguiente
                    continue
                # allow commas/spaces
                parts = [p for p in re.split(r"[\s,;]+", line) if p]
                ctry = inst = None
                if len(parts) == 1:
                    token = parts[0].strip()
                    token = token.replace('"', '').replace("'", '').strip()
                    if len(token) >= 3 and token[:2].isalpha() and token[2:].isdigit():
                        ctry, inst = token[:2].upper(), token[2:]
                    elif token.isdigit():
                        ctry, inst = dc, token
                    else:
                        # try ES 5499266 style stuck together with dash
                        m = re.fullmatch(r"([A-Za-z]{2})[-_]?([0-9]+)", token)
                        if m:
                            ctry, inst = m.group(1).upper(), m.group(2)
                        # Si no es válida, simplemente seguimos (no break)
                else:
                    # ES 5499266
                    if len(parts[0]) == 2 and parts[0].isalpha() and parts[1].isdigit():
                        ctry, inst = parts[0].upper(), parts[1]
                    elif parts[0].isdigit():
                        ctry, inst = dc, parts[0]
                    # Si no es válida, simplemente seguimos (no break)
                if not ctry or len(ctry) != 2 or not ctry.isalpha() or not inst or not inst.isdigit():
                    continue
                key = (ctry, inst)
                if key in seen:
                    continue
                seen.add(key)
                out.append(key)
            return out

        def _refresh_buttons(*_):
            unlocked = _is_unlocked()
            has_token = bool(_resolve_auth_now())
            has_ok = bool(last_checked.get('ok')) and any(it.get('ok') for it in (last_checked.get('items') or []))
            busy = bool(running.get('active'))
            try:
                btn_check.configure(state=('disabled' if busy else 'normal'))
            except Exception:
                pass
            try:
                btn_send.configure(state=('normal' if (not busy and unlocked and has_token and has_ok) else 'disabled'))
            except Exception:
                pass
            try:
                btn_stop.configure(state=('normal' if busy else 'disabled'))
            except Exception:
                pass

        def _check_all():
            if running.get('active'):
                return
            stop_flag['stop'] = False
            running['active'] = True
            running['mode'] = 'check'
            last_checked['ok'] = False
            last_checked['items'] = []
            last_checked['stale_reason'] = ''
            _refresh_buttons()

            def worker():
                try:
                    log_q.put(('clear', 'CHECK ALL: starting...'))

                    base = _resolve_base_now()
                    if not base:
                        raise ValueError(
                            "APIM base URL is not configured. Set DCR_APIM_URL (or DCR_APIM_BASE_URL) in Ultrajson/.env or as an environment variable."
                        )
                    auth = _resolve_auth_now()
                    if not auth:
                        raise ValueError(
                            "No APIM token found.\n\n"
                            "Set it in Ultrajson/.env as DCR_APIM_BEARER_TOKEN=... (or DCR_APIM_AUTHORIZATION=Bearer ...)\n"
                            "or copy the token (or Authorization: Bearer ...) to clipboard and try again."
                        )
                    self._remote_edit_last_auth = auth  # cachear para reusar en single edit

                    ctry_default = (country_var.get() or '').strip().upper()
                    ids = _parse_installation_lines(inst_txt.get('1.0', 'end'), ctry_default)
                    if not ids:
                        raise ValueError('No valid installations found in the list.')

                    ok_ptr, ptr, ptr_reason = self._dotpath_to_pointer_simple(path_var.get().strip())
                    if not ok_ptr:
                        raise ValueError(ptr_reason)

                    serial = serial_var.get().strip()
                    items = []
                    log_q.put(('append', f"CHECK ALL: {len(ids)} installations"))
                    log_q.put(('append', f"Path: {ptr}  | Mode: {VALUE_MODE_FIXED} (fixed) | persistent: {str(bool(PERSISTENT_FIXED)).lower()} (fixed)"))
                    if serial:
                        log_q.put(('append', f"Node serialNumber: {self._normalize_serial(serial)}"))
                    log_q.put(('append', ''))

                    for ctry, ins in ids:
                        if stop_flag.get('stop'):
                            log_q.put(('append', 'CHECK ALL: stopped.'))
                            break
                        installation_id = f"{ctry}{ins}"
                        try:
                            get_url = f"{base.rstrip('/')}/device-support/device-config-repository/v2.0/installation/{installation_id}"
                            resp = requests.get(get_url, headers={"Authorization": auth}, timeout=20)
                            if resp.status_code >= 400:
                                raise ValueError(f"GET failed HTTP {resp.status_code}: {resp.text[:200]}")
                            data = resp.json()

                            full_ptr = ptr
                            target_desc = 'Target: root JSON'
                            if ptr.startswith('/nodes/'):
                                target_desc = 'Target: absolute pointer'
                            elif serial:
                                idx, node_reason = self._find_node_index_by_serial(data, serial)
                                if idx is None:
                                    raise ValueError(node_reason)
                                full_ptr = self._pointer_join(f"/nodes/{idx}", ptr)
                                target_desc = f"Target: node serialNumber={self._normalize_serial(serial)} -> nodes[{idx}]"

                            exists, old_val, why = self._get_by_pointer(data, full_ptr)
                            op, path_note = _validate_patch_target(data, full_ptr, exists, old_val, why)
                            new_val, parse_mode, _parse_note = _parse_new_value(new_value_var.get(), exists, old_val)
                            payload = {
                                "persistent": bool(PERSISTENT_FIXED),
                                "request": [{"op": op, "path": full_ptr, "value": new_val}],
                            }
                            patch_url = f"{base.rstrip('/')}/device-support/device-config-repository/v2.0/config/changes"
                            headers = {
                                "Authorization": auth,
                                "Content-Type": "application/json",
                                "x-sd-requested-by-user": REQ_BY_USER_FIXED,
                            }
                            if req_by_app_default:
                                headers["x-sd-requested-by-app"] = req_by_app_default

                            items.append(
                                {
                                    "ok": True,
                                    "installation_id": installation_id,
                                    "get_url": get_url,
                                    "url": patch_url,
                                    "headers": headers,
                                    "params": {"installationId": installation_id},
                                    "payload": payload,
                                    "full_ptr": full_ptr,
                                    "expected_new_val": new_val,
                                }
                            )

                            cur_type = self._json_type_name(old_val) if exists else 'missing'
                            log_q.put(('append', f"OK  {installation_id} | {target_desc} | {full_ptr} | {path_note} | cur={cur_type} -> new={self._json_type_name(new_val)} ({parse_mode})"))
                            if bool(show_cur_var.get()):
                                if exists:
                                    log_q.put(('append', f"    current={self._format_value_preview(old_val)}"))
                                else:
                                    log_q.put(('append', "    current=<missing>"))
                        except Exception as e:
                            items.append({"ok": False, "installation_id": installation_id, "error": str(e)})
                            log_q.put(('append', f"FAIL {installation_id} | {e}"))

                    ok_count = sum(1 for it in items if it.get('ok'))
                    fail_count = len(items) - ok_count
                    last_checked['items'] = items
                    last_checked['ok'] = ok_count > 0 and not stop_flag.get('stop')
                    log_q.put(('append', ''))
                    log_q.put(('append', f"CHECK ALL: done. OK={ok_count} FAIL={fail_count}"))
                    if fail_count:
                        log_q.put(('append', "Tip: fix failing installations and run Check all again."))
                except Exception as e:
                    last_checked['ok'] = False
                    last_checked['items'] = []
                    log_q.put(('clear', "CHECK ALL: FAIL\n" + str(e) + "\n\n" + traceback.format_exc()))
                finally:
                    running['active'] = False
                    running['mode'] = ''
                    try:
                        win.after(0, _refresh_buttons)
                    except Exception:
                        pass

            threading.Thread(target=worker, daemon=True).start()

        def _send_all_ok():
            if running.get('active'):
                return
            if not _is_unlocked():
                messagebox.showwarning("Bulk Remote Edit", "Editing is locked. Check 'Enable editing' and type EDIT.", parent=win)
                return
            items = [it for it in (last_checked.get('items') or []) if it.get('ok')]
            if not items:
                messagebox.showinfo("Bulk Remote Edit", "Run Check all first (need at least one OK).", parent=win)
                return

            if not messagebox.askyesno(
                "Confirm Bulk PATCH",
                f"Send PATCH to {len(items)} installations (only those with CHECK OK)?\n\nProceed?",
                parent=win,
            ):
                return

            stop_flag['stop'] = False
            running['active'] = True
            running['mode'] = 'send'
            _refresh_buttons()

            def worker():
                try:
                    log_q.put(('append', ''))
                    log_q.put(('append', f"SEND: starting... ({len(items)} installations)"))
                    ok = 0
                    fail = 0

                    try:
                        verify_timeout_s = int(str(verify_timeout_s_var.get() or '').strip() or '0')
                    except Exception:
                        verify_timeout_s = 8
                    if verify_timeout_s < 0:
                        verify_timeout_s = 0
                    if verify_timeout_s > 120:
                        verify_timeout_s = 120

                    for it in items:
                        if stop_flag.get('stop'):
                            log_q.put(('append', 'SEND: stopped.'))
                            break
                        installation_id = it.get('installation_id')
                        try:
                            resp = requests.patch(
                                it['url'],
                                params=it.get('params') or {},
                                headers=it.get('headers') or {},
                                json=it.get('payload') or {},
                                timeout=20,
                            )
                            if resp.status_code >= 400:
                                fail += 1
                                log_q.put(('append', f"FAIL {installation_id} | HTTP {resp.status_code} | {resp.text[:200]}"))
                            else:
                                ok += 1
                                # Provide response preview even for HTTP 2xx (helps diagnose async/queued changes)
                                resp_preview = ''
                                try:
                                    resp_preview = json.dumps(resp.json(), ensure_ascii=False)
                                except Exception:
                                    resp_preview = (resp.text or '')
                                resp_preview = (resp_preview or '').replace('\r', ' ').replace('\n', ' ').strip()
                                if len(resp_preview) > 180:
                                    resp_preview = resp_preview[:179] + '…'

                                if bool(verify_after_send_var.get()):
                                    try:
                                        get_url = str(it.get('get_url') or '').strip()
                                        full_ptr = str(it.get('full_ptr') or '').strip()
                                        expected_val = it.get('expected_new_val')
                                        auth = ((it.get('headers') or {}).get('Authorization') or '').strip()
                                        last_err = ''
                                        verified = False
                                        cur_preview = ''
                                        deadline = time.time() + float(verify_timeout_s or 0)
                                        attempt = 0
                                        while True:
                                            if stop_flag.get('stop'):
                                                break
                                            attempt += 1
                                            if not get_url or not auth or not full_ptr:
                                                last_err = 'verify not configured'
                                                break
                                            vresp = requests.get(get_url, headers={"Authorization": auth}, timeout=20)
                                            if vresp.status_code >= 400:
                                                last_err = f"GET HTTP {vresp.status_code}: {vresp.text[:120]}"
                                            else:
                                                vdata = vresp.json()
                                                v_exists, v_val, _v_why = self._get_by_pointer(vdata, full_ptr)
                                                cur_preview = self._format_value_preview(v_val) if v_exists else '<missing>'
                                                if v_exists and v_val == expected_val:
                                                    verified = True
                                                    break
                                                last_err = f"mismatch (current={cur_preview})"

                                            if (verify_timeout_s or 0) <= 0:
                                                break
                                            if time.time() >= deadline:
                                                break
                                            time.sleep(1.0)

                                        if verified:
                                            if resp_preview:
                                                log_q.put(('append', f"OK  {installation_id} | HTTP {resp.status_code} | verify=OK | current={cur_preview} | resp={resp_preview}"))
                                            else:
                                                log_q.put(('append', f"OK  {installation_id} | HTTP {resp.status_code} | verify=OK | current={cur_preview}"))
                                        else:
                                            # Verification failed, but accept the successful PATCH response as OK
                                            try:
                                                exp_json = json.dumps(expected_val, ensure_ascii=False)
                                            except Exception:
                                                try:
                                                    exp_json = str(expected_val)
                                                except Exception:
                                                    exp_json = '<unserializable expected value>'
                                            if len(exp_json) > 400:
                                                exp_preview = exp_json[:400] + '…'
                                            else:
                                                exp_preview = exp_json
                                            if resp_preview:
                                                log_q.put(('append', f"OK  {installation_id} | HTTP {resp.status_code} | verify=ACCEPTED (mismatch) | expected={exp_preview} | {last_err} | resp={resp_preview}"))
                                            else:
                                                log_q.put(('append', f"OK  {installation_id} | HTTP {resp.status_code} | verify=ACCEPTED (mismatch) | expected={exp_preview} | {last_err}"))
                                    except Exception as ve:
                                        if resp_preview:
                                            log_q.put(('append', f"WARN {installation_id} | HTTP {resp.status_code} | verify=ERROR | {ve} | resp={resp_preview}"))
                                        else:
                                            log_q.put(('append', f"WARN {installation_id} | HTTP {resp.status_code} | verify=ERROR | {ve}"))
                                else:
                                    if resp_preview:
                                        log_q.put(('append', f"OK  {installation_id} | HTTP {resp.status_code} | resp={resp_preview}"))
                                    else:
                                        log_q.put(('append', f"OK  {installation_id} | HTTP {resp.status_code}"))
                        except Exception as e:
                            fail += 1
                            log_q.put(('append', f"FAIL {installation_id} | {e}"))

                    log_q.put(('append', ''))
                    log_q.put(('append', f"SEND: done. OK={ok} FAIL={fail}"))
                except Exception as e:
                    log_q.put(('append', "SEND: FAIL\n" + str(e) + "\n\n" + traceback.format_exc()))
                finally:
                    running['active'] = False
                    running['mode'] = ''
                    try:
                        win.after(0, _refresh_buttons)
                    except Exception:
                        pass

            threading.Thread(target=worker, daemon=True).start()

        def _stop():
            stop_flag['stop'] = True

        # Inputs invalidate check
        try:
            country_var.trace_add('write', lambda *_a: _invalidate_check('country changed'))
        except Exception:
            pass
        try:
            serial_var.trace_add('write', lambda *_a: _invalidate_check('serial changed'))
        except Exception:
            pass
        try:
            path_var.trace_add('write', lambda *_a: _invalidate_check('path changed'))
        except Exception:
            pass
        try:
            new_value_var.trace_add('write', lambda *_a: _invalidate_check('new value changed'))
        except Exception:
            pass
        try:
            allow_add_var.trace_add('write', lambda *_a: _invalidate_check('allow-add changed'))
        except Exception:
            pass
        try:
            show_cur_var.trace_add('write', lambda *_a: _invalidate_check('show-current changed'))
        except Exception:
            pass
        try:
            verify_after_send_var.trace_add('write', lambda *_a: _invalidate_check('verify-after-send changed'))
        except Exception:
            pass
        try:
            verify_timeout_s_var.trace_add('write', lambda *_a: _invalidate_check('verify-timeout changed'))
        except Exception:
            pass
        try:
            inst_txt.bind('<KeyRelease>', lambda _e: _invalidate_check('installations list changed'))
        except Exception:
            pass

        # Buttons
        btns = ttk.Frame(frm)
        btns.grid(row=r, column=0, columnspan=4, sticky='we', pady=(10, 0))
        btn_check = ttk.Button(btns, text="Check all", command=_check_all)
        btn_check.pack(side='left')
        btn_send = ttk.Button(btns, text="Send PATCH (OK only)", command=_send_all_ok)
        btn_send.pack(side='left', padx=(8, 0))
        btn_stop = ttk.Button(btns, text="Stop", command=_stop)
        btn_stop.pack(side='left', padx=(8, 0))
        ttk.Button(btns, text="Close", command=win.destroy).pack(side='right')

        enable_var.trace_add('write', _refresh_buttons)
        phrase_var.trace_add('write', _refresh_buttons)

        # Initial state
        _write(
            "Paste installations (one per line), set Path and New value, then run Check all.\n"
            "Send is enabled only when unlocked + token available + at least one CHECK OK.\n"
            "\nHeaders used:\n"
            f"- x-sd-requested-by-app: {req_by_app_default or '(<none)'}\n"
            f"- x-sd-requested-by-user: {REQ_BY_USER_FIXED}\n"
        )
        _refresh_buttons()

        try:
            frm.columnconfigure(1, weight=1)
        except Exception:
            pass


    def _normalize_serial(self, s: str) -> str:
        return str(s or "").strip().upper()

    @staticmethod
    def _read_dotenv_file(path: str) -> dict:
        """Lee un fichero .env y devuelve un dict con sus claves/valores."""
        out: dict = {}
        try:
            if not path or not os.path.isfile(path):
                return out
            with open(path, 'r', encoding='utf-8') as f:
                for line in f:
                    line = (line or '').strip()
                    if not line or line.startswith('#'):
                        continue
                    if line.lower().startswith('export '):
                        line = line[7:].strip()
                    if '=' in line:
                        k, v = line.split('=', 1)
                    elif ':' in line:
                        k, v = line.split(':', 1)
                    else:
                        continue
                    k = (k or '').strip()
                    v = (v or '').strip()
                    if not k:
                        continue
                    if (len(v) >= 2) and ((v[0] == v[-1]) and v[0] in ('"', "'")):
                        v = v[1:-1]
                    out[k] = v
        except Exception:
            return out
        return out

    @staticmethod
    def _normalize_auth(raw: str) -> str:
        """Normaliza un token/cabecera de autorización añadiendo el prefijo 'Bearer ' si falta."""
        raw = str(raw or '').strip()
        if not raw:
            return ''
        if raw.lower().startswith('bearer '):
            return raw
        if ' ' not in raw:
            return 'Bearer ' + raw
        return raw

    @staticmethod
    def _format_value_preview(v: Any, max_len: int = 140) -> str:
        """Devuelve una representación corta de un valor JSON para mostrar en UI."""
        try:
            if v is None:
                return 'null'
            if isinstance(v, bool):
                return 'true' if v else 'false'
            if isinstance(v, (int, float)) and not isinstance(v, bool):
                return str(v)
            if isinstance(v, str):
                s = json.dumps(v, ensure_ascii=False)
                if len(s) > max_len:
                    s = s[:max_len - 1] + '…'
                return s
            if isinstance(v, list):
                return f"<array len={len(v)}>"
            if isinstance(v, dict):
                return f"<object keys={len(v)}>"
            s = str(v)
            if len(s) > max_len:
                s = s[:max_len - 1] + '…'
            return s
        except Exception:
            return '<unavailable>'

    @staticmethod
    def _format_value_for_input(val: Any) -> str:
        """Convierte un valor JSON a cadena editable para un Entry de la UI."""
        if isinstance(val, bool):
            return 'true' if val else 'false'
        if val is None:
            return 'null'
        if isinstance(val, (int, float)) and not isinstance(val, bool):
            return str(val)
        if isinstance(val, str):
            return val
        try:
            return json.dumps(val, ensure_ascii=False)
        except Exception:
            return str(val)

    @staticmethod
    def _strip_wrapping_quotes(s: str) -> str:
        """Elimina comillas sencillas o dobles que envuelvan un string."""
        s = str(s or '').strip()
        if len(s) >= 2 and s[0] == s[-1] and s[0] in ('"', "'"):
            return s[1:-1].strip()
        return s

    @staticmethod
    def _looks_like_raw_token(s: str) -> bool:
        """Devuelve True si el string parece un token JWT o UUID sin el prefijo Bearer."""
        s = InstalacionGUI._strip_wrapping_quotes(s)
        if not s:
            return False
        if s.count('.') == 2 and len(s) > 40:
            return True
        if re.fullmatch(r"[0-9a-fA-F]{8}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-[0-9a-fA-F]{12}", s):
            return True
        return False

    @staticmethod
    def _extract_auth_from_text(text: str) -> str:
        """Extrae un valor de cabecera Authorization desde texto arbitrario (portapapeles, env, JSON de Postman)."""
        text = str(text or '').strip()
        if not text:
            return ''

        try:
            parsed = json.loads(text)
        except Exception:
            parsed = None

        def _from_any(obj: Any) -> str:
            if obj is None:
                return ''
            if isinstance(obj, str):
                s = InstalacionGUI._strip_wrapping_quotes(obj)
                if s.lower().startswith('bearer '):
                    return 'Bearer ' + s.split(None, 1)[1].strip() if len(s.split(None, 1)) == 2 else s
                m = re.search(r"(?i)\\bbearer\\s+([^\\s\"']+)", s)
                if m:
                    return 'Bearer ' + m.group(1).strip()
                if InstalacionGUI._looks_like_raw_token(s):
                    return 'Bearer ' + s
                return ''
            if isinstance(obj, dict):
                for k in ('Authorization', 'authorization', 'token', 'access_token', 'bearer', 'Bearer'):
                    if k in obj:
                        v = _from_any(obj.get(k))
                        if v:
                            return v
                for v in obj.values():
                    vv = _from_any(v)
                    if vv:
                        return vv
            if isinstance(obj, list):
                for it in obj:
                    vv = _from_any(it)
                    if vv:
                        return vv
            return ''

        auth = _from_any(parsed) if parsed is not None else ''
        if auth:
            return auth

        t = InstalacionGUI._strip_wrapping_quotes(text)
        m = re.search(r"(?i)\\bauthorization\\s*[:=]\\s*(bearer\\s+[^\\r\\n\"']+)", t)
        if m:
            v = m.group(1).strip()
            return v if v.lower().startswith('bearer ') else ('Bearer ' + v)

        m = re.search(r"(?i)\\bbearer\\s+([^\\s\"']+)", t)
        if m:
            return 'Bearer ' + m.group(1).strip()

        if InstalacionGUI._looks_like_raw_token(t):
            return 'Bearer ' + t

        return ''

    def _resolve_initial_auth(self, dotenv: dict) -> str:
        """Resuelve el token de autorización inicial desde caché, env, .env o portapapeles."""
        cached = self._normalize_auth(str(getattr(self, '_remote_edit_last_auth', '') or ''))
        if cached:
            return cached
        for k in ("DCR_APIM_AUTHORIZATION", "DCR_AUTHORIZATION", "APIM_AUTHORIZATION"):
            v = self._normalize_auth((os.getenv(k, '') or '').strip())
            if v:
                return v
        for k in ("DCR_APIM_AUTHORIZATION", "DCR_AUTHORIZATION", "APIM_AUTHORIZATION"):
            v = self._normalize_auth(str(dotenv.get(k, '') or '').strip())
            if v:
                return v
        for k in ("DCR_APIM_BEARER_TOKEN", "DCR_BEARER_TOKEN", "DCR_APIM_TOKEN", "DCR_TOKEN", "APIM_BEARER_TOKEN", "APIM_TOKEN"):
            v = (os.getenv(k, '') or '').strip()
            if v:
                return self._normalize_auth(v)
        for k in ("DCR_APIM_BEARER_TOKEN", "DCR_BEARER_TOKEN", "DCR_APIM_TOKEN", "DCR_TOKEN", "APIM_BEARER_TOKEN", "APIM_TOKEN"):
            v = str(dotenv.get(k, '') or '').strip()
            if v:
                return self._normalize_auth(v)
        try:
            clip = self.root.clipboard_get()
            ex = self._extract_auth_from_text(clip)
            if ex:
                return ex
        except Exception:
            pass
        return ''

    def _json_type_name(self, v: Any) -> str:
        if v is None:
            return "null"
        if isinstance(v, bool):
            return "boolean"
        if isinstance(v, (int, float)) and not isinstance(v, bool):
            return "number"
        if isinstance(v, str):
            return "string"
        if isinstance(v, list):
            return "array"
        if isinstance(v, dict):
            return "object"
        return type(v).__name__

    def _unescape_json_pointer_token(self, token: str) -> str:
        # RFC 6901
        return token.replace('~1', '/').replace('~0', '~')

    def _escape_json_pointer_token(self, token: str) -> str:
        # RFC 6901
        return str(token).replace('~', '~0').replace('/', '~1')

    def _pointer_join(self, *parts: str) -> str:
        out = ""
        for p in parts:
            p = str(p or "")
            if not p:
                continue
            if not p.startswith('/'):
                p = '/' + p
            out += p
        return out or '/'

    def _get_by_pointer(self, obj: Any, pointer: str) -> Tuple[bool, Any, str]:
        """Return (exists, value, reason)."""
        pointer = str(pointer or "").strip()
        if pointer in ("", "/"):
            return True, obj, ""
        if not pointer.startswith('/'):
            return False, None, "Pointer must start with '/'"

        cur = obj
        for raw_tok in pointer.split('/')[1:]:
            tok = self._unescape_json_pointer_token(raw_tok)
            if isinstance(cur, dict):
                if tok in cur:
                    cur = cur[tok]
                else:
                    return False, None, f"Missing key: {tok}"
            elif isinstance(cur, list):
                if tok == '-':
                    return False, None, "'-' is only valid for add at the end"
                try:
                    idx = int(tok)
                except Exception:
                    return False, None, f"Invalid list index: {tok}"
                if idx < 0 or idx >= len(cur):
                    return False, None, f"Index out of range: {idx}"
                cur = cur[idx]
            else:
                return False, None, "Cannot traverse into scalar"
        return True, cur, ""

    def _dotpath_to_pointer_simple(self, dot_path: str) -> Tuple[bool, str, str]:
        """Convert a simple dot path like 'config.entryType' or 'a.b[0].c' to JSON Pointer.

        Limitations: does not support selectors like '[type=101]' or multi-match syntax.
        """
        s = str(dot_path or "").strip()
        if not s:
            return False, "", "Path is empty"
        if s.startswith('/'):
            return True, s, ""
        if '[' in s and ']' in s:
            # Support [0] indices, reject selectors
            if re.search(r"\[[^\d\]]+\]", s):
                return False, "", "Selectors like [field=value] are not supported here. Use a JSON Pointer path instead."

        tokens: List[str] = []
        i = 0
        buf = ""
        while i < len(s):
            ch = s[i]
            if ch == '.':
                if buf:
                    tokens.append(buf)
                    buf = ""
                i += 1
                continue
            if ch == '[':
                # flush current key
                if buf:
                    tokens.append(buf)
                    buf = ""
                j = s.find(']', i)
                if j < 0:
                    return False, "", "Unclosed [index]"
                inner = s[i+1:j].strip()
                if not inner.isdigit():
                    return False, "", "Only numeric [index] is supported in this editor"
                tokens.append(inner)
                i = j + 1
                continue
            buf += ch
            i += 1
        if buf:
            tokens.append(buf)

        ptr = "".join('/' + self._escape_json_pointer_token(t) for t in tokens)
        return True, ptr or '/', ""

    def _find_node_index_by_serial(self, data: Any, serial: str) -> Tuple[Optional[int], str]:
        serial_n = self._normalize_serial(serial)
        if not serial_n:
            return None, "Serial number is empty"
        if not isinstance(data, dict):
            return None, "JSON root is not an object"
        nodes = data.get('nodes')
        if not isinstance(nodes, list):
            return None, "No 'nodes' array found in JSON"
        matches = []
        for idx, node in enumerate(nodes):
            if not isinstance(node, dict):
                continue
            sn = node.get('serialNumber')
            if self._normalize_serial(sn) == serial_n:
                matches.append(idx)
        if not matches:
            return None, f"No node found with serialNumber={serial_n}"
        if len(matches) > 1:
            return None, f"Multiple nodes found with serialNumber={serial_n}: {matches}"
        return matches[0], ""

    def open_remote_edit_apim(self):
        """Open a separate, gated remote edit window (APIM PATCH config/changes)."""
        if self._last_json_data is None:
            messagebox.showinfo("Remote Edit", "No JSON loaded. Query an installation first.")
            return
        if requests is None:
            messagebox.showerror("Remote Edit", f"requests is not available: {_requests_import_error}")
            return

        try:
            if getattr(self, '_remote_edit_win', None) is not None and self._remote_edit_win.winfo_exists():
                try:
                    self._remote_edit_win.lift()
                    return
                except Exception:
                    pass
        except Exception:
            pass

        win = tk.Toplevel(self.root)
        self._remote_edit_win = win
        win.title("Remote Edit (APIM)")
        try:
            win.transient(self.root)
        except Exception:
            pass

        frm = ttk.Frame(win, padding=10)
        frm.pack(fill='both', expand=True)

        ttk.Label(frm, text="Remote Edit (APIM)", font=('Segoe UI', 11, 'bold')).grid(row=0, column=0, columnspan=4, sticky='w')
        ttk.Label(frm, text="This window is gated to avoid accidental changes.").grid(row=1, column=0, columnspan=4, sticky='w', pady=(2, 10))

        # Gate
        enable_var = tk.BooleanVar(value=False)
        phrase_var = tk.StringVar(value="")
        chk_enable = ttk.Checkbutton(frm, text="Enable editing (danger)", variable=enable_var)
        chk_enable.grid(row=2, column=0, columnspan=2, sticky='w')
        ttk.Label(frm, text="Type EDIT to unlock:").grid(row=2, column=2, sticky='e')
        ent_phrase = ttk.Entry(frm, textvariable=phrase_var, width=10)
        ent_phrase.grid(row=2, column=3, sticky='w')

        # Defaults
        dotenv = {}
        try:
            dotenv.update(self._read_dotenv_file(os.path.join(os.path.dirname(__file__), '.env')))
        except Exception:
            pass
        try:
            dotenv.update(self._read_dotenv_file(os.path.join(os.getcwd(), '.env')))
        except Exception:
            pass

        default_base = (
            os.getenv("DCR_APIM_URL", "").strip()
            or os.getenv("DCR_APIM_BASE_URL", "").strip()
            or str(dotenv.get("DCR_APIM_URL", "")).strip()
            or str(dotenv.get("DCR_APIM_BASE_URL", "")).strip()
            or os.getenv("DCR_BASE_URL", DEFAULT_DCR_BASE_URL).strip()
        )

        VALUE_MODE_FIXED = "Auto (infer)"
        PERSISTENT_FIXED = True

        def _resolve_base_now() -> str:
            return (default_base or "").strip()
        country_var = tk.StringVar(value=(self.country_var.get() or "").strip().upper())
        inst_var = tk.StringVar(value=(self.install_var.get() or "").strip())
        serial_var = tk.StringVar(value="")
        path_var = tk.StringVar(value="/config/entryType")
        new_value_var = tk.StringVar(value="1")
        allow_add_var = tk.BooleanVar(value=False)
        current_value_var = tk.StringVar(value="")
        current_value_info_var = tk.StringVar(value="")

        auth_var = tk.StringVar(value=self._resolve_initial_auth(dotenv))
        req_by_app_default = (os.getenv("DCR_REQUESTED_BY_APP", "").strip() or str(dotenv.get("DCR_REQUESTED_BY_APP", "")).strip() or "Rnd").strip()
        # Fixed default for now (per request).
        req_by_user_var = tk.StringVar(value="Rnd")

        r = 3
        ttk.Separator(frm, orient='horizontal').grid(row=r, column=0, columnspan=4, sticky='we', pady=(10, 10)); r += 1

        ttk.Label(frm, text="x-sd-requested-by-user:").grid(row=r, column=0, sticky='w')
        ent_user = ttk.Entry(frm, textvariable=req_by_user_var, width=30)
        ent_user.grid(row=r, column=1, sticky='w', pady=(0, 6)); r += 1

        ttk.Label(frm, text="Country:").grid(row=r, column=0, sticky='w')
        ent_country = ttk.Entry(frm, textvariable=country_var, width=6)
        ent_country.grid(row=r, column=1, sticky='w', pady=(0, 6))
        ttk.Label(frm, text="Installation:").grid(row=r, column=2, sticky='e')
        ent_inst = ttk.Entry(frm, textvariable=inst_var, width=14)
        ent_inst.grid(row=r, column=3, sticky='w', pady=(0, 6)); r += 1

        ttk.Label(frm, text="Node serialNumber:").grid(row=r, column=0, sticky='w')
        ent_serial = ttk.Entry(frm, textvariable=serial_var, width=24)
        ent_serial.grid(row=r, column=1, sticky='w', pady=(0, 6))
        ttk.Label(frm, text="Path (pointer or simple dot):").grid(row=r, column=2, sticky='e')
        ent_path = ttk.Entry(frm, textvariable=path_var, width=28)
        ent_path.grid(row=r, column=3, sticky='w', pady=(0, 6)); r += 1

        ttk.Label(frm, text="Current value:").grid(row=r, column=0, sticky='w')
        ent_cur = ttk.Entry(frm, textvariable=current_value_var, width=70, state='readonly')
        ent_cur.grid(row=r, column=1, columnspan=3, sticky='we', pady=(0, 2)); r += 1
        ttk.Label(frm, textvariable=current_value_info_var, foreground="#666666").grid(row=r, column=1, columnspan=3, sticky='w', pady=(0, 6)); r += 1

        ttk.Label(frm, text="New value:").grid(row=r, column=0, sticky='w')
        ent_val = ttk.Entry(frm, textvariable=new_value_var, width=24)
        ent_val.grid(row=r, column=1, sticky='w', pady=(0, 6))
        chk_allow_add = ttk.Checkbutton(frm, text="Allow add (create new key)", variable=allow_add_var)
        chk_allow_add.grid(row=r, column=2, columnspan=2, sticky='w', pady=(0, 6)); r += 1

        ttk.Label(frm, text="Preview / Response:").grid(row=r, column=0, sticky='nw', pady=(6, 0))
        txt = tk.Text(frm, width=88, height=14, wrap='word', font=('Consolas', 9))
        txt.grid(row=r, column=1, columnspan=3, sticky='we', pady=(6, 0)); r += 1

        def _write(s: str):
            try:
                txt.delete('1.0', 'end')
                txt.insert('1.0', s)
            except Exception:
                pass

        def _is_unlocked() -> bool:
            return bool(enable_var.get()) and phrase_var.get().strip().upper() == 'EDIT'

        def _resolve_auth_now() -> str:
            """Resuelve el token bearer desde memoria, env, .env o portapapeles."""
            cached = self._normalize_auth(auth_var.get())
            if cached:
                return cached
            return self._resolve_initial_auth(dotenv)

        def _parse_new_value(raw: str, mode: str, exists: bool, old_val: Any) -> Tuple[Any, str, str]:
            raw = str(raw or "").strip()
            mode = str(mode or "").strip().lower()
            if raw == "":
                raise ValueError("New value is empty")

            def _try_json_load(s: str) -> Tuple[bool, Any]:
                try:
                    return True, json.loads(s)
                except Exception:
                    return False, None

            # Force modes
            if mode.startswith('json'):
                ok, val = _try_json_load(raw)
                if not ok:
                    raise ValueError("Mode=JSON: value must be valid JSON (e.g. true/false/null, 123, \"text\")")
                return val, "json", ""
            if mode.startswith('string'):
                # If user typed a JSON string, decode escapes.
                if len(raw) >= 2 and raw[0] == raw[-1] and raw[0] in ('\"', "'"):
                    ok, val = _try_json_load(raw)
                    if ok and isinstance(val, str):
                        return val, "string(json)", ""
                return raw, "string", ""

            # Auto mode
            if exists:
                # boolean
                if isinstance(old_val, bool):
                    rl = raw.strip().lower()
                    if rl in ('true', 'false'):
                        return (rl == 'true'), "auto:boolean", ""
                    if raw in ('True', 'False'):
                        return (raw == 'True'), "auto:boolean", ""
                    raise ValueError("Expected a boolean for this path. Use true/false (or True/False).")

                # number (avoid bool which is also int)
                if isinstance(old_val, (int, float)) and not isinstance(old_val, bool):
                    ok, val = _try_json_load(raw)
                    if ok and isinstance(val, (int, float)) and not isinstance(val, bool):
                        return val, "auto:number(json)", ""
                    try:
                        if any(c in raw for c in ('.', 'e', 'E')):
                            return float(raw), "auto:number", ""
                        return int(raw), "auto:number", ""
                    except Exception:
                        raise ValueError("Expected a number for this path (e.g. 10 or 3.14).")

                # string
                if isinstance(old_val, str):
                    if len(raw) >= 2 and raw[0] == raw[-1] and raw[0] in ('\"', "'"):
                        ok, val = _try_json_load(raw)
                        if ok and isinstance(val, str):
                            return val, "auto:string(json)", ""
                    return raw, "auto:string", ""

                # null
                if old_val is None:
                    rl = raw.strip().lower()
                    if rl in ('null', 'none'):
                        return None, "auto:null", ""
                    raise ValueError("Expected null for this path. Use null.")

                # object/array: require JSON
                if isinstance(old_val, (dict, list)):
                    ok, val = _try_json_load(raw)
                    if not ok:
                        raise ValueError("Expected a JSON object/array for this path. Use Mode=JSON and paste valid JSON.")
                    if isinstance(old_val, dict) and isinstance(val, dict):
                        return val, "auto:object(json)", ""
                    if isinstance(old_val, list) and isinstance(val, list):
                        return val, "auto:array(json)", ""
                    raise ValueError("Type mismatch: current value is %s but new value is %s" % (self._json_type_name(old_val), self._json_type_name(val)))

            # If missing, fall back to JSON-then-string (but block common non-JSON literals)
            ok, val = _try_json_load(raw)
            if ok:
                return val, "json", ""
            if raw in ("True", "False", "None"):
                raise ValueError(
                    "Value is not valid JSON. Use true/false/null (lowercase) or set Mode=String if you really want a string."
                )
            return raw, "string", ""

        def _compute_full_ptr_for_ui() -> Tuple[bool, str, str]:
            """Return (ok, full_ptr, target_desc_or_error)."""
            ok_ptr, ptr, ptr_reason = self._dotpath_to_pointer_simple(path_var.get().strip())
            if not ok_ptr:
                return False, "", ptr_reason

            serial = serial_var.get().strip()
            idx = None
            if serial:
                idx, node_reason = self._find_node_index_by_serial(self._last_json_data, serial)
                if idx is None:
                    return False, "", node_reason

            if ptr.startswith('/nodes/'):
                return True, ptr, "Target: absolute pointer"
            if idx is not None:
                full_ptr = self._pointer_join(f"/nodes/{idx}", ptr)
                return True, full_ptr, f"Target: node serialNumber={self._normalize_serial(serial)} -> nodes[{idx}]"
            return True, ptr, "Target: root JSON"

        def _refresh_current_value(*_):
            try:
                ok, full_ptr, target_or_err = _compute_full_ptr_for_ui()
                if not ok:
                    current_value_var.set("")
                    current_value_info_var.set(f"{target_or_err}")
                    return

                exists, old_val, why = self._get_by_pointer(self._last_json_data, full_ptr)
                if exists:
                    current_value_var.set(self._format_value_for_input(old_val))
                    current_value_info_var.set(f"{target_or_err} • type={self._json_type_name(old_val)}")
                else:
                    current_value_var.set("<missing>")
                    current_value_info_var.set(f"{target_or_err} • {why}")
            except Exception as e:
                try:
                    current_value_var.set("")
                    current_value_info_var.set(f"Error reading current value: {e}")
                except Exception:
                    pass

        last_checked = {"ok": False, "url": None, "headers": None, "payload": None, "stale_reason": ""}

        def _invalidate_check(reason: str = "inputs changed") -> None:
            """Disable Send until user re-runs Check after editing any inputs."""
            try:
                if last_checked.get("ok") or last_checked.get("payload"):
                    last_checked["ok"] = False
                    last_checked["url"] = None
                    last_checked["headers"] = None
                    last_checked["payload"] = None
                    last_checked["stale_reason"] = str(reason or "inputs changed")
                else:
                    # Keep a reason if we don't have one yet
                    if not last_checked.get("stale_reason"):
                        last_checked["stale_reason"] = str(reason or "inputs changed")
            except Exception:
                pass
            try:
                _refresh_buttons()
            except Exception:
                pass

        def _parent_pointer(ptr: str) -> str:
            ptr = str(ptr or '').strip()
            if ptr in ('', '/'):
                return '/'
            if not ptr.startswith('/'):
                return '/'
            parts = ptr.split('/')
            if len(parts) <= 2:
                return '/'
            out = '/'.join(parts[:-1])
            return out if out else '/'

        def _last_pointer_token(ptr: str) -> str:
            ptr = str(ptr or '').strip()
            if not ptr or ptr == '/':
                return ''
            parts = ptr.split('/')
            raw = parts[-1] if parts else ''
            try:
                return self._unescape_json_pointer_token(raw)
            except Exception:
                return raw

        def _validate_patch_target(full_ptr: str, exists: bool, old_val: Any, why: str) -> Tuple[str, str]:
            """Return (op, note). Raises ValueError if path is not patchable."""
            if exists:
                return 'replace', 'OK: existing path (replace)'

            # Safer default: do not allow creating new keys unless explicitly enabled.
            if not bool(allow_add_var.get()):
                raise ValueError(
                    "Path does not exist. For safety, this tool blocks creating new keys by default. "
                    "Enable 'Allow add (create new key)' if you really want to add it."
                )

            # Missing target => allow 'add' only if parent exists and is a container.
            parent_ptr = _parent_pointer(full_ptr)
            p_exists, p_val, p_why = self._get_by_pointer(self._last_json_data, parent_ptr)
            if not p_exists:
                raise ValueError(f"Invalid path: parent is missing at {parent_ptr} ({p_why})")
            if not isinstance(p_val, (dict, list)):
                raise ValueError(f"Invalid path: parent {parent_ptr} is not an object/array (type={self._json_type_name(p_val)})")

            # If parent is list, validate index or '-'
            if isinstance(p_val, list):
                tok = _last_pointer_token(full_ptr)
                if tok == '-':
                    return 'add', 'OK: append to array (add /-)'
                try:
                    idx = int(tok)
                except Exception:
                    raise ValueError("Invalid path: array target must be a numeric index or '-' for append")
                if idx < 0 or idx > len(p_val):
                    raise ValueError(f"Invalid path: array index out of range for add (idx={idx}, len={len(p_val)})")
                return 'add', 'OK: array index add'

            # parent is dict
            return 'add', 'OK: missing leaf under existing object (add)'

        def _build_and_check() -> None:
            try:
                self.status_var.set("Remote Edit: verificando...")
                ok, full_ptr, _target_desc = _compute_full_ptr_for_ui()
                if not ok:
                    last_checked["ok"] = False
                    _write(f"CHECK: FAIL\nPath error: {_target_desc}")
                    _refresh_buttons()
                    self.status_var.set("Check: FAIL")
                    return

                exists, old_val, why = self._get_by_pointer(self._last_json_data, full_ptr)
                # Validate the target/path so we don't enable Send when it's not patchable
                try:
                    op, path_note = _validate_patch_target(full_ptr, exists, old_val, why)
                except Exception as e:
                    last_checked["ok"] = False
                    _write(f"CHECK: FAIL\nPath validation error: {e}")
                    _refresh_buttons()
                    return

                # Validación estricta de tipo y campo editable
                try:
                    new_val, parse_mode, parse_note = _parse_new_value(new_value_var.get(), VALUE_MODE_FIXED, exists, old_val)
                except Exception as e:
                    last_checked["ok"] = False
                    _write(f"CHECK: FAIL\nValue error: {e}")
                    _refresh_buttons()
                    return

                # Si el campo existe, solo permitir si el tipo coincide exactamente
                if exists:
                    tipo_actual = type(old_val)
                    tipo_nuevo = type(new_val)
                    if isinstance(old_val, (int, float)) and isinstance(new_val, (int, float)) and not isinstance(old_val, bool) and not isinstance(new_val, bool):
                        pass  # ok
                    elif tipo_actual != tipo_nuevo:
                        last_checked["ok"] = False
                        _write(f"CHECK: FAIL\nTipo incompatible: el campo actual es {tipo_actual.__name__} y el nuevo valor es {tipo_nuevo.__name__}. Deben coincidir exactamente.")
                        _refresh_buttons()
                        return
                    if isinstance(old_val, bool) and not isinstance(new_val, bool):
                        last_checked["ok"] = False
                        _write("CHECK: FAIL\nTipo incompatible: el campo actual es booleano y el nuevo valor no lo es. Usa true/false.")
                        _refresh_buttons()
                        return
                    if isinstance(old_val, (dict, list)):
                        raw_val = new_value_var.get().strip()
                        try:
                            val_json = json.loads(raw_val)
                            if not isinstance(val_json, tipo_actual):
                                raise Exception()
                        except Exception:
                            last_checked["ok"] = False
                            _write(f"CHECK: FAIL\nValue must be valid JSON of type {tipo_actual.__name__} for this path.")
                            _refresh_buttons()
                            return
                # Si el campo no existe, bloquear salvo que se permita crear
                if not exists and not bool(allow_add_var.get()):
                    last_checked["ok"] = False
                    _write("CHECK: FAIL\nEl campo no existe y no está permitido crearlo (habilita 'Allow add' para forzar).")
                    _refresh_buttons()
                    return

                payload = {
                    "persistent": bool(PERSISTENT_FIXED),
                    "request": [
                        {
                            "op": op,
                            "path": full_ptr,
                            "value": new_val,
                        }
                    ],
                }

                preview = ["CHECK: OK"]
                preview.append(_target_desc)
                preview.append(f"Full path: {full_ptr}")
                if exists:
                    preview.append(f"Current: {old_val!r}  (type={self._json_type_name(old_val)})")
                else:
                    preview.append(f"Current: <missing>  ({why})")
                preview.append(path_note)
                preview.append(f"Mode: {VALUE_MODE_FIXED} (fixed)")
                preview.append(f"persistent: {str(bool(PERSISTENT_FIXED)).lower()} (fixed)")
                preview.append(f"New: {new_val!r}  (type={self._json_type_name(new_val)}, input={parse_mode})")
                if parse_note:
                    preview.append(parse_note)
                preview.append(f"Recommended op: {op}")
                preview.append("")
                preview.append("Request JSON:")
                preview.append(json.dumps(payload, indent=2, ensure_ascii=False))
                _write("\n".join(preview))

                # Cache for send
                base = _resolve_base_now()
                if not base:
                    last_checked["ok"] = False
                    _write("CHECK: FAIL\nAPIM base URL is not configured. Set DCR_APIM_URL (or DCR_APIM_BASE_URL) in Ultrajson/.env or as an environment variable.")
                    _refresh_buttons()
                    return
                url = f"{base.rstrip('/')}/device-support/device-config-repository/v2.0/config/changes"

                ctry = (country_var.get() or "").strip().upper()
                ins = (inst_var.get() or "").strip()
                if not ins.isdigit():
                    last_checked["ok"] = False
                    _write("CHECK: FAIL\nInstallation must be numeric")
                    _refresh_buttons()
                    return
                if not ctry or len(ctry) != 2 or not ctry.isalpha():
                    last_checked["ok"] = False
                    _write("CHECK: FAIL\nCountry must be a 2-letter code")
                    _refresh_buttons()
                    return

                last_checked["ok"] = True
                last_checked["url"] = (url, f"{ctry}{ins}")
                last_checked["headers"] = None
                last_checked["payload"] = payload
                last_checked["stale_reason"] = ""
                _refresh_buttons()
                self.status_var.set(f"Check OK — {ctry}{ins} {full_ptr}")
            except Exception as e:
                last_checked["ok"] = False
                _write("CHECK: FAIL\n" + "Error during check: " + str(e) + "\n\n" + traceback.format_exc())
                _refresh_buttons()
                self.status_var.set("Check: FAIL")

                # --- Botón para editar experimental ---
            def _edit_experimental():
                try:
                    exp_path = "/config/experimental"
                    exists, old_val, _ = self._get_by_pointer(self._last_json_data, exp_path)
                    old_json = ""
                    if exists:
                        try:
                            old_json = json.dumps(old_val, indent=2, ensure_ascii=False)
                        except Exception:
                            old_json = str(old_val)
                    else:
                        old_json = "{}"  # Por defecto objeto vacío, pero editable
                    edit_win = tk.Toplevel(win)
                    edit_win.title("Editar experimental (JSON libre)")
                    edit_win.geometry("600x400")
                    txt_exp = tk.Text(edit_win, font=("Consolas", 10))
                    txt_exp.pack(fill="both", expand=True, padx=8, pady=8)
                    txt_exp.insert("1.0", old_json)
                    msg_var = tk.StringVar(value="Pegue aquí el JSON experimental. Puede ser objeto o lista. Solo se valida la sintaxis.")
                    lbl = ttk.Label(edit_win, textvariable=msg_var, foreground="#666666")
                    lbl.pack(anchor="w", padx=8)
                    def guardar():
                        raw = txt_exp.get("1.0", "end").strip()
                        if not raw:
                            nuevo = {}  # Por defecto objeto vacío
                        else:
                            try:
                                nuevo = json.loads(raw)
                            except Exception as e:
                                msg_var.set(f"Error de sintaxis JSON: {e}")
                                return
                        # Actualizar el valor en memoria respetando el tipo
                        ptrs = exp_path.strip("/").split("/")
                        obj = self._last_json_data
                        for p in ptrs[:-1]:
                            p = self._unescape_json_pointer_token(p)
                            if isinstance(obj, dict):
                                obj = obj.setdefault(p, {})
                            else:
                                msg_var.set("No se pudo navegar hasta experimental (estructura inesperada)")
                                return
                        last = self._unescape_json_pointer_token(ptrs[-1])
                        obj[last] = nuevo
                        msg_var.set("Experimental actualizado en memoria. Cierra la ventana y usa 'Check' y 'Send' para aplicar.")
                    btn_guardar = ttk.Button(edit_win, text="Guardar experimental", command=guardar)
                    btn_guardar.pack(pady=8)
                except Exception as e:
                    messagebox.showerror("Experimental", f"Error: {e}", parent=win)


        def _copy_request() -> None:
            try:
                if not last_checked.get("payload"):
                    messagebox.showinfo("Copy", "Run Check first", parent=win)
                    return
                data = json.dumps(last_checked["payload"], indent=2, ensure_ascii=False)
                self.root.clipboard_clear()
                self.root.clipboard_append(data)
                try:
                    self.root.update_idletasks()
                except Exception:
                    pass
                self.status_var.set("Remote Edit: request JSON copied")
            except Exception as e:
                messagebox.showerror("Copy", f"Error copying: {e}", parent=win)

        def _refresh_json() -> None:
            """Requery installation JSON and refresh current-value display."""
            try:
                base = _resolve_base_now()
                if not base:
                    messagebox.showwarning(
                        "Refresh",
                        "APIM base URL is not configured. Set DCR_APIM_URL (or DCR_APIM_BASE_URL) in Ultrajson/.env or as an environment variable.",
                        parent=win,
                    )
                    return

                ctry = (country_var.get() or "").strip().upper()
                ins = (inst_var.get() or "").strip()
                if not ins.isdigit():
                    messagebox.showwarning("Refresh", "Installation must be numeric.", parent=win)
                    return
                if not ctry or len(ctry) != 2 or not ctry.isalpha():
                    messagebox.showwarning("Refresh", "Country must be a 2-letter code.", parent=win)
                    return

                url = f"{base.rstrip('/')}/device-support/device-config-repository/v2.0/installation/{ctry}{ins}"

                headers = {}
                auth = _resolve_auth_now()
                if auth:
                    headers["Authorization"] = auth

                resp = requests.get(url, headers=headers, timeout=20)
                resp.raise_for_status()
                data = resp.json()

                # Update the live JSON used by the editor
                self._last_json_data = data

                # Update main viewer too (best-effort)
                try:
                    self.text.delete('1.0', 'end')
                    self.text.insert('1.0', json.dumps(data, indent=2, ensure_ascii=False))
                except Exception:
                    pass

                # After refresh, require a new Check
                _invalidate_check("refreshed JSON")
                _refresh_current_value()

                _write(f"REFRESH: OK\nHTTP {resp.status_code}\nFetched {ctry}{ins}\n\nTip: run Check again before Send.")
            except Exception as e:
                _write("REFRESH: FAIL\n" + str(e) + "\n\n" + traceback.format_exc())

        def _send() -> None:
            try:
                if not _is_unlocked():
                    messagebox.showwarning("Remote Edit", "Editing is locked. Check 'Enable editing' and type EDIT.", parent=win)
                    return
                if not last_checked.get("ok"):
                    messagebox.showinfo("Remote Edit", "Run Check first", parent=win)
                    return

                url, installation_id = last_checked["url"]
                payload = last_checked["payload"]

                auth = _resolve_auth_now()
                if not auth:
                    messagebox.showwarning(
                        "Remote Edit",
                        "No APIM token found.\n\n"
                        "Set it in Ultrajson/.env as DCR_APIM_BEARER_TOKEN=... (or DCR_APIM_AUTHORIZATION=Bearer ...)\n"
                        "or copy the token (or Authorization: Bearer ...) to clipboard and try again.",
                        parent=win,
                    )
                    return

                try:
                    auth_var.set(auth)
                    self._remote_edit_last_auth = auth
                except Exception:
                    pass

                headers = {
                    "Authorization": auth,
                    "Content-Type": "application/json",
                }
                app_h = (req_by_app_default or "").strip()
                user_h = (req_by_user_var.get() or "").strip()
                if not user_h:
                    messagebox.showwarning("Remote Edit", "x-sd-requested-by-user is required (who performs the change).", parent=win)
                    return
                if app_h:
                    headers["x-sd-requested-by-app"] = app_h
                headers["x-sd-requested-by-user"] = user_h

                if not messagebox.askyesno(
                    "Confirm Remote PATCH",
                    f"Send PATCH to:\n{url}\n\ninstallationId={installation_id}\n\nProceed?",
                    parent=win,
                ):
                    return

                self.status_var.set(f"Enviando PATCH {installation_id}...")
                resp = requests.patch(
                    url,
                    params={"installationId": installation_id},
                    headers=headers,
                    json=payload,
                    timeout=20,
                )
                out = [f"HTTP {resp.status_code}"]
                try:
                    out.append(json.dumps(resp.json(), indent=2, ensure_ascii=False))
                except Exception:
                    out.append(resp.text)
                _write("\n\n".join(out))
                self.status_var.set(f"PATCH {installation_id} → HTTP {resp.status_code}")
            except Exception as e:
                _write("Error sending PATCH: " + str(e) + "\n\n" + traceback.format_exc())
                self.status_var.set("PATCH: error")

        def _refresh_buttons(*_):
            unlocked = _is_unlocked()
            checked_ok = bool(last_checked.get("ok"))
            has_token = bool(_resolve_auth_now())
            try:
                btn_send.configure(state=('normal' if (unlocked and checked_ok and has_token) else 'disabled'))
            except Exception:
                pass

            try:
                missing = []
                if not unlocked:
                    missing.append("unlock (Enable editing + EDIT)")
                if not checked_ok:
                    if last_checked.get("stale_reason"):
                        missing.append("Check (inputs changed)")
                    else:
                        missing.append("Check")
                if not has_token:
                    missing.append("APIM token")
                if missing:
                    lbl_send_state.configure(text="Send disabled: missing " + ", ".join(missing))
                else:
                    lbl_send_state.configure(text="Send ready")
            except Exception:
                pass

        enable_var.trace_add('write', _refresh_buttons)
        phrase_var.trace_add('write', _refresh_buttons)

        # Keep the "Current value" display in sync with inputs.
        try:
            path_var.trace_add('write', _refresh_current_value)
        except Exception:
            pass
        try:
            serial_var.trace_add('write', _refresh_current_value)
        except Exception:
            pass

        # Initial fill
        _refresh_current_value()

        # Any edits to inputs invalidate the last successful Check.
        for v, why in (
            (country_var, "country changed"),
            (inst_var, "installation changed"),
            (serial_var, "node serialNumber changed"),
            (path_var, "path changed"),
            (new_value_var, "new value changed"),
            (req_by_user_var, "requested-by-user changed"),
        ):
            try:
                v.trace_add('write', lambda *_a, _why=why: _invalidate_check(_why))
            except Exception:
                pass
        try:
            allow_add_var.trace_add('write', lambda *_a: _invalidate_check("allow-add changed"))
        except Exception:
            pass

        # If user copies token after opening the window, allow auto-fill on focus
        def _maybe_autofill_on_focus(_evt=None):
            try:
                if (auth_var.get() or '').strip():
                    return
                clip = self.root.clipboard_get()
                extracted = self._extract_auth_from_text(clip)
                if extracted:
                    auth_var.set(extracted.strip())
                    try:
                        _refresh_buttons()
                    except Exception:
                        pass
            except Exception:
                pass

        try:
            win.bind('<FocusIn>', _maybe_autofill_on_focus)
        except Exception:
            pass

        # --- Forzar la barra de botones SIEMPRE en una fila nueva ---
        r = frm.grid_size()[1]  # Siempre ponerla en la siguiente fila libre
        btns = ttk.Frame(frm)
        btns.grid(row=r, column=0, columnspan=4, sticky='we', pady=(10, 0))

        # --- Definir _edit_experimental antes de usarlo ---
        def _edit_experimental():
            try:
                exp_path = "/config/experimental"
                exists, old_val, _ = self._get_by_pointer(self._last_json_data, exp_path)
                old_json = ""
                if exists:
                    try:
                        old_json = json.dumps(old_val, indent=2, ensure_ascii=False)
                    except Exception:
                        old_json = str(old_val)
                else:
                    old_json = "{}"  # Por defecto objeto vacío, pero editable
                edit_win = tk.Toplevel(win)
                edit_win.title("Editar experimental (JSON libre)")
                edit_win.geometry("600x400")
                txt_exp = tk.Text(edit_win, font=("Consolas", 10))
                txt_exp.pack(fill="both", expand=True, padx=8, pady=8)
                txt_exp.insert("1.0", old_json)
                msg_var = tk.StringVar(value="Pegue aquí el JSON experimental. Puede ser objeto o lista. Solo se valida la sintaxis.")
                lbl = ttk.Label(edit_win, textvariable=msg_var, foreground="#666666")
                lbl.pack(anchor="w", padx=8)

                def guardar():
                    raw = txt_exp.get("1.0", "end").strip()
                    if not raw:
                        nuevo = {}  # Por defecto objeto vacío
                    else:
                        try:
                            nuevo = json.loads(raw)
                        except Exception as e:
                            err = str(e)
                            extra = ""
                            if (
                                "Expecting property name enclosed in double quotes" in err
                                or "trailing comma" in err
                                or "char" in err and (",}" in raw or ",]" in raw)
                            ):
                                extra = "\nSugerencia: Revisa si hay comas de más al final de objetos o arrays. En JSON no puede haber una coma después del último elemento."
                            msg_var.set(f"Error de sintaxis JSON: {e}{extra}")
                            return
                    # Actualizar el valor en memoria respetando el tipo
                    ptrs = exp_path.strip("/").split("/")
                    obj = self._last_json_data
                    for p in ptrs[:-1]:
                        p = self._unescape_json_pointer_token(p)
                        if isinstance(obj, dict):
                            obj = obj.setdefault(p, {})
                        else:
                            msg_var.set("No se pudo navegar hasta experimental (estructura inesperada)")
                            return
                    last = self._unescape_json_pointer_token(ptrs[-1])
                    obj[last] = nuevo
                    msg_var.set("Experimental actualizado en memoria. Cierra la ventana y usa 'Check' y 'Send' para aplicar.")
                    # Refrescar valor actual en la ventana principal
                    try:
                        _refresh_current_value()
                    except Exception:
                        pass
                    # Si el path actual es experimental, actualiza también el campo New value
                    try:
                        if path_var.get().strip().replace("/","").lower() == "experimental":
                            new_value_var.set(json.dumps(nuevo, ensure_ascii=False, indent=None))
                    except Exception:
                        pass
                    try:
                        _invalidate_check("experimental edit")
                        _write("Experimental actualizado. Haz 'Check' para validar y enviar si es necesario.")
                    except Exception:
                        pass

                def anadir_objeto():
                    # Permite pegar un objeto y lo añade al array actual de experimental
                    raw = txt_exp.get("1.0", "end").strip()
                    if not raw:
                        msg_var.set("Pega aquí el objeto JSON que quieres añadir al array experimental.")
                        return
                    try:
                        nuevo_obj = json.loads(raw)
                    except Exception as e:
                        msg_var.set(f"Error de sintaxis JSON: {e}")
                        return
                    # Leer SIEMPRE el array real actual desde self._last_json_data
                    ptrs = exp_path.strip("/").split("/")
                    obj = self._last_json_data
                    for p in ptrs[:-1]:
                        p = self._unescape_json_pointer_token(p)
                        if isinstance(obj, dict):
                            obj = obj.setdefault(p, {})
                        else:
                            msg_var.set("No se pudo navegar hasta experimental (estructura inesperada)")
                            return
                    last = self._unescape_json_pointer_token(ptrs[-1])
                    arr = []
                    if last in obj and isinstance(obj[last], list):
                        arr = list(obj[last])
                    # Si el usuario pega un array, añadir todos los elementos
                    if isinstance(nuevo_obj, list):
                        arr.extend(nuevo_obj)
                    else:
                        arr.append(nuevo_obj)
                    obj[last] = arr
                    msg_var.set("Objeto añadido al array experimental. Cierra la ventana y usa 'Check' y 'Send' para aplicar.")
                    # Refrescar valor actual en la ventana principal
                    try:
                        _refresh_current_value()
                    except Exception:
                        pass
                    try:
                        if path_var.get().strip().replace("/","").lower() == "experimental":
                            new_value_var.set(json.dumps(arr, ensure_ascii=False, indent=None))
                    except Exception:
                        pass
                    try:
                        _invalidate_check("experimental add object")
                        _write("Experimental actualizado. Haz 'Check' para validar y enviar si es necesario.")
                    except Exception:
                        pass

                btn_guardar = ttk.Button(edit_win, text="Guardar experimental", command=guardar)
                btn_guardar.pack(pady=8)
                btn_anadir = ttk.Button(edit_win, text="Añadir objeto al array", command=anadir_objeto)
                btn_anadir.pack(pady=2)
            except Exception as e:
                messagebox.showerror("Experimental", f"Error: {e}", parent=win)

        btn_check = ttk.Button(btns, text="Check", command=_build_and_check)
        btn_check.pack(side='left')
        # --- Asegurar botón Editar experimental ---
        btn_exp = ttk.Button(btns, text="Editar experimental", command=_edit_experimental)
        btn_exp.pack(side='left', padx=(8, 0))
        # --- Fin botón Editar experimental ---
        def _simulate():
            try:
                ok, full_ptr, _target_desc = _compute_full_ptr_for_ui()
                if not ok:
                    _write(f"SIMULAR: FAIL\nPath error: {_target_desc}")
                    return
                exists, old_val, why = self._get_by_pointer(self._last_json_data, full_ptr)
                try:
                    new_val, parse_mode, parse_note = _parse_new_value(new_value_var.get(), VALUE_MODE_FIXED, exists, old_val)
                except Exception as e:
                    _write(f"SIMULAR: FAIL\nValue error: {e}")
                    return
                import copy
                json_sim = copy.deepcopy(self._last_json_data)
                # Navegar hasta el padre
                ptr_parts = full_ptr.strip('/').split('/')
                parent = json_sim
                for p in ptr_parts[:-1]:
                    p = self._unescape_json_pointer_token(p)
                    if isinstance(parent, dict):
                        parent = parent.get(p, {})
                    elif isinstance(parent, list):
                        try:
                            idx = int(p)
                            parent = parent[idx]
                        except Exception:
                            _write("SIMULAR: FAIL\nNo se pudo navegar hasta el padre del campo.")
                            return
                last = self._unescape_json_pointer_token(ptr_parts[-1])
                # Simular cambio
                if isinstance(parent, dict):
                    parent[last] = new_val
                elif isinstance(parent, list):
                    try:
                        idx = int(last)
                        parent[idx] = new_val
                    except Exception:
                        _write("SIMULAR: FAIL\nNo se pudo modificar el índice de la lista.")
                        return
                else:
                    _write("SIMULAR: FAIL\nNo se pudo modificar el campo (no es dict ni lista).")
                    return
                _write("SIMULAR: OK\nVista previa del JSON resultante:\n" + json.dumps(json_sim, indent=2, ensure_ascii=False))
            except Exception as e:
                _write(f"SIMULAR: FAIL\n{e}")

        ttk.Button(btns, text="Simular", command=_simulate).pack(side='left', padx=(8, 0))
        ttk.Button(btns, text="Refresh JSON", command=_refresh_json).pack(side='left', padx=(8, 0))
        ttk.Button(btns, text="Copy request JSON", command=_copy_request).pack(side='left', padx=(8, 0))
        btn_send = ttk.Button(btns, text="Send PATCH", command=_send)
        btn_send.pack(side='left', padx=(8, 0))
        lbl_send_state = ttk.Label(btns, text="", foreground="#666666")
        lbl_send_state.pack(side='left', padx=(10, 0))
        ttk.Button(btns, text="Close", command=win.destroy).pack(side='right')

        # Initialize send button state
        _refresh_buttons()

        try:
            frm.columnconfigure(1, weight=1)
        except Exception:
            pass

        try:
            ent_serial.focus_set()
        except Exception:
            pass

    def extraer_valor_ruta_gui(self):
        if self._last_json_data is None:
            messagebox.showinfo("Extract Path Value", "No JSON loaded. Query an installation first.")
            return
        # Extraction is a single-pane operation
        self._set_compare_mode(False)
        ruta = self.search_var.get().strip()
        if not ruta:
            messagebox.showinfo("Extract Path Value", "Enter a path in the search field.")
            return
        # Support for multi-path separated by commas
        rutas = [r.strip() for r in ruta.split(',') if r.strip()]
        self.text.delete('1.0','end')
        if len(rutas) == 1:
            query = rutas[0]
            valor = self.extraer_valor_ruta(self._last_json_data, query)

            def is_path_value_matches(v: Any) -> bool:
                if not isinstance(v, list) or not v:
                    return False
                first = v[0]
                return (
                    isinstance(first, (tuple, list))
                    and len(first) == 2
                    and isinstance(first[0], str)
                )

            if is_path_value_matches(valor):
                matches: List[Tuple[str, Any]] = valor  # type: ignore
                self.text.insert('end', f"# {len(matches)} matches for '{query}'\n\n")
                for p, v in matches:
                    if isinstance(v, dict) and list(v.keys()) == ['time'] and isinstance(v.get('time'), (str, int, float, bool)):
                        # Common case for antibounce: show the exact scalar path.
                        self.text.insert('end', f"Path: {p}.time\nValue: {v.get('time')}\n\n")
                    elif isinstance(v, (dict, list)):
                        self.text.insert('end', f"Path: {p}\nValue (object):\n" + json.dumps(v, indent=2, ensure_ascii=False) + "\n\n")
                    else:
                        self.text.insert('end', f"Path: {p}\nValue: {v}\n\n")
                self.text.see('1.0')
                return

            # If user typed a bare key (e.g. 'simNumber'), search the entire JSON for that key.
            if valor is None and not self._es_patron_ruta(query):
                matches = find_key_matches(self._last_json_data, query)
                if not matches:
                    self.text.insert('end', f"Path/key: {query}\nExtracted value: None")
                    self.text.see('end')
                    return
                if len(matches) == 1:
                    mpath, mval = matches[0]
                    if isinstance(mval, (dict, list)):
                        self.text.insert('end', f"Path: {mpath}\nExtracted value (object):\n" + json.dumps(mval, indent=2, ensure_ascii=False))
                    else:
                        self.text.insert('end', f"Path: {mpath}\nExtracted value: {mval}")
                    self.text.see('end')
                    return

                # Multiple matches -> show a compact table
                ancho_ruta = max(len(p) for p, _ in matches)
                self.text.insert('end', f"# Key matches for '{query}' ({len(matches)})\n")
                header = f"{'PATH'.ljust(ancho_ruta)} | VALUE"
                self.text.insert('end', header + "\n" + ('-' * len(header)) + "\n")
                for p, v in matches:
                    if isinstance(v, (dict, list)):
                        v_repr = f"<{type(v).__name__}>"
                    else:
                        v_repr = str(v)
                    self.text.insert('end', f"{p.ljust(ancho_ruta)} | {v_repr}\n")
                self.text.see('end')
                return

            if isinstance(valor,(dict,list)):
                self.text.insert('end', f"Path: {query}\nExtracted value (object):\n" + json.dumps(valor, indent=2, ensure_ascii=False))
            else:
                self.text.insert('end', f"Path: {query}\nExtracted value: {valor}")
            self.text.see('end'); return
        # Multi-path: table
        resultados = []
        for r in rutas:
            v = self.extraer_valor_ruta(self._last_json_data, r)
            # If the suffix-search returned (path,value) matches, keep it compact in the table.
            try:
                if isinstance(v, list) and v and isinstance(v[0], (tuple, list)) and len(v[0]) == 2 and isinstance(v[0][0], str):
                    v = f"MATCHES({len(v)})"
            except Exception:
                pass
            if v is None and not self._es_patron_ruta(r):
                ms = find_key_matches(self._last_json_data, r)
                if len(ms) == 1:
                    v = ms[0][1]
                elif len(ms) > 1:
                    v = f"MULTIPLE_MATCHES({len(ms)})"
            if isinstance(v,(dict,list)):
                v_repr = json.dumps(v, ensure_ascii=False)
            else:
                v_repr = str(v)
            resultados.append((r, v_repr))
        # Render
        ancho_ruta = max(len(r) for r,_ in resultados) if resultados else 10
        self.text.insert('end', f"# Multi-path ({len(resultados)})\n")
        header = f"{'PATH'.ljust(ancho_ruta)} | VALUE"
        self.text.insert('end', header + "\n" + ('-'*len(header)) + "\n")
        for r,v in resultados:
            self.text.insert('end', f"{r.ljust(ancho_ruta)} | {v}\n")
        self.text.see('end')
    # === Actions ===
    def consultar(self):
        inst = self.install_var.get().strip()
        country = self.country_var.get().strip().upper()
        if not inst:
            messagebox.showwarning("Installation", "Enter number")
            return
        if country not in PAISES:
            messagebox.showwarning("Country", "Invalid code")
            return
        # Keep compare country in sync by default (user can change it)
        try:
            if not self.compare_country_var.get().strip():
                self.compare_country_var.set(country)
        except Exception:
            pass
        self.status_var.set("Downloading...")
        self._start_busy()
        self._set_compare_mode(False)
        self.text.delete('1.0','end')
        threading.Thread(target=self._worker, args=(inst, country, self.save_var.get(), self.view_mode.get()), daemon=True).start()

    def compare_installations(self):
        """Compare currently loaded JSON vs another installation (downloaded)."""
        if self._last_json_data is None:
            messagebox.showinfo("Compare", "No JSON loaded. Query an installation first.")
            return

        inst_b = self.compare_install_var.get().strip()
        country_b = self.compare_country_var.get().strip().upper()
        if not inst_b:
            messagebox.showwarning("Compare", "Enter the installation number to compare against.")
            return
        if country_b not in PAISES:
            messagebox.showwarning("Compare", "Invalid comparison country code")
            return

        inst_a = self.install_var.get().strip() or "(current)"
        country_a = self.country_var.get().strip().upper() or "??"
        base_data = self._last_json_data

        self.status_var.set("Comparing...")
        self._start_busy()
        # Split the UI immediately when Compare is clicked
        self._set_compare_mode(True)
        # Keep last compare state; it will be filled when the worker returns.
        self._last_compare_state = None
        self.text.delete('1.0', 'end')
        try:
            self.compare_text.delete('1.0', 'end')
        except Exception:
            pass
        threading.Thread(
            target=self._worker_compare,
            args=(base_data, inst_a, country_a, inst_b, country_b, self.save_var.get()),
            daemon=True,
        ).start()

    def _set_compare_mode(self, enabled: bool):
        """Enable/disable side-by-side compare view."""
        self._compare_active = bool(enabled)
        if self._compare_active:
            # Compare mode uses the text panel on the left and a dedicated panel on the right
            try:
                self.tree_visible.set(False)
            except Exception:
                pass
            try:
                # IMPORTANT: if Tree View was previously raised, its frame can cover the right panel.
                # Hide it explicitly in Compare mode.
                try:
                    self.tree_frame.grid_remove()
                except Exception:
                    pass
                self.text_frame.grid(row=0, column=0, columnspan=1, sticky='nsew')
                self.compare_right_frame.grid(row=0, column=1, sticky='nsew')
                # Ensure both panes are above any previously raised full-width frames.
                self.compare_right_frame.tkraise()
                self.text_frame.tkraise()
            except Exception:
                pass
            return

        # Normal mode: hide the right compare panel and let main views span the full width
        try:
            # Avoid confusion: diff-only view only makes sense while comparing.
            self.compare_only_diffs_var.set(False)
        except Exception:
            pass
        try:
            self.compare_right_frame.grid_remove()
        except Exception:
            pass
        try:
            self.text_frame.grid(row=0, column=0, columnspan=2, sticky='nsew')
            self.tree_frame.grid(row=0, column=0, columnspan=2, sticky='nsew')
        except Exception:
            pass
        if self.tree_visible.get():
            try:
                self.tree_frame.tkraise()
            except Exception:
                pass
        else:
            try:
                self.text_frame.tkraise()
            except Exception:
                pass

    @staticmethod
    def _safe_value_repr(value: Any, max_len: int = 800) -> str:
        try:
            if isinstance(value, (dict, list)):
                s = json.dumps(value, ensure_ascii=False, sort_keys=True)
            else:
                s = str(value)
        except Exception:
            s = repr(value)
        if len(s) > max_len:
            return s[:max_len] + "… (truncated)"
        return s

    @classmethod
    def _diff_json(cls, a: Any, b: Any, path: str = ""):
        """Return (changed, added, removed) as lists of tuples.

        - changed: (path, old, new)
        - added: (path, value)
        - removed: (path, value)
        """
        changed = []
        added = []
        removed = []

        def join(base: str, key: str) -> str:
            return f"{base}.{key}" if base else key

        def rec(x: Any, y: Any, p: str):
            if isinstance(x, dict) and isinstance(y, dict):
                keys = set(x.keys()) | set(y.keys())
                for k in sorted(keys, key=lambda z: str(z)):
                    kp = join(p, str(k))
                    if k not in x:
                        added.append((kp, y.get(k)))
                        continue
                    if k not in y:
                        removed.append((kp, x.get(k)))
                        continue
                    rec(x.get(k), y.get(k), kp)
                return
            if isinstance(x, list) and isinstance(y, list):
                m = max(len(x), len(y))
                for i in range(m):
                    kp = f"{p}[{i}]" if p else f"[{i}]"
                    if i >= len(x):
                        added.append((kp, y[i]))
                        continue
                    if i >= len(y):
                        removed.append((kp, x[i]))
                        continue
                    rec(x[i], y[i], kp)
                return
            if x != y:
                changed.append((p, x, y))

        rec(a, b, path)
        return changed, added, removed

    class _JsonSpanRenderer:
        """Pretty-print JSON while recording path->(line,start,end) spans for highlighting."""

        def __init__(self, indent: int = 2):
            self.indent = indent
            self.lines: List[str] = []
            self.spans: dict = {}

        def _append(self, line: str):
            self.lines.append(line)

        def _add_span(self, path: str, line_no: int, start: int, end: int):
            if not path:
                return
            self.spans.setdefault(path, []).append((line_no, start, end))

        @staticmethod
        def _is_scalar(v: Any) -> bool:
            return isinstance(v, (str, int, float, bool)) or v is None

        @staticmethod
        def _scalar_repr(v: Any) -> str:
            try:
                return json.dumps(v, ensure_ascii=False)
            except Exception:
                return json.dumps(str(v), ensure_ascii=False)

        def _render_dict_body(self, d: dict, path: str, indent_open: int, comma: bool):
            items = list(d.items())
            for idx, (k, v) in enumerate(items):
                key = str(k)
                key_json = json.dumps(key, ensure_ascii=False)
                child_path = f"{path}.{key}" if path else key
                is_last = (idx == len(items) - 1)
                need_comma = (not is_last)

                prefix = (" " * (indent_open + self.indent)) + key_json + ": "
                line_no = len(self.lines) + 1
                key_start = indent_open + self.indent
                key_end = key_start + len(key_json)

                if self._is_scalar(v):
                    line = prefix + self._scalar_repr(v)
                    if need_comma:
                        line += ","
                    self._append(line)
                    self._add_span(child_path, line_no, key_start, key_end)
                    continue

                if isinstance(v, dict):
                    self._append(prefix + "{")
                    self._add_span(child_path, line_no, key_start, key_end)
                    self._render_dict_body(v, child_path, indent_open + self.indent, comma=need_comma)
                    continue

                if isinstance(v, list):
                    self._append(prefix + "[")
                    self._add_span(child_path, line_no, key_start, key_end)
                    self._render_list_body(v, child_path, indent_open + self.indent, comma=need_comma)
                    continue

                # Fallback for unknown types
                line = prefix + self._scalar_repr(str(v))
                if need_comma:
                    line += ","
                self._append(line)
                self._add_span(child_path, line_no, key_start, key_end)

            closing = (" " * indent_open) + "}"
            if comma:
                closing += ","
            self._append(closing)

        def _render_list_body(self, arr: list, path: str, indent_open: int, comma: bool):
            for i, v in enumerate(arr):
                child_path = f"{path}[{i}]" if path else f"[{i}]"
                is_last = (i == len(arr) - 1)
                need_comma = (not is_last)
                prefix = " " * (indent_open + self.indent)
                line_no = len(self.lines) + 1

                if self._is_scalar(v):
                    line = prefix + self._scalar_repr(v)
                    if need_comma:
                        line += ","
                    self._append(line)
                    self._add_span(child_path, line_no, indent_open + self.indent, len(line))
                    continue

                if isinstance(v, dict):
                    self._append(prefix + "{")
                    self._add_span(child_path, line_no, indent_open + self.indent, indent_open + self.indent + 1)
                    self._render_dict_body(v, child_path, indent_open + self.indent, comma=need_comma)
                    continue

                if isinstance(v, list):
                    self._append(prefix + "[")
                    self._add_span(child_path, line_no, indent_open + self.indent, indent_open + self.indent + 1)
                    self._render_list_body(v, child_path, indent_open + self.indent, comma=need_comma)
                    continue

                line = prefix + self._scalar_repr(str(v))
                if need_comma:
                    line += ","
                self._append(line)
                self._add_span(child_path, line_no, indent_open + self.indent, len(line))

            closing = (" " * indent_open) + "]"
            if comma:
                closing += ","
            self._append(closing)

        def render(self, value: Any):
            self.lines.clear()
            self.spans.clear()
            if isinstance(value, dict):
                self._append("{")
                self._render_dict_body(value, "", 0, comma=False)
                return
            if isinstance(value, list):
                self._append("[")
                self._render_list_body(value, "", 0, comma=False)
                return
            self._append(self._scalar_repr(value))

        def text(self) -> str:
            return "\n".join(self.lines) + "\n"

    @staticmethod
    def _best_span_lookup(spans: dict, path: str):
        """Try path, then parent paths, until a span exists."""
        p = path
        while p:
            if p in spans:
                return spans[p]
            # Strip last list index
            if p.endswith(']') and '[' in p:
                p = p[:p.rfind('[')]
                continue
            # Strip last dict key
            if '.' in p:
                p = p[:p.rfind('.')]
                continue
            break
        return None

    def _render_json_into(self, widget: tk.Text, data: Any) -> dict:
        renderer = self._JsonSpanRenderer(indent=2)
        renderer.render(data)
        widget.delete('1.0', 'end')
        widget.insert('1.0', renderer.text())
        return renderer.spans

    def _apply_diff_highlights(self, widget: tk.Text, spans: dict, paths: set[str]):
        try:
            widget.tag_remove('diff_field', '1.0', 'end')
        except Exception:
            return
        for p in paths:
            span_list = self._best_span_lookup(spans, p)
            if not span_list:
                continue
            for line_no, start, end in span_list:
                try:
                    widget.tag_add('diff_field', f"{line_no}.{start}", f"{line_no}.{end}")
                except Exception:
                    pass

    def _render_compare_diff_only(self, changed, added, removed, inst_a: str, country_a: str, inst_b: str, country_b: str):
        """Render a compact view listing only differing fields into both panes."""
        # Sort for stable output
        changed_sorted = sorted(changed, key=lambda t: str(t[0]))
        added_sorted = sorted(added, key=lambda t: str(t[0]))
        removed_sorted = sorted(removed, key=lambda t: str(t[0]))

        left_lines: List[str] = []
        right_lines: List[str] = []

        header = f"# ONLY DIFFS: {inst_a}({country_a}) vs {inst_b}({country_b})"
        left_lines.append(header)
        right_lines.append(header)
        left_lines.append(f"# changed={len(changed_sorted)} added={len(added_sorted)} removed={len(removed_sorted)}")
        right_lines.append(f"# changed={len(changed_sorted)} added={len(added_sorted)} removed={len(removed_sorted)}")
        left_lines.append("")
        right_lines.append("")

        def add_section(title: str):
            left_lines.append(f"== {title} ==")
            right_lines.append(f"== {title} ==")

        def add_pair(path: str, a_val: Any, b_val: Any):
            left_lines.append(f"{path} = {self._safe_value_repr(a_val)}")
            right_lines.append(f"{path} = {self._safe_value_repr(b_val)}")

        if changed_sorted:
            add_section("CHANGED")
            for p, old, new in changed_sorted:
                add_pair(p, old, new)
            left_lines.append("")
            right_lines.append("")

        if added_sorted:
            add_section("ADDED")
            for p, val in added_sorted:
                add_pair(p, "<missing>", val)
            left_lines.append("")
            right_lines.append("")

        if removed_sorted:
            add_section("REMOVED")
            for p, val in removed_sorted:
                add_pair(p, val, "<missing>")
            left_lines.append("")
            right_lines.append("")

        if not (changed_sorted or added_sorted or removed_sorted):
            left_lines.append("No differences detected.")
            right_lines.append("No differences detected.")

        # Paint
        try:
            self.text.delete('1.0', 'end')
            self.compare_text.delete('1.0', 'end')
        except Exception:
            pass
        self.text.insert('1.0', "\n".join(left_lines) + "\n")
        self.compare_text.insert('1.0', "\n".join(right_lines) + "\n")

        # Make only the diff lines red (not headers)
        try:
            self.text.tag_remove('diff_field', '1.0', 'end')
            self.compare_text.tag_remove('diff_field', '1.0', 'end')
        except Exception:
            return

        def tag_diff_lines(widget: tk.Text, lines: List[str]):
            # Lines are 1-based in Tk Text indices
            for i, line in enumerate(lines, start=1):
                if not line or line.startswith('#') or line.startswith('=='):
                    continue
                if ' = ' in line and not line.startswith('No differences'):
                    try:
                        widget.tag_add('diff_field', f"{i}.0", f"{i}.end")
                    except Exception:
                        pass

        tag_diff_lines(self.text, left_lines)
        tag_diff_lines(self.compare_text, right_lines)

    def _toggle_compare_only_diffs(self):
        """Toggle between full JSON compare view and diff-only view."""
        if not getattr(self, '_compare_active', False):
            try:
                self.compare_only_diffs_var.set(False)
            except Exception:
                pass
            return

        state = getattr(self, '_last_compare_state', None)
        if not state:
            # No compare results yet
            return

        (base_data, data_b, changed, added, removed, inst_a, country_a, inst_b, country_b) = state
        if self.compare_only_diffs_var.get():
            self._render_compare_diff_only(changed, added, removed, inst_a, country_a, inst_b, country_b)
            self.status_var.set(f"Diff-only view: {inst_a}({country_a}) vs {inst_b}({country_b})")
            return

        # Restore full JSON views + highlights
        changed_paths = {p for (p, _, _) in changed if p}
        added_paths = {p for (p, _) in added if p}
        removed_paths = {p for (p, _) in removed if p}
        spans_left = self._render_json_into(self.text, base_data)
        spans_right = self._render_json_into(self.compare_text, data_b)
        self._apply_diff_highlights(self.text, spans_left, set(changed_paths) | set(removed_paths))
        self._apply_diff_highlights(self.compare_text, spans_right, set(changed_paths) | set(added_paths))
        self.status_var.set(f"Comparison ready: {inst_a}({country_a}) vs {inst_b}({country_b})")

    def close_compare_view(self):
        """User action: close side-by-side Compare view and return to normal layout."""
        try:
            self.compare_only_diffs_var.set(False)
        except Exception:
            pass
        self._set_compare_mode(False)
        self.status_var.set("Compare closed")
        self._update_action_states()

    def _worker_compare(self, base_data: Any, inst_a: str, country_a: str, inst_b: str, country_b: str, keep_files: bool):
        try:
            data_b, filename_b = descargar_instalacion(inst_b, country_b)
            # Remove downloaded file if user does not want to save JSONs
            if not keep_files and os.path.isfile(filename_b):
                try:
                    os.remove(filename_b)
                except OSError:
                    pass

            changed, added, removed = self._diff_json(base_data, data_b)
            payload = (base_data, data_b, changed, added, removed, inst_a, country_a, inst_b, country_b)
            self._q.put(("cmpview", payload))
        except Exception as e:
            self._q.put(("error", self._format_user_error(e)))

    def _format_user_error(self, exc: Exception) -> str:
        """Return a short actionable message for common expected errors."""
        msg = str(exc) if exc is not None else "Unknown error"

        if "Offline mode is enabled" in msg or "DCR_OFFLINE" in msg:
            return (
                msg
                + "\n\nTip: unset DCR_OFFLINE (or set it to 0) to download online."
            )

        if isinstance(exc, (RuntimeError, ValueError)):
            return msg

        return f"{msg}\n\n{traceback.format_exc()}"

    def _build_summary(self, data: Any) -> str:
        """Build the same human summary used by `analizar_json.py` (console output), but as a string."""
        # Device type mapping (copied from analizar_json.py)
        device_type_mapping = {
            "101": "MAGNETIC",
            "102": "CAMPIR",
            "104": "ZEROVISION",
            "120": "SPB",
            "121": "SMOKE",
            "122": "WATER",
            "130": "SMARTPLUG",
            "140": "HOMEPANEL",
            "141": "SMARTDOT",
            "142": "SVK",
            "162": "KEYFOB",
            "105": "IPCAMERA",
            "103": "CROPTEX",
            "106": "ORION",
            "107": "AQUILA",
            "109": "AQUILABUSINESS",
            "143": "MOK",
            "108": "NOX",
            "123": "SENTINEL1",
            "124": "SENTINEL2",
            "163": "LOCK",
            "131": "WRE",
            "110": "MC3",
        }

        # In the GUI we can keep accents; keep this as identity.
        def sin_acentos(texto: Any) -> Any:
            return texto

        def get_dict(obj: Any) -> dict:
            return obj if isinstance(obj, dict) else {}

        def get_list(obj: Any) -> list:
            return obj if isinstance(obj, list) else []

        data = get_dict(data)
        lines: List[str] = []

        # Counters and extracted fields
        device_count: dict = {}
        total_devices = 0
        user_count = 0
        global_tag_count = 0
        user_tag_count = 0
        cu_serial_number = "unknown"
        installation_number = "unknown"
        nodes_info: List[str] = []

        cu = get_dict(data.get('cu'))
        if cu:
            cu_serial_number = cu.get('serialNumber', 'unknown')

        platform = get_dict(data.get('platform'))
        if platform:
            installation_number = platform.get('installationNumber', 'unknown')

        nodes = get_list(data.get('nodes'))
        for device in nodes:
            if not isinstance(device, dict):
                continue
            total_devices += 1
            device_type_code = str(device.get('type', ''))
            device_type_name = device_type_mapping.get(device_type_code, "UNKNOWN")
            serial_number = device.get('serialNumber', 'unknown')
            zone_id = device.get('zoneId', 'unknown')
            mac = None
            if device_type_name in ["ORION", "AQUILA"]:
                config = get_dict(device.get('config'))
                mac = config.get('mac')
            node_str = f"{device_type_name}: {serial_number}, Zone ID: {zone_id}"
            if mac:
                node_str += f", MAC: {mac}"
            nodes_info.append(node_str)
            device_count[device_type_name] = device_count.get(device_type_name, 0) + 1

        users = get_list(data.get('users'))
        user_count = len(users)
        for user in users:
            if isinstance(user, dict) and isinstance(user.get('tags'), list):
                user_tag_count += len(user.get('tags'))

        tags = get_list(data.get('tags'))
        global_tag_count = len(tags)

        # Header summary (matches original style)
        lines.append(sin_acentos("============================================"))
        lines.append(sin_acentos("           INSTALLATION SUMMARY              "))
        lines.append(sin_acentos("============================================"))
        lines.append(sin_acentos(f"\nInstallation number  : {installation_number}"))
        lines.append(sin_acentos(f"Serial Number         : {cu_serial_number}"))
        lines.append(sin_acentos(f"Created               : {data.get('createdAt', 'N/A')}"))
        lines.append(sin_acentos(f"Last updated          : {data.get('lastUpdatedDate', 'N/A')}"))
        lines.append(sin_acentos("\n--------------------------------------------"))
        lines.append(sin_acentos(f"Total devices         : {total_devices}"))
        lines.append(sin_acentos("Count by type:"))
        for device_type in sorted(device_count.keys()):
            lines.append(sin_acentos(f"   - {device_type:12}: {device_count[device_type]}"))

        lines.append(sin_acentos("\nNodes:"))
        for node_info in nodes_info:
            lines.append(sin_acentos(f"   • {node_info}"))

        lines.append(sin_acentos(f"\nTotal users           : {user_count}"))
        lines.append(sin_acentos(f"Total tags            : {global_tag_count}"))

        # Extended info
        lines.append(sin_acentos("\n============================================"))
        lines.append(sin_acentos("         EXTENDED INFORMATION                "))
        lines.append(sin_acentos("============================================"))

        lines.append(sin_acentos("\nCU:"))
        lines.append(sin_acentos(f"   Label      : {cu.get('label', 'N/A')}"))
        lines.append(sin_acentos(f"   Location   : {cu.get('location', 'N/A')}"))
        lines.append(sin_acentos(f"   Alias      : {cu.get('alias', 'N/A')}"))
        lines.append(sin_acentos(f"   SSID       : {cu.get('ssid', 'N/A')}"))
        feature_flags = get_dict(cu.get('featureFlags'))
        lines.append(sin_acentos("   Feature Flags:"))
        for flag, value in feature_flags.items():
            lines.append(sin_acentos(f"      - {str(flag):12}: {value}"))
        voip = get_dict(cu.get('voip'))
        lines.append(sin_acentos(f"   VoIP SIM   : {get_dict(voip).get('simNumber', 'N/A')}"))

        # Users
        try:
            if users:
                lines.append(sin_acentos("\nUsers:"))
                for user in users:
                    if not isinstance(user, dict):
                        continue
                    lines.append(sin_acentos(
                        f"   • ID: {user.get('id')} | Label: {user.get('label')} | Owner: {user.get('isOwner')} | Admin: {user.get('isAdmin')} | TagID: {user.get('tagId')}"
                    ))
        except Exception as e:
            lines.append(sin_acentos(f"Unexpected error while processing users: {e}"))

        # Tags
        if tags:
            lines.append(sin_acentos("\nTags:"))
            for tag in tags:
                if not isinstance(tag, dict):
                    continue
                lines.append(sin_acentos(
                    f"   • ID: {tag.get('id')} | Label: {tag.get('label')} | Color: {tag.get('color')} | Active: {tag.get('isActive')}"
                ))

        # Alarm partitions
        alarm_partitions = get_list(data.get('alarmPartitions'))
        if alarm_partitions:
            lines.append(sin_acentos("\nAlarm partitions:"))
            for part in alarm_partitions:
                if not isinstance(part, dict):
                    continue
                lines.append(sin_acentos(
                    f"   • ID: {part.get('id')} | Name: {part.get('name')} | Associated nodes: {part.get('associatedNodes')}"
                ))
                arm_modes = get_list(part.get('armModes'))
                for mode in arm_modes:
                    if not isinstance(mode, dict):
                        continue
                    lines.append(sin_acentos(
                        f"      - Mode: {mode.get('label')} | Entry: {mode.get('entryTime')}s | Exit: {mode.get('exitTime')}s"
                    ))

        # Providers
        providers = get_list(data.get('providers'))
        if providers:
            lines.append(sin_acentos("\nNetwork providers:"))
            for prov in providers:
                if not isinstance(prov, dict):
                    continue
                mcc = prov.get('mcc', 'N/A')
                if mcc != 'N/A':
                    lines.append(sin_acentos(
                        f"   • Name: {prov.get('name', 'N/A')} | Type: {prov.get('type', 'N/A')} | MCC: {mcc}"
                    ))

        # Endpoints
        lines.append(sin_acentos("\n============================================"))
        lines.append(sin_acentos("            RELEVANT ENDPOINTS               "))
        lines.append(sin_acentos("============================================"))
        endpoints = data.get('endpoints')
        if isinstance(endpoints, dict):
            for key in [
                'dcr', 'keyMaster', 'elk', 'orquestadorInst', 'orquestadorServ',
                'fotaBackend', 'voipRegister', 'keyVault', 'debugFiles'
            ]:
                if key in endpoints:
                    val = endpoints.get(key)
                    url = val.get('url') if isinstance(val, dict) else val
                    lines.append(sin_acentos(f"   • {key:16}: {url}"))

        return "\n".join(str(x) for x in lines)

    # (Old duplicate version of buscar_en_lista removed)
    def _worker(self, inst, country, guardar, modo):
        try:
            data, filename = descargar_instalacion(inst, country)
            contenido = json.dumps(data, indent=2, ensure_ascii=False) if modo=='json' else self._build_summary(data)
            # Always keep the raw JSON for path extraction even if only the summary is shown
            if not guardar and os.path.isfile(filename):
                try: os.remove(filename)
                except OSError: pass
            if modo=='json':
                self._q.put(("okjson", contenido, filename if guardar else None, data))
            else:
                self._q.put(("ok", contenido, filename if guardar else None, data))
        except Exception as e:
            self._q.put(("error", self._format_user_error(e)))
    # Auto export removed

    def _poll_queue(self):
        try:
            while True:
                estado, *rest = self._q.get_nowait()
                if estado=='ok':
                    self._set_compare_mode(False)
                    contenido, filename, data_obj = rest
                    self.text.insert('end', contenido)
                    self.status_var.set("Summary" + (f" saved to {filename}" if filename else " loaded"))
                    # Mantener JSON para extracción de rutas y filtrados posteriores
                    self._last_json_data = data_obj
                    self._stop_busy()
                    self._update_action_states()
                elif estado=='okjson':
                    self._set_compare_mode(False)
                    contenido, filename, data_obj = rest
                    self.text.insert('end', contenido)
                    self.status_var.set("JSON" + (f" saved to {filename}" if filename else " loaded"))
                    self._last_json_data = data_obj
                    if self.tree_visible.get(): self._cargar_tree(data_obj)
                    self._stop_busy()
                    self._update_action_states()
                elif estado=='csvprog':
                    self._set_compare_mode(False)
                    linea, stat = rest
                    self.text.insert('end', linea+"\n"); self.text.see('end'); self.status_var.set(stat)
                elif estado=='csvok':
                    self._set_compare_mode(False)
                    contenido, stats = rest
                    self.text.insert('end', contenido); self.status_var.set(stats)
                elif estado=='jsonforce':
                    self._set_compare_mode(False)
                    filename = rest[0]; self.status_var.set(f"JSON saved to {filename}"); self._stop_busy()
                elif estado=='batchsearchprog':
                    self._set_compare_mode(False)
                    linea, stat = rest
                    self.text.insert('end', linea+"\n"); self.text.see('end'); self.status_var.set(stat)
                elif estado=='batchsearchdone':
                    self._set_compare_mode(False)
                    resumen = rest[0]
                    self.text.insert('end',"\n"+resumen+"\n")
                    self.status_var.set("Batch search completed")
                    self._set_batch_running(False)
                    self._stop_busy()
                elif estado=='batchcancelled':
                    self._set_compare_mode(False)
                    resumen = rest[0]
                    self.text.insert('end',"\n"+resumen+"\n")
                    self.status_var.set("Batch cancelled")
                    self._set_batch_running(False)
                    self._stop_busy()
                elif estado=='cmpview':
                    (base_data, data_b, changed, added, removed, inst_a, country_a, inst_b, country_b) = rest[0]
                    self._set_compare_mode(True)
                    # Persist last compare info so user can toggle diff-only view without re-downloading.
                    self._last_compare_state = (base_data, data_b, changed, added, removed, inst_a, country_a, inst_b, country_b)

                    if self.compare_only_diffs_var.get():
                        self._render_compare_diff_only(changed, added, removed, inst_a, country_a, inst_b, country_b)
                        self.status_var.set(f"Diff-only view: {inst_a}({country_a}) vs {inst_b}({country_b})")
                    else:
                        spans_left = self._render_json_into(self.text, base_data)
                        spans_right = self._render_json_into(self.compare_text, data_b)
                        # Highlight any differing paths in red
                        changed_paths = {p for (p, _, _) in changed if p}
                        added_paths = {p for (p, _) in added if p}
                        removed_paths = {p for (p, _) in removed if p}
                        self._apply_diff_highlights(self.text, spans_left, set(changed_paths) | set(removed_paths))
                        self._apply_diff_highlights(self.compare_text, spans_right, set(changed_paths) | set(added_paths))
                        self.status_var.set(f"Comparison ready: {inst_a}({country_a}) vs {inst_b}({country_b})")
                    self._stop_busy()
                    self._update_action_states()
                elif estado=='error':
                    err = rest[0]; self.status_var.set("Error"); self._stop_busy(); messagebox.showerror("Error", err)
        except queue.Empty:
            pass
        finally:
            self.root.after(300, self._poll_queue)

    def limpiar(self):
        self._set_compare_mode(False)
        self.text.delete('1.0','end')
        try:
            self.compare_text.delete('1.0', 'end')
        except Exception:
            pass
        # Also reset search state so Clear is a true full reset
        try:
            self.text.tag_remove('search_match','1.0','end')
            self.text.tag_remove('current_match','1.0','end')
        except Exception:
            pass
        try:
            self._search_results = []
            self._search_pos = -1
        except Exception:
            pass
        try:
            self._tree_limpiar_busqueda()
        except Exception:
            pass
        try:
            self._original_text = None
        except Exception:
            pass
        try:
            self.search_var.set('')
        except Exception:
            pass
        self.status_var.set("Ready")
        self._last_json_data = None
        self._last_compare_state = None
        try:
            self.compare_only_diffs_var.set(False)
        except Exception:
            pass
        self._update_action_states()

    def copiar_texto(self):
        # Copy what the user is effectively viewing.
        data = ""
        try:
            in_compare = bool(getattr(self, '_compare_active', False))
        except Exception:
            in_compare = False

        if in_compare:
            left = self.text.get('1.0', 'end-1c').rstrip()
            right = self.compare_text.get('1.0', 'end-1c').rstrip()
            if not left.strip() and not right.strip():
                messagebox.showinfo("Copy", "Nothing to copy")
                return
            label_left = "LEFT"
            label_right = "RIGHT"
            try:
                if getattr(self, '_last_compare_state', None):
                    (_, _, _, _, _, inst_a, country_a, inst_b, country_b) = self._last_compare_state
                    label_left = f"{inst_a}({country_a})"
                    label_right = f"{inst_b}({country_b})"
            except Exception:
                pass
            data = (
                f"### COMPARE COPY\n"
                f"### {label_left}\n{left}\n\n"
                f"### {label_right}\n{right}\n"
            )
        else:
            try:
                if self.tree_visible.get():
                    if self._last_json_data is None:
                        messagebox.showinfo("Copy", "No JSON loaded")
                        return
                    data = json.dumps(self._last_json_data, indent=2, ensure_ascii=False)
                else:
                    data = self.text.get('1.0', 'end-1c')
            except Exception:
                data = self.text.get('1.0', 'end-1c')

        if not data.strip():
            messagebox.showinfo("Copy", "Nothing to copy")
            return
        longitud = len(data)
        # Direct copy threshold: ~200k chars (~200 KB)
        UMBRAL = 200_000
        if longitud <= UMBRAL:
            try:
                self.root.clipboard_clear()
                self.root.clipboard_append(data)
                try:
                    # Helps clipboard persistence on some systems
                    self.root.update_idletasks()
                except Exception:
                    pass
                # Solo notificar en tamaños moderados para evitar retrasos de UI
                if longitud < 20_000:
                    messagebox.showinfo("Copy", f"Copied ({longitud} chars)")
                else:
                    self.status_var.set(f"Copied to clipboard ({longitud} chars)")
            except Exception as e:
                messagebox.showerror("Copy", f"Error copying: {e}")
            return
        # For large outputs, offer saving to file instead of copying
        resp = messagebox.askyesno("Copy", f"Output is very large ({longitud} chars). Save to a file instead of copying?")
        if resp:
            ruta = filedialog.asksaveasfilename(title="Save output", defaultextension=".txt", filetypes=[("Text","*.txt"),("All files","*.*")])
            if not ruta:
                self.status_var.set("Save canceled")
                return
            try:
                with open(ruta,'w',encoding='utf-8') as f:
                    f.write(data)
                self.status_var.set(f"Output saved to {os.path.basename(ruta)}")
            except Exception as e:
                messagebox.showerror("Save", f"Error saving: {e}")
                # Auto export removed
        else:
            # Partial copy if the user insists on copying
            fragmento = data[:UMBRAL]
            try:
                self.root.clipboard_clear(); self.root.clipboard_append(fragmento)
                self.status_var.set(f"Partial copy ({UMBRAL} chars of {longitud})")
            except Exception as e:
                messagebox.showerror("Copy", f"Error copying partially: {e}")

    # (Copiar rápido eliminado; lógica integrada en copiar_texto)

    def forzar_descarga_json(self):
        inst = self.install_var.get().strip(); country = self.country_var.get().strip().upper()
        if not inst or country not in PAISES:
            messagebox.showwarning("Download","Invalid installation or country"); return
        def _dl():
            try:
                data, filename = descargar_instalacion(inst, country)
                # Show in panel according to the current mode and refresh the tree
                contenido = json.dumps(data, indent=2, ensure_ascii=False) if self.view_mode.get()=="json" else self._build_summary(data)
                # Publish content first to refresh UI and _last_json_data
                if self.view_mode.get()=="json":
                    self._q.put(("okjson", contenido, filename, data))
                else:
                    self._q.put(("ok", contenido, filename, data))
                # Then explicitly update the save state
                self._q.put(("jsonforce", filename))
            except Exception as e:
                self._q.put(("error", str(e)))
        self.status_var.set("Downloading raw JSON..."); self._start_busy(); threading.Thread(target=_dl, daemon=True).start()

    def cargar_json_local(self):
        ruta = filedialog.askopenfilename(title="Open JSON", filetypes=[("JSON","*.json"),("All files","*.*")])
        if not ruta:
            return
        try:
            with open(ruta,'r',encoding='utf-8') as f:
                data = json.load(f)
        except Exception as e:
            messagebox.showerror("Open JSON", f"Error loading file: {e}")
            return
        self._last_json_data = data
        self.text.delete('1.0','end')
        if self.view_mode.get()=="resumen":
            try:
                contenido = self._build_summary(data)
            except Exception:
                contenido = json.dumps(data, indent=2, ensure_ascii=False)
        else:
            contenido = json.dumps(data, indent=2, ensure_ascii=False)
        self.text.insert('end', contenido)
        if self.tree_visible.get():
            try:
                self._cargar_tree(data)
            except Exception:
                pass
        self.status_var.set(f"JSON loaded: {os.path.basename(ruta)}")
        self._update_action_states()

    def _open_json_folder(self):
        """Open the folder where downloaded JSON files are stored."""
        try:
            download_dir = os.getenv("ULTRAJSON_DOWNLOAD_DIR", "").strip()
        except Exception:
            download_dir = ""
        if not download_dir:
            download_dir = os.path.join(tempfile.gettempdir(), "Ultrajson")

        try:
            os.makedirs(download_dir, exist_ok=True)
        except Exception:
            pass

        try:
            os.startfile(download_dir)
        except Exception as e:
            messagebox.showerror("Open JSON", f"Error opening folder:\n{download_dir}\n\n{e}")

    # === Menú y utilidades ===
    def _build_menu(self):
        menubar = tk.Menu(self.root)
        # File
        menu_file = tk.Menu(menubar, tearoff=0)
        menu_file.add_command(label="Open JSON (Ctrl+O)", command=self.cargar_json_local)
        # Guardar referencias para poder habilitar/deshabilitar esta acción
        menu_file.add_command(label="Download JSON", command=self.forzar_descarga_json)
        # Índice actual del último ítem (Download JSON)
        try:
            self._menu_file = menu_file
            self._menu_download_idx = menu_file.index("end")
        except Exception:
            self._menu_file = None
            self._menu_download_idx = None
        menu_file.add_separator()
        menu_file.add_command(label="Exit", command=self.root.destroy)
        menubar.add_cascade(label="File", menu=menu_file)
        # Edit
        menu_edit = tk.Menu(menubar, tearoff=0)
        menu_edit.add_command(label="Copy", command=self.copiar_texto)
        menu_edit.add_command(label="Clear Panel", command=self.limpiar)
        menu_edit.add_separator()
        menu_edit.add_command(label="Search (Ctrl+F)", command=self.buscar)
        menu_edit.add_command(label="Next (F3)", command=self.buscar_siguiente)
        menu_edit.add_command(label="Previous (Shift+F3)", command=self.buscar_anterior)
        menubar.add_cascade(label="Edit", menu=menu_edit)
        # View
        menu_view = tk.Menu(menubar, tearoff=0)
        menu_view.add_checkbutton(label="Tree View", variable=self.tree_visible, command=self.toggle_tree_view)
        if sv_ttk:
            theme_menu = tk.Menu(menu_view, tearoff=0)
            theme_menu.add_radiobutton(label="Light", value="light", variable=self.theme_var, command=lambda: self._apply_theme("light"))
            theme_menu.add_radiobutton(label="Dark", value="dark", variable=self.theme_var, command=lambda: self._apply_theme("dark"))
            menu_view.add_cascade(label="Theme", menu=theme_menu)
        menubar.add_cascade(label="View", menu=menu_view)
        # Tools
        menu_tools = tk.Menu(menubar, tearoff=0)
        menu_tools.add_command(label="Filter CSV (Batch)", command=self.buscar_en_lista)
        menu_tools.add_command(label="Show Paths", command=self.mostrar_rutas)
        menu_tools.add_command(label="Extract Path Value", command=self.extraer_valor_ruta_gui)
        menu_tools.add_command(label="Remote Edit (APIM)…", command=self.open_remote_edit_apim)
        menubar.add_cascade(label="Tools", menu=menu_tools)
        # Help
        menu_help = tk.Menu(menubar, tearoff=0)
        menu_help.add_command(label="Quick Start", command=self._show_quick_start)
        menu_help.add_command(label="Keyboard Shortcuts", command=self._show_shortcuts)
        menu_help.add_command(label="Open Ultrajson README", command=self._open_ultrajson_readme)
        menubar.add_cascade(label="Help", menu=menu_help)
        self.root.config(menu=menubar)

    def _show_quick_start(self):
        text = (
            "Quick Start\n\n"
            "1) Query: enter Installation + Country → Query\n"
            "2) Open JSON: File → Open JSON (Ctrl+O)\n"
            "3) Search highlights; Filter reduces output\n"
            "4) Tree View: browse JSON structure\n"
            "5) Compare: use 'Compare to' → Compare\n"
            "   - Only diffs: show only changed/added/removed fields\n"
            "   - Close Compare: return to normal view\n\n"
            "Path examples: cu.serialNumber | nodes[0].serialNumber | behaviours[id=autolock].config.timeout"
        )
        messagebox.showinfo("Quick Start", text)

    def _show_shortcuts(self):
        text = (
            "Keyboard Shortcuts\n\n"
            "Ctrl+O  Open JSON\n"
            "Ctrl+F  Focus search box\n"
            "F3      Next match\n"
            "Shift+F3 Previous match\n"
            "Esc     Clear search highlights\n\n"
            "Compare mode:\n"
            "Mouse wheel scrolls both panes\n"
            "Shift + wheel scrolls horizontally"
        )
        messagebox.showinfo("Shortcuts", text)

    def _open_ultrajson_readme(self):
        try:
            here = os.path.dirname(os.path.abspath(__file__))
            readme = os.path.join(here, 'README.md')
            if os.path.isfile(readme):
                os.startfile(readme)
                return
            messagebox.showinfo("README", "README.md not found in Ultrajson folder")
        except Exception as e:
            messagebox.showerror("README", f"Error opening README: {e}")

    def _apply_theme(self, which: str):
        try:
            if not sv_ttk:
                return
            if which == 'light':
                sv_ttk.set_theme('light')
            else:
                sv_ttk.set_theme('dark')
        except Exception:
            pass

    def _focus_search(self):
        try:
            self.ent_search.focus_set()
            self.ent_search.select_range(0, 'end')
        except Exception:
            pass

    def _start_busy(self):
        try:
            if not self.progress.winfo_ismapped():
                self.progress.pack(side='right')
            self.progress.start(12)
        except Exception:
            pass

    def _stop_busy(self):
        try:
            self.progress.stop()
            self.progress.configure(value=0)
            if self.progress.winfo_ismapped():
                self.progress.pack_forget()
        except Exception:
            pass
        self._update_action_states()

    def _set_batch_running(self, running: bool):
        self._batch_running = bool(running)
        try:
            if self.btn_stop is not None:
                self.btn_stop.configure(state=('normal' if self._batch_running else 'disabled'))
        except Exception:
            pass
        if not self._batch_running:
            try:
                self._batch_cancel_event.clear()
            except Exception:
                pass

    def stop_batch(self):
        """Request cancellation of the current Filter CSV batch run."""
        if not getattr(self, '_batch_running', False):
            return
        try:
            self._batch_cancel_event.set()
            self.status_var.set("Stopping batch...")
        except Exception:
            pass
        # Disable immediately to prevent repeated clicks; worker will exit ASAP.
        try:
            if self.btn_stop is not None:
                self.btn_stop.configure(state='disabled')
        except Exception:
            pass

    def _update_download_button(self):
        """Hide 'Save JSON' when redundant.

        If "Save JSON" is enabled, "Query" already saves the JSON (both in Summary and JSON views),
        so this button becomes a duplicate. When "Save JSON" is disabled, keep it available as a
        one-off "save now" action.
        """
        try:
            save_enabled = bool(self.save_var.get())
            redundant = save_enabled
            # Always keep a clear label
            if hasattr(self, 'btn_download') and self.btn_download:
                self.btn_download.configure(text='Save JSON')
                if redundant:
                    # Query already saves (in any view) when Save JSON is enabled
                    if self.btn_download.winfo_ismapped():
                        self.btn_download.pack_forget()
                else:
                    if not self.btn_download.winfo_ismapped():
                        # Reinsert always in the same position (before Open JSON)
                        before_widget = getattr(self, 'btn_open_json', None)
                        if before_widget is not None:
                            self.btn_download.pack(side='left', before=before_widget)
                        else:
                            self.btn_download.pack(side='left')
            # Update File menu action state in sync
            if getattr(self, '_menu_file', None) is not None and getattr(self, '_menu_download_idx', None) is not None:
                try:
                    state = 'disabled' if redundant else 'normal'
                    self._menu_file.entryconfig(self._menu_download_idx, state=state)
                except Exception:
                    pass
        except Exception:
            pass

    # === Búsqueda texto ===
    def buscar(self):
        patron = self.search_var.get().strip()
        patrones = [p.strip() for p in patron.split(',') if p.strip()]
        if self.tree_visible.get():
            self._tree_buscar(patron)
            return
        # No eliminar los tags previos, así se acumulan los resultados
        if not hasattr(self, '_search_results') or not isinstance(self._search_results, list):
            self._search_results = []
        if not hasattr(self, '_search_pos'):
            self._search_pos = -1
        if not patrones:
            return
        self._search_results = buscar_patrones_en_texto(self.text, patrones)
        if not self._search_results:
            self.status_var.set("No matches found")
            return
        self._search_pos = 0
        self._marcar_actual()
        self.status_var.set(f"{len(self._search_results)} matches (accumulated)")

    def _marcar_actual(self):
        self.text.tag_remove('current_match','1.0','end')
        if self._search_pos<0 or self._search_pos>=len(self._search_results): return
        ini,fin = self._search_results[self._search_pos]; self.text.tag_add('current_match',ini,fin); self.text.see(ini); self.text.mark_set('insert',ini)

    def buscar_siguiente(self):
        if self.tree_visible.get():
            if not self._tree_search_results: self.buscar(); return
            self._tree_search_pos = (self._tree_search_pos+1) % len(self._tree_search_results)
            self._tree_marcar_actual()
            self.status_var.set(f"Node {self._tree_search_pos+1}/{len(self._tree_search_results)}")
            return
        if not self._search_results: self.buscar(); return
        self._search_pos = (self._search_pos+1)%len(self._search_results); self._marcar_actual(); self.status_var.set(f"Match {self._search_pos+1}/{len(self._search_results)}")
    def buscar_anterior(self):
        if self.tree_visible.get():
            if not self._tree_search_results: self.buscar(); return
            self._tree_search_pos = (self._tree_search_pos-1) % len(self._tree_search_results)
            self._tree_marcar_actual()
            self.status_var.set(f"Node {self._tree_search_pos+1}/{len(self._tree_search_results)}")
            return
        if not self._search_results: self.buscar(); return
        self._search_pos = (self._search_pos-1)%len(self._search_results); self._marcar_actual(); self.status_var.set(f"Match {self._search_pos+1}/{len(self._search_results)}")
    def limpiar_busqueda(self):
        # Limpiar búsqueda y restaurar contenido original si fue reemplazado por un filtrado / rutas / extracción
        if self.tree_visible.get():
            self._tree_limpiar_busqueda()
            self.search_var.set('')  # también limpiar campo de entrada
            self.status_var.set("Ready")
            return
        # Modo texto
        self.text.tag_remove('search_match','1.0','end')
        self.text.tag_remove('current_match','1.0','end')
        self._search_results=[]
        self._search_pos=-1
        # Restaurar snapshot si existe
        if self._original_text is not None:
            self.text.delete('1.0','end')
            self.text.insert('end', self._original_text)
            self._original_text = None
        self.search_var.set('')
        self.status_var.set("Ready")

    # === Filtrado por ruta / contenido ===
    def filtrar_resultados(self):
        patron = self.search_var.get().strip()
        if not patron:
            self.status_var.set("Empty pattern")
            return
        # Guardar snapshot inicial del panel si todavía no se ha guardado
        if self._original_text is None:
            try:
                self._original_text = self.text.get('1.0','end')
            except Exception:
                self._original_text = None
        patrones = [p.strip() for p in patron.split(',') if p.strip()]
        # Modo árbol: aplicar enfoque de ruta o filtro estructural
        if self.tree_visible.get():
            if self._last_json_data is None:
                self.status_var.set("No JSON loaded")
                return
            if self._es_patron_ruta(patron):
                if self._tree_focus_path(patron):
                    self.status_var.set("Path focused")
                    return
                self.status_var.set("Path not found. Applying partial filter...")
            self._tree_apply_filter(patron)
            return
        # Modo texto
        if self._last_json_data is None:
            contenido = self.text.get('1.0','end').splitlines()
            if not patrones:
                self.status_var.set("Empty pattern")
                return
            resultados = filtrar_lineas_por_patrones(contenido, patrones)
            self.text.tag_remove('search_match','1.0','end'); self.text.tag_remove('current_match','1.0','end')
            self.text.delete('1.0','end')
            if resultados:
                self.text.insert('end', f"# {len(resultados)} lines containing '{patron}'\n\n" + "\n".join(resultados))
                self.status_var.set(f"{len(resultados)} lines filtered")
            else:
                self.text.insert('end', f"No matches for: {patron}")
                self.status_var.set("No matches")
            return
        data = self._last_json_data
        # Multi-ruta separada por comas: delega a la lógica completa de Extract Path Value
        _rutas_split = [r.strip() for r in patron.split(',') if r.strip()]
        if len(_rutas_split) > 1:
            self.extraer_valor_ruta_gui()
            self.status_var.set(f"Multi-path: {len(_rutas_split)} paths")
            return
        if self._es_patron_ruta(patron):
            valor = self.extraer_valor_ruta(data, patron)
            if valor is not None:
                self.text.delete('1.0','end')
                if isinstance(valor, (dict, list)):
                    self.text.insert('end', f"# Exact path: {patron}\n" + json.dumps(valor, indent=2, ensure_ascii=False))
                else:
                    self.text.insert('end', f"Path: {patron}\nValue: {valor}")
                self.status_var.set("Path value")
                return
            self.status_var.set("Path not found. Searching matches...")
        resultados = self._resolve_json_path(data, patron)
        if not resultados:
            self.text.delete('1.0','end')
            self.text.insert('end', f"No matches for: {patron}")
            self.status_var.set("No matches")
            return
        self.text.delete('1.0','end')
        self.text.insert('end', f"# {len(resultados)} matches for '{patron}'\n\n")
        for bloque in resultados:
            self.text.insert('end', bloque + "\n\n")
        self.status_var.set(f"{len(resultados)} matches")
    def _es_patron_ruta(self, patron: str) -> bool:
        return ('[' in patron and ']' in patron) or ('.' in patron)

    def _resolve_json_path(self, data: object, ruta: str):
        MAX = 100
        resultados = []
        patron_lower = ruta.lower().strip()
        vistos = set()
        def agregar(ruta_k, k, v):
            if ruta_k in vistos:
                return
            vistos.add(ruta_k)
            resultados.append(f"### Path: {ruta_k}\n" + json.dumps({k: v}, indent=2, ensure_ascii=False))
        def recorrer(nodo, ruta):
            if len(resultados) >= MAX:
                return
            if isinstance(nodo, list):
                for idx, el in enumerate(nodo):
                    ruta_el = f"{ruta}[{idx}]" if ruta else f"[{idx}]"
                    if isinstance(el, dict) and 'id' in el:
                        idv = str(el.get('id'))
                        if patron_lower in idv.lower():
                            resultados.append(f"### Path: {ruta}[id={idv}]\n" + json.dumps(el, indent=2, ensure_ascii=False))
                            continue
                    recorrer(el, ruta_el)
                return
            if isinstance(nodo, dict):
                for k, v in nodo.items():
                    ruta_k = f"{ruta}.{k}" if ruta else k
                    clave_low = k.lower()
                    # Coincidencia parcial o total en la clave
                    if patron_lower and patron_lower in clave_low:
                        agregar(ruta_k, k, v)
                        if len(resultados) >= MAX: return
                    # Coincidencia en valor escalar
                    if isinstance(v, (str, int, float, bool)) and patron_lower in str(v).lower():
                        agregar(ruta_k, k, v)
                        if len(resultados) >= MAX: return
                    # Recursión para estructuras
                    if isinstance(v, (dict, list)):
                        recorrer(v, ruta_k)
        try:
            recorrer(data, '')
        except Exception as e:
            return []
        return resultados

    # === CSV / Excel batch (filtrar lista) ===
    def buscar_en_lista(self):
        ruta = filedialog.askopenfilename(
            title="List (CSV/XLS/XLSX) with installation + country columns",
            filetypes=[("CSV","*.csv"),("Text","*.txt"),("Excel","*.xls;*.xlsx"),("All files","*.*")]
        )
        if not ruta: return
        patron = self.search_var.get().strip();
        if not patron:
            messagebox.showinfo(
                "Filter CSV",
                "File selected, but Search is empty.\n\n"
                "Enter a search pattern or a JSON path in the Search field (e.g. 'timeout' or 'cu.voip.simNumber') and click Filter CSV again."
            )
            return
        self.text.delete('1.0','end'); self.status_var.set("Batch searching..."); self._start_busy()
        # Batch-to-CSV streaming removed
        try:
            self._batch_cancel_event.clear()
        except Exception:
            pass
        self._set_batch_running(True)
        threading.Thread(target=self._worker_batch_search, args=(ruta, patron, self._es_patron_ruta(patron)), daemon=True).start()

    def _cargar_filas(self, ruta):
        filas = []
        low = ruta.lower()
        try:
            if low.endswith(('.xlsx', '.xls')):
                if low.endswith('.xlsx'):
                    import openpyxl
                    wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
                    ws = wb.active
                    for row in ws.iter_rows(values_only=True):
                        filas.append(["" if c is None else str(c).strip() for c in row])
                else:
                    try:
                        import xlrd
                        wb = xlrd.open_workbook(ruta)
                        sh = wb.sheet_by_index(0)
                        for rx in range(sh.nrows):
                            filas.append([str(sh.cell_value(rx, cx)).strip() for cx in range(sh.ncols)])
                    except Exception as e:
                        self._q.put(("error", f"XLS reading requires xlrd<=1.2.0 installed: {e}"))
                        return filas
            else:
                # Detect delimiter for CSV/TSV (be flexible: ';', ',', '\t')
                try:
                    with open(ruta, 'r', encoding='utf-8', errors='ignore') as f:
                        muestra = f.read(4096)
                    try:
                        dialect = csv.Sniffer().sniff(muestra)
                        delim = dialect.delimiter
                    except Exception:
                        # Heuristic fallback: pick the most frequent delimiter in the sample.
                        candidatos = [';', ',', '\t']
                        conteos = {c: muestra.count(c) for c in candidatos}
                        delim = max(conteos, key=conteos.get) if any(conteos.values()) else ','
                except Exception:
                    delim = ','
                with open(ruta, 'r', encoding='utf-8', errors='ignore') as f:
                    reader = csv.reader(f, delimiter=delim)
                    for row in reader:
                        # Ignorar filas vacías o incompletas
                        row = [str(cell).replace('\ufeff','').replace('\u00a0',' ').strip() for cell in row]
                        if not row or len(row) < 2 or all(not cell for cell in row):
                            continue
                        filas.append(row)
        except Exception as e:
            self._q.put(("error", f"Error reading list: {e}"))
            return []
        return filas

    def _worker_batch_search(self, ruta: str, patron: str, es_ruta: bool):
        filas=self._cargar_filas(ruta)
        # Soporte multi patrón separados por coma. Ej: "cu.featureFlags.VoLTE,cu.featureFlags.voip".
        def _batch_pattern_alias(p: str) -> str:
            s = (p or "").strip()
            low = s.lower()
            # Convenience aliases (common user intent)
            if low in ("installation", "install", "instalacion", "instalación", "id_install", "idinstall", "installation_id"):
                return "sd.installation.number"
            if low in ("country", "pais", "país", "country_iso", "country_iso_code", "iso", "iso_code"):
                return "geo.country_iso_code"
            return s

        patron = _batch_pattern_alias(patron)
        patrones=[_batch_pattern_alias(p.strip()) for p in patron.split(',') if p.strip()]
        if not patrones:
            patrones=[]
        patron_lower=patron.lower()
        total=ok=errores=rutas_inval=matches=0
        header_processed=False
        idx_country=None
        idx_install=None
        # Batch-to-CSV streaming removed
        def extraer_timeout_autolock(data):
            try:
                for b in data.get('behaviours',[]):
                    if isinstance(b,dict) and str(b.get('id','')).lower()=='autolock':
                        cfg=b.get('config',{}); return cfg.get('timeout') if isinstance(cfg,dict) else None
            except Exception as e:
                pass
            return None
        def buscar_patron(data):
            resultados=[]
            # Caso multi-ruta: todos los patrones tienen pinta de ruta
            if patrones and all(self._es_patron_ruta(p) for p in patrones):
                for ruta_pat in patrones:
                    valor=self.extraer_valor_ruta(data, ruta_pat)
                    if valor is not None:
                        resultados.append((ruta_pat, valor))
                return resultados
            # Caso una sola ruta
            if es_ruta and len(patrones)==1:
                valor = self.extraer_valor_ruta(data, patrones[0])
                return [(patrones[0], valor)] if valor is not None else []
            # Patrones genéricos (OR): si hay varios, unimos resultados de cada uno
            gen_pats = patrones if patrones else [patron]
            for pat in gen_pats:
                plow = pat.lower().strip()
                # Soporte especial autolock
                if plow=='autolock':
                    t=extraer_timeout_autolock(data)
                    if t is not None:
                        resultados.append(("behaviours[id=autolock].config.timeout", t))
                    continue
                parcial=[]
                def rec(nodo,r):
                    if isinstance(nodo,dict):
                        if 'id' in nodo and isinstance(nodo['id'],(str,int)) and plow in str(nodo['id']).lower(): parcial.append((r+"[id="+str(nodo['id'])+"]", nodo))
                        for k,v in nodo.items():
                            rk=f"{r}.{k}" if r else k
                            if plow==k.lower(): parcial.append((rk,v))
                            elif isinstance(v,(str,int,float,bool)) and plow in str(v).lower() and plow!=k.lower(): parcial.append((rk,v))
                            else: rec(v,rk)
                    elif isinstance(nodo,list):
                        for i,el in enumerate(nodo): rec(el,f"{r}[{i}]" if r else f"[{i}]")
                rec(data,'')
                # Filtro especial timeout
                if plow=='timeout':
                    parcial=[(r,v) for r,v in parcial if r.lower().endswith('timeout') and not isinstance(v,(dict,list))]
                resultados.extend(parcial)
            # Eliminar duplicados exactos (ruta, valor) manteniendo orden
            vistos=set()
            dedup=[]
            for r,v in resultados:
                clave=(r, json.dumps(v, sort_keys=True, ensure_ascii=False) if isinstance(v,(dict,list)) else v)
                if clave in vistos: continue
                vistos.add(clave); dedup.append((r,v))
            return dedup
        # Fixed rate limit to avoid flooding endpoints: 1 request per second.
        min_interval = 1.0
        last_request_ts = 0.0

        def throttle_if_needed() -> bool:
            """Sleep as needed to respect RPS; return False if cancelled."""
            nonlocal last_request_ts
            # First request: no sleep
            if last_request_ts <= 0:
                last_request_ts = time.time()
                return True
            target = last_request_ts + min_interval
            while True:
                try:
                    if self._batch_cancel_event.is_set():
                        return False
                except Exception:
                    pass
                now = time.time()
                remaining = target - now
                if remaining <= 0:
                    last_request_ts = time.time()
                    return True
                time.sleep(min(0.1, remaining))

        cancelled = False
        for row in filas:
            try:
                if self._batch_cancel_event.is_set():
                    cancelled = True
                    break
            except Exception:
                pass

            # Ignorar filas completamente vacías o con solo espacios
            if not row or all((str(cell).strip() == "" for cell in row)):
                continue
            if len(row) < 2:
                continue

            def normalize_cell(v) -> str:
                s = "" if v is None else str(v)
                # Remove common invisible chars coming from Excel/exports
                s = (s
                     .replace('\ufeff', '')
                     .replace('\u00a0', ' ')
                     .replace('\u200b', '')
                     .replace('\u200e', '')
                     .replace('\u200f', '')
                     .strip())
                return s

            def normalize_country(v) -> str:
                s = normalize_cell(v).upper()
                # Keep only letters (handles NBSP etc)
                return re.sub(r'[^A-Z]', '', s)

            def normalize_installation(v) -> str:
                s = normalize_cell(v)
                # Excel sometimes gives floats like 1996808.0; also remove any non-digits
                if s.endswith('.0') and s.replace('.', '', 1).isdigit():
                    s = s[:-2]
                return re.sub(r'\D', '', s)

            def looks_like_installation_token(token: str) -> bool:
                token = (token or "").strip()
                return token.isdigit() and len(token) >= 5

            def infer_indices_from_row(rw) -> tuple[Optional[int], Optional[int]]:
                # Scan the entire row for best country/install candidates
                country_candidates = []
                install_candidates = []
                for i, cell in enumerate(rw):
                    ctry = normalize_country(cell)
                    inst = normalize_installation(cell)
                    if ctry and map_country(ctry) is not None:
                        country_candidates.append(i)
                    if inst and looks_like_installation_token(inst):
                        install_candidates.append(i)
                # Choose first non-conflicting pair
                for ci in country_candidates:
                    for ii in install_candidates:
                        if ci != ii:
                            return ci, ii
                return None, None

            # Detección robusta de columnas por nombre o posición
            if not header_processed:
                head = [str(c).strip().upper() for c in row]
                idx_country = None
                idx_install = None
                for i, col in enumerate(head):
                    norm = col.replace('_','').replace(' ','').replace('.','')
                    if norm in ("DSCOUNTRYSHORT","COUNTRY","PAIS","GEOCOUNTRYISOCODE","COUNTRYISOCODE"):
                        idx_country = i
                    if norm in ("IDINSTALL","INSTALL","ID","INSTALLATION","SDINSTALLATIONNUMBER","INSTALLATIONNUMBER"):
                        idx_install = i
                header_processed = True

                # If headers were detected, skip this header row.
                if idx_country is not None and idx_install is not None:
                    continue

                # No recognized headers: decide whether this first row is data.
                ci, ii = infer_indices_from_row(row)
                if ci is not None and ii is not None:
                    idx_country, idx_install = ci, ii
                    # fall through to process this row as data
                else:
                    # Unknown header row: skip it and infer indices from the first data row later.
                    idx_country, idx_install = None, None
                    continue

            # If indices are still unknown (unknown header or weird files), infer from this row.
            if idx_country is None or idx_install is None:
                ci, ii = infer_indices_from_row(row)
                if ci is not None and ii is not None:
                    idx_country, idx_install = ci, ii
                else:
                    # Last resort: assume country then installation
                    idx_country = 0
                    idx_install = 1 if len(row) > 1 else 0
            # Limpieza de espacios y mayúsculas/minúsculas en datos
            inst = normalize_installation(row[idx_install]) if idx_install is not None and idx_install < len(row) else None
            country_raw = normalize_country(row[idx_country]) if idx_country is not None and idx_country < len(row) else None
            # Solo cuenta como total si la fila no está vacía
            total += 1
            mapped = map_country(country_raw) if country_raw else None
            # Solo marca como inválido si realmente falta país o instalación Y la fila no está vacía
            if not inst or not mapped or mapped not in PAISES:
                self._q.put(("batchsearchprog", f"{inst or '-'}, {country_raw or '-'} invalid", f"{total} processed"))
                rutas_inval += 1
                continue
            try:
                try:
                    if self._batch_cancel_event.is_set():
                        cancelled = True
                        break
                except Exception:
                    pass

                if not throttle_if_needed():
                    cancelled = True
                    break

                data, filename = descargar_instalacion(inst, mapped)
                ok += 1
                resultados_locales = buscar_patron(data)
                if resultados_locales:
                    for ruta_match, valor in resultados_locales:
                        if isinstance(valor, list):
                            vtxt = json.dumps(valor, ensure_ascii=False)
                            if len(vtxt) > 500:
                                vtxt = vtxt[:500] + f"... (len={len(valor)})"
                            self._q.put(("batchsearchprog", f"{inst}({mapped}) -> {ruta_match} = {vtxt}", f"{total} processed"))
                        elif isinstance(valor, dict):
                            self._q.put(("batchsearchprog", f"{inst}({mapped}) -> {ruta_match} = <dict>", f"{total} processed"))
                        else:
                            self._q.put(("batchsearchprog", f"{inst}({mapped}) -> {ruta_match} = {valor}", f"{total} processed"))
                        matches += 1
                else:
                    self._q.put(("batchsearchprog", f"{inst}({mapped}) -> no matches", f"{total} processed"))
                if not self.save_var.get() and os.path.isfile(filename):
                    try:
                        os.remove(filename)
                    except OSError:
                        pass
            except Exception as e:
                errores += 1
                self._q.put(("batchsearchprog", f"{inst},{mapped}: ERROR {e}", f"{total} processed"))
        prefix = "Cancelled" if cancelled else "Total"
        resumen=f"{prefix} rows processed:{total}, OK:{ok}, Errors:{errores}, Invalid:{rutas_inval}, Matches:{matches}"
        if cancelled:
            self._q.put(("batchcancelled", resumen))
        else:
            self._q.put(("batchsearchdone", resumen))

    # === Vista árbol ===
    def toggle_tree_view(self):
        # Tree view is a single-pane mode; exit Compare mode if active
        self._set_compare_mode(False)
        if self.tree_visible.get():
            self._ensure_treeview_style(True)
            # Switch view without repacking/reflowing the layout
            try:
                self.tree_frame.tkraise()
            except Exception:
                pass
            if self._last_json_data is not None:
                self._cargar_tree(self._last_json_data)
            else:
                self._show_tree_placeholder()
        else:
            self._ensure_treeview_style(False)
            try:
                self.text_frame.tkraise()
            except Exception:
                pass

    def _cargar_tree(self, data):
        for it in self.tree.get_children(): self.tree.delete(it)
        self._tree_path_map.clear()
        self._tree_original_labels.clear(); self._tree_search_results=[]; self._tree_search_pos=-1
        root_id = self.tree.insert('', 'end', text='ROOT', open=True); self._tree_path_map['']=root_id
        def add(parent_id, value, path):
            if isinstance(value, dict):
                for k,v in value.items():
                    child_path = f"{path}.{k}" if path else k
                    if isinstance(v,(dict,list)):
                        nid=self.tree.insert(parent_id,'end',text=k,open=False)
                        self._tree_path_map[child_path]=nid
                        add(nid,v,child_path)
                    else:
                        # hoja directa
                        leaf_text = f"{k} = {v}"
                        nid=self.tree.insert(parent_id,'end',text=leaf_text,open=False)
                        self._tree_path_map[child_path]=nid
                        self._tree_original_labels[nid]=leaf_text
                return
            if isinstance(value,list):
                for i,el in enumerate(value):
                    child_path = f"{path}[{i}]" if path else f"[{i}]"
                    if isinstance(el,(dict,list)):
                        nid=self.tree.insert(parent_id,'end',text=f"[{i}]",open=False)
                        self._tree_path_map[child_path]=nid
                        add(nid,el,child_path)
                    else:
                        leaf_text = f"[{i}] = {el}"
                        nid=self.tree.insert(parent_id,'end',text=leaf_text,open=False)
                        self._tree_path_map[child_path]=nid
                        self._tree_original_labels[nid]=leaf_text
                return
            if not isinstance(value,(dict,list)):
                leaf_text = f"{path or '<root>'} = {value}"
                nid=self.tree.insert(parent_id,'end',text=leaf_text,open=False)
                self._tree_original_labels[nid]=leaf_text
        add(root_id, data, '')
        for c in self.tree.get_children(root_id): self.tree.item(c, open=False)

    # === Rutas ===
    def mostrar_rutas(self):
        if self._last_json_data is None:
            messagebox.showinfo("Paths","No JSON loaded")
            return
        rutas = []
        def rec(nodo, path):
            if isinstance(nodo, dict):
                for k,v in nodo.items():
                    cp = f"{path}.{k}" if path else k
                    if isinstance(v,(dict,list)):
                        rec(v, cp)
                    else:
                        rutas.append(f"{cp} = {v}")
            elif isinstance(nodo, list):
                for i,el in enumerate(nodo):
                    cp = f"{path}[{i}]" if path else f"[{i}]"
                    if isinstance(el,(dict,list)):
                        rec(el, cp)
                    else:
                        rutas.append(f"{cp} = {el}")
            else:
                rutas.append(f"{path or '<root>'} = {nodo}")
        rec(self._last_json_data, '')
        if not rutas:
            messagebox.showinfo("Paths","No scalar leaves found")
            return
        # Mostrar en Text (guardando original si no se había guardado ya)
        if self._original_text is None:
            self._original_text = self.text.get('1.0','end')
        self.text.delete('1.0','end')
        self.text.insert('end', "# Paths ("+str(len(rutas))+" leaves)\n"+"\n".join(rutas))
        self.status_var.set("Paths shown in panel")

    # === Búsqueda en Treeview ===
    def _tree_limpiar_busqueda(self):
        for item,label in self._tree_original_labels.items():
            try: self.tree.item(item, text=label)
            except Exception as e: pass
        self._tree_search_results=[]; self._tree_search_pos=-1

    def _tree_buscar(self, patron: str):
        self._tree_limpiar_busqueda()
        patron = patron.strip()
        if not patron:
            self.status_var.set("Empty pattern")
            return
        patron_lower = patron.lower()
        # Recorrer todos los items
        def collect(parent=''):
            for item in self.tree.get_children(parent):
                text = self.tree.item(item,'text')
                if patron_lower in text.lower():
                    self._tree_search_results.append(item)
                collect(item)
        collect('')
        if not self._tree_search_results:
            self.status_var.set("No matches")
            return
        # Marcar visualmente
        for item in self._tree_search_results:
            original = self.tree.item(item,'text')
            if item not in self._tree_original_labels:
                self._tree_original_labels[item]=original
            if not original.startswith('>> '):
                self.tree.item(item, text='>> '+original)
        self._tree_search_pos=0
        self._tree_marcar_actual()
        self.status_var.set(f"{len(self._tree_search_results)} matches")

    def _tree_marcar_actual(self):
        if self._tree_search_pos<0 or self._tree_search_pos>=len(self._tree_search_results): return
        item = self._tree_search_results[self._tree_search_pos]
        # Expandir ancestros
        anc = []
        pid = self.tree.parent(item)
        while pid:
            anc.append(pid); pid = self.tree.parent(pid)
        for a in anc: self.tree.item(a, open=True)
        self.tree.selection_set(item)
        self.tree.focus(item)
        # Asegurar visibilidad: no hay scroll horizontal programático fácil; se asume vertical visible.

    def _tree_focus_path(self, ruta: str) -> bool:
        ruta = ruta.strip()
        if ruta in self._tree_path_map:
            item = self._tree_path_map[ruta]
            pid = self.tree.parent(item)
            while pid:
                self.tree.item(pid, open=True)
                pid = self.tree.parent(pid)
            self.tree.selection_set(item); self.tree.focus(item)
            return True
        # Intento tolerante: quitar partes finales hasta encontrar
        partes = ruta.split('.')
        while partes:
            parcial = '.'.join(partes)
            if parcial in self._tree_path_map:
                item = self._tree_path_map[parcial]
                pid = self.tree.parent(item)
                while pid:
                    self.tree.item(pid, open=True)
                    pid = self.tree.parent(pid)
                self.tree.selection_set(item); self.tree.focus(item)
                return True
            partes.pop()
        # Selector con id dentro de []: behaviours[id=autolock].config.timeout
        if '[' in ruta and ']' in ruta:
            base, _, resto = ruta.partition('[')
            selector = '[' + resto
            if selector.lower().startswith('[id=') and selector.endswith(']'):
                posible = base.strip()
                # Recorrer hijos del posible para buscar id
                if posible in self._tree_path_map:
                    parent_item = self._tree_path_map[posible]
                    objetivo = selector[4:-1].lower()
                    for child in self.tree.get_children(parent_item):
                        text = self.tree.item(child,'text').lower()
                        if ('id = ' in text and objetivo in text) or (text.startswith('id =') and text.endswith(objetivo)):
                            self.tree.item(parent_item, open=True)
                            self.tree.selection_set(child); self.tree.focus(child)
                            return True
        return False

    def _tree_apply_filter(self, patron: str):
        patron_lower = patron.lower().strip()
        if not patron_lower:
            self._cargar_tree(self._last_json_data)
            self.status_var.set("Empty filter")
            return
        # Construir subconjunto
        def matches_scalar(val):
            if isinstance(val,(str,int,float,bool)):
                return patron_lower in str(val).lower()
            return False
        def build(node):
            if isinstance(node,dict):
                nuevo={}
                for k,v in node.items():
                    if patron_lower in k.lower() or matches_scalar(v):
                        nuevo[k]=v; continue
                    sub=build(v)
                    if sub is not None:
                        nuevo[k]=sub
                return nuevo if nuevo else None
            if isinstance(node,list):
                nuevo=[]
                for el in node:
                    if matches_scalar(el):
                        nuevo.append(el); continue
                    sub=build(el)
                    if sub is not None:
                        nuevo.append(sub)
                return nuevo if nuevo else None
            return node if matches_scalar(node) else None
        filtrado = build(self._last_json_data)
        if filtrado is None:
            # Nothing matches -> empty tree
            for it in self.tree.get_children(): self.tree.delete(it)
            root_id = self.tree.insert('', 'end', text='ROOT', open=True)
            self.status_var.set("No matches")
            return
        self._cargar_tree(filtrado)
        self.status_var.set("Filtered tree")
    # Auto export eliminado
    # === Auto export helper ===
    # Método de auto export eliminado

    def run(self):
        self.root.minsize(780,520); self.root.mainloop()


def main():
    """Entry point with clean interrupt handling.
    Ctrl+C in the console does not show a traceback; exits cleanly."""
    root = tk.Tk()
    app = InstalacionGUI(root)
    try:
        app.run()
    except KeyboardInterrupt:
        # Cierre limpio sin traceback
        try:
            root.destroy()
        except Exception:
            pass
        print("Manual interrupt. Clean exit.")
    except Exception as e:
        # Log mínimo para errores inesperados
        print("Error inesperado:", e)
        print(traceback.format_exc())

if __name__ == '__main__':
    main()
