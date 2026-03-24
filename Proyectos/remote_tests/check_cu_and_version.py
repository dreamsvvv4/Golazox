#!/usr/bin/env python3
"""Find CU IP (by MAC/serial) and run CheckVersion.sh via SSH.

Usage examples:
  # Find by MAC and run check script using key file
  python check_cu_and_version.py --mac 00:23:c1:2d:d9:1c --key "Proyectos/cu_devkey_private"

  # Find by serial saved in network_scanner.MAC_INFO
  python check_cu_and_version.py --serial 26QXL99B --key "Proyectos/cu_devkey_private"

  # Direct host
  python check_cu_and_version.py --host 192.168.1.155 --key "Proyectos/cu_devkey_private"

The script prefers Paramiko; falls back to system ssh with rsa options if needed.
Prints JSON with results for each target.
"""
from __future__ import annotations
import argparse
import importlib.util
import json
import re
import os
import shlex
import subprocess
import sys
import contextlib
from typing import Optional, List, Tuple
import difflib
import unicodedata


def load_network_scanner(module_path: str):
    spec = importlib.util.spec_from_file_location('network_scanner', module_path)
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


def find_ip_for_mac(scanner_mod, mac: str, subnet: Optional[str] = None) -> Optional[str]:
    # normalize mac
    if not subnet:
        subnet = scanner_mod.detectar_subred_local()
    return scanner_mod.find_ip_from_mac(mac, subnet)


def find_ip_for_serial(scanner_mod, serial: str) -> Optional[Tuple[str, str]]:
    # Look up MAC_INFO for serial -> mac, then find ip
    for mac, info in getattr(scanner_mod, 'MAC_INFO', {}).items():
        if info.get('Serial Number') == serial or info.get('Numero de instalacion') == serial:
            ip = find_ip_for_mac(scanner_mod, mac, scanner_mod.detectar_subred_local())
            return (ip, mac)
    return None


def run_paramiko_cmd(host: str, user: str, port: int, key_path: Optional[str], passphrase: Optional[str], use_agent: bool, remote_cmd: str, timeout: int = 20):
    try:
        import paramiko
    except Exception:
        return {'error': 'paramiko-missing'}

    client = paramiko.SSHClient()
    client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    try:
        pkey = None
        if key_path:
            # Let paramiko handle key file detection
            pass

        client.connect(hostname=host, port=port, username=user, key_filename=key_path, allow_agent=use_agent, look_for_keys=use_agent, timeout=timeout)
        stdin, stdout, stderr = client.exec_command(remote_cmd, timeout=timeout)
        out = stdout.read().decode('utf-8', errors='ignore').strip()
        err = stderr.read().decode('utf-8', errors='ignore').strip()
        rc = stdout.channel.recv_exit_status()
        client.close()
        return {'rc': rc, 'stdout': out, 'stderr': err}
    except Exception as e:
        try:
            client.close()
        except Exception:
            pass
        return {'error': str(e)}


def run_ssh_subprocess(host: str, user: str, port: int, key_path: Optional[str], use_agent: bool, remote_cmd: str, timeout: int = 20):
    # Build ssh command with rsa accept options to support older devices
    opts = [
        '-o', 'HostKeyAlgorithms=+ssh-rsa',
        '-o', 'PubkeyAcceptedKeyTypes=+ssh-rsa',
        '-o', 'StrictHostKeyChecking=no',
        '-o', 'UserKnownHostsFile=NUL'
    ]
    cmd = ['ssh'] + opts + ['-p', str(port)]
    if key_path:
        cmd += ['-i', key_path]
    cmd += [f'{user}@{host}', remote_cmd]
    try:
        out = subprocess.run(cmd, capture_output=True, text=True, timeout=timeout)
        return {'rc': out.returncode, 'stdout': out.stdout.strip(), 'stderr': out.stderr.strip()}
    except Exception as e:
        return {'error': str(e)}


def _parse_version(v: Optional[str]):
    if not v:
        return None
    v = str(v).strip()
    try:
        from packaging.version import parse as _p
        return _p(v)
    except Exception:
        parts = re.findall(r"\d+", v)
        if not parts:
            return v
        return tuple(int(x) for x in parts)


def _compare_versions(a: Optional[str], b: Optional[str]) -> Optional[int]:
    """Return 1 if a>b, 0 if equal, -1 if a<b, None if not comparable"""
    if a is None or b is None:
        return None
    pa = _parse_version(a)
    pb = _parse_version(b)
    try:
        if pa is None or pb is None:
            return None
        if pa > pb:
            return 1
        if pa < pb:
            return -1
        return 0
    except Exception:
        try:
            if str(a) > str(b):
                return 1
            if str(a) < str(b):
                return -1
            return 0
        except Exception:
            return None


def load_latest_from_excel(path: str):
    """Load versions mapping and global latest from an Excel file.
    Returns (mapping_by_serial, global_latest)
    mapping_by_serial: dict keyed by serial (lowercased) -> version string
    global_latest: best effort latest version string found in file
    """
    mapping = {}
    versions = []
    def _clean_cell(x):
        if x is None:
            return None
        s = str(x).strip()
        # remove surrounding quotes if present
        if (s.startswith("'") and s.endswith("'")) or (s.startswith('"') and s.endswith('"')):
            s = s[1:-1].strip()
        # strip any stray leading/trailing quotes/spaces
        s = s.strip().strip('\'"')
        return s
    # if a directory provided, pick the first Excel file inside
    try:
        if os.path.isdir(path):
            files = [f for f in os.listdir(path) if f.lower().endswith(('.xlsx', '.xls'))]
            if files:
                path = os.path.join(path, files[0])
            else:
                return mapping, None
    except Exception:
        pass
    try:
        try:
            from openpyxl import load_workbook
        except Exception:
            try:
                import pandas as pd
                df = pd.read_excel(path, sheet_name=0)
            except Exception:
                return mapping, None
            cols = [ _clean_cell(c).lower() for c in df.columns.astype(str) ]
            ver_cols = [i for i, c in enumerate(cols) if 'version' in c or 'fw' in c or 'tag' in c]
            serial_cols = [i for i, c in enumerate(cols) if 'serial' in c or c == 'sn' or 'installation' in c or 'instalacion' in c or 'numero' in c or 'column' in c or 'component' in c or 'device' in c]
            for idx, row in df.iterrows():
                ver = None
                ser = None
                if ver_cols:
                    raw = row.iloc[ver_cols[0]]
                    ver = _clean_cell(raw) if pd.notna(raw) else None
                if serial_cols:
                    raw = row.iloc[serial_cols[0]]
                    ser = _clean_cell(raw) if pd.notna(raw) else None
                if ver:
                    versions.append(ver)
                if ser and ver:
                    mapping[str(ser).strip().lower()] = str(ver).strip()
            if versions:
                best = versions[0]
                for v in versions[1:]:
                    c = _compare_versions(v, best)
                    if c is None:
                        continue
                    if c == 1:
                        best = v
                return mapping, best
            return mapping, None

        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = ws.iter_rows(values_only=True)
        header = None
        for row in rows:
            if not header:
                header = [ _clean_cell(c).lower() if c is not None else '' for c in row ]
                ver_idx = None
                serial_idx = None
                model_idx = None
                for i, h in enumerate(header):
                    if any(x in h for x in ('version', 'vers', 'fw', 'tag')) and ver_idx is None:
                        ver_idx = i
                    if any(x in h for x in ('serial', 'sn', 'numero', 'installation', 'instalacion')) and serial_idx is None:
                        serial_idx = i
                # Prefer explicit 'column'/'column1' header for model, otherwise use other heuristics
                for i, h in enumerate(header):
                    if 'column' in h:
                        model_idx = i
                        break
                if model_idx is None:
                    for i, h in enumerate(header):
                        if any(x in h for x in ('component', 'device', 'model', 'teamslink')):
                            model_idx = i
                            break
                # fallback: use last column as model if nothing else matched
                if model_idx is None and len(header) > 0:
                    model_idx = len(header) - 1
                continue
            if not any(row):
                continue
            ver = None
            ser = None
            if 'ver_idx' in locals() and ver_idx is not None and ver_idx < len(row):
                ver = _clean_cell(row[ver_idx])
            else:
                for cell in row:
                    if cell and re.search(r"\d+\.\d+", str(cell)):
                        ver = _clean_cell(cell)
                        break
            if 'serial_idx' in locals() and serial_idx is not None and serial_idx < len(row):
                ser = _clean_cell(row[serial_idx])
            model = None
            if 'model_idx' in locals() and model_idx is not None and model_idx < len(row):
                model = _clean_cell(row[model_idx])
            if ver is not None:
                versions.append(str(ver).strip())
            if ser is not None and ver is not None:
                mapping[str(ser).strip().lower()] = str(ver).strip()
            if model is not None and ver is not None:
                mapping[str(model).strip().lower()] = str(ver).strip()
        best = None
        for v in versions:
            if best is None:
                best = v
                continue
            c = _compare_versions(v, best)
            if c is None:
                continue
            if c == 1:
                best = v
        return mapping, best
    except Exception:
        return mapping, None


def load_device_version_json(path: Optional[str] = None):
    """Load Device Version.json and return list of entries.
    Default path is next to this script: 'Device Version.json'
    """
    if not path:
        path = os.path.join(os.path.dirname(__file__), 'Device Version.json')
    try:
        with open(path, 'r', encoding='utf-8') as f:
            data = json.load(f)
            return data
    except Exception:
        return []


def main(argv=None):
    ap = argparse.ArgumentParser(description='Find CU IP and run CheckVersion.sh')
    ap.add_argument('--host', help='Direct host IP')
    ap.add_argument('--mac', help='MAC address to search (AA:BB:CC:DD:EE:FF)')
    ap.add_argument('--serial', help='Serial number as in MAC_INFO')
    ap.add_argument('--key', help='Path to private key file', default=os.path.join(os.path.dirname(__file__), 'cu_devkey_private'))
    ap.add_argument('--user', default='root')
    ap.add_argument('--port', type=int, default=22)
    ap.add_argument('--use-agent', action='store_true')
    ap.add_argument('--verbose', action='store_true', help='Show detailed scanner output (disabled by default)')
    # Remote commands embedded in code (no need to pass at runtime)
    # The script will try these in order and use the first that returns output.
    REMOTE_CMDS = ['checkVersion.sh']
    # Command to list nodes from the CU sqlite DB and busctl (runs on the CU)
    NODE_LIST_CMD = r'''db=/mnt/extra/vsdb_cu.sqlite3; fmt="%-10s %-6s %-12s %-10s %-8s %-9s %-18s
"; printf "$fmt" sn zone config_label fw hw is_active location; sqlite3 -noheader -batch "$db" "SELECT label AS sn, COALESCE(zone,''), COALESCE(config_label,''), COALESCE(is_active,0), COALESCE(location,'') FROM devices ORDER BY sn;" | while IFS='|' read -r sn zone cfg active location; do fw=$(busctl call com.verisure.cuxs-core /node com.verisure.Node GetNodeFirmwareVersion s "$sn" 2>/dev/null | awk -F'"' '{print $2}'); hw=$(busctl call com.verisure.cuxs-core /node com.verisure.Node GetNodeHardwareVersion s "$sn" 2>/dev/null | awk -F'"' '{print $2}'); printf "$fmt" "$sn" "$zone" "$cfg" "$fw" "$hw" "$active" "$location"; done; echo'''
    ap.add_argument('--timeout', type=int, default=20)
    ap.add_argument('--show-tag-only', action='store_true', help='Print only TAG_VERSION per host')
    ap.add_argument('--pretty', action='store_true', help='Print a pretty table of Host, Serial, Installation, TAG_VERSION')
    ap.add_argument('--export', help='Write full results to JSON file')
    ap.add_argument('--releases-excel', help='Path to Excel file listing releases/versions (optional)')
    ap.add_argument('--match-cutoff', type=float, default=0.6, help='Fuzzy match cutoff for matching device names (0.0-1.0)')
    ap.add_argument('--dump-scanner', action='store_true', help='Print network_scanner.MAC_INFO and exit (debug)')
    args = ap.parse_args(argv)

    def _normalize_for_match(s: Optional[str]):
        if not s:
            return ''
        if not isinstance(s, str):
            s = str(s)
        s = s.strip().lower()
        # remove accents
        s = unicodedata.normalize('NFKD', s)
        s = ''.join(ch for ch in s if not unicodedata.combining(ch))
        # keep only alphanum
        s = re.sub(r'[^a-z0-9]', '', s)
        return s

    # If the script is executed without flags (no host/mac/serial and no output flags),
    # default to pretty output so it can be run directly.
    if not any([args.host, args.mac, args.serial, args.pretty, args.show_tag_only, args.export, args.verbose]):
        args.pretty = True

    scanner_path = os.path.join(os.path.dirname(__file__), 'network_scanner.py')
    scanner = load_network_scanner(scanner_path)

    # If user didn't provide --releases-excel, try a sensible default in OneDrive
    if not args.releases_excel:
        try:
            home = os.path.expanduser('~')
            default_candidates = [
                os.path.join(home, 'OneDrive - Verisure', 'Releases'),
                os.path.join(os.path.dirname(__file__), '..', 'Releases')
            ]
            for cand in default_candidates:
                try:
                    if os.path.isdir(cand):
                        files = [f for f in os.listdir(cand) if f.lower().endswith(('.xlsx', '.xls'))]
                        if files:
                            args.releases_excel = cand
                            break
                except Exception:
                    continue
        except Exception:
            pass

    if args.dump_scanner:
        try:
            info = getattr(scanner, 'MAC_INFO', {})
            print(json.dumps({'MAC_INFO_len': len(info), 'sample_keys': list(info.keys())[:50]}, indent=2))
        except Exception as e:
            print(json.dumps({'error': str(e)}))
        return 0

    targets = []
    if args.host:
        targets.append({'host': args.host, 'mac': None})
    elif args.mac:
        if args.verbose:
            ip = find_ip_for_mac(scanner, args.mac)
        else:
            devnull = open(os.devnull, 'w')
            with contextlib.redirect_stdout(devnull), contextlib.redirect_stderr(devnull):
                ip = find_ip_for_mac(scanner, args.mac)
            devnull.close()
        targets.append({'host': ip, 'mac': args.mac})
    elif args.serial:
        found = find_ip_for_serial(scanner, args.serial)
        if found:
            ip, mac = found
            targets.append({'host': ip, 'mac': mac})
        else:
            print(json.dumps({'error': 'serial-not-found'}))
            return 2
    else:
        # Use only stored serials/MACs in network_scanner.MAC_INFO
        for mac, info in getattr(scanner, 'MAC_INFO', {}).items():
            # find_ip_for_mac (and the scanner) prints a lot; silence unless verbose
            if args.verbose:
                ip = find_ip_for_mac(scanner, mac)
            else:
                devnull = open(os.devnull, 'w')
                with contextlib.redirect_stdout(devnull), contextlib.redirect_stderr(devnull):
                    ip = find_ip_for_mac(scanner, mac)
                devnull.close()
            # Append discovered target (ip may be None if not found)
            targets.append({'host': ip, 'mac': mac, 'serial': info.get('Serial Number'), 'installation': info.get('Numero de instalacion')})

    # Load releases/versions mapping from Excel if requested
    latest_map = {}
    global_latest = None
    latest_source = None
    if args.releases_excel:
        rx = args.releases_excel
        # If a directory provided, try to pick the first Excel file inside
        if os.path.isdir(rx):
            files = [f for f in os.listdir(rx) if f.lower().endswith(('.xlsx', '.xls'))]
            if files:
                rx = os.path.join(rx, files[0])
        mapping, best = load_latest_from_excel(rx)
        latest_map = mapping or {}
        # normalize excel-derived keys for robust matching
        try:
            latest_map = { _normalize_for_match(k): v for k, v in (latest_map or {}).items() }
        except Exception:
            latest_map = { str(k).strip().lower(): v for k, v in (latest_map or {}).items() }
        global_latest = best
        latest_source = f"excel:{rx}"

        # Try to enrich mapping using Device Version.json entries (Entity / FOTAxS)
        devs = load_device_version_json()
        # load optional aliases file to map entities to excel keys
        aliases = {}
        try:
            aliases_path = os.path.join(os.path.dirname(__file__), 'device_aliases.json')
            if os.path.isfile(aliases_path):
                with open(aliases_path, 'r', encoding='utf-8') as af:
                    aliases = json.load(af)
        except Exception:
            aliases = {}
        # sensible defaults when aliases not provided
        default_aliases = {
            'orion': 'camera1',
            'aquila': 'camera2',
            'cu': 'cuxs'
        }
        if devs:
            # normalize keys from excel mapping (again) to ensure consistent form
            excel_keys = { _normalize_for_match(k): v for k, v in (latest_map or {}).items() }
            for entry in devs:
                ent = (entry.get('Entity / FOTAxS') or entry.get('Entity') or '').strip()
                name = (entry.get('Name') or '').strip()
                if not ent and not name:
                    continue
                ent_k = _normalize_for_match(ent)
                name_k = _normalize_for_match(name)
                found = None
                # direct match against excel keys
                if ent_k and ent_k in excel_keys:
                    found = excel_keys[ent_k]
                elif name_k and name_k in excel_keys:
                    found = excel_keys[name_k]
                else:
                    # substring match
                    for k in excel_keys.keys():
                        if ent_k and (ent_k in k or k in ent_k):
                            found = excel_keys[k]
                            break
                        if name_k and (name_k in k or k in name_k):
                            found = excel_keys[k]
                            break
                # fuzzy match
                if not found and ent_k:
                    km = difflib.get_close_matches(ent_k, list(excel_keys.keys()), n=1, cutoff=args.match_cutoff)
                    if km:
                        found = excel_keys[km[0]]
                if not found and name_k:
                    km = difflib.get_close_matches(name_k, list(excel_keys.keys()), n=1, cutoff=args.match_cutoff)
                    if km:
                        found = excel_keys[km[0]]

                if found:
                    latest_map[ent_k] = found
                    if name_k:
                        latest_map[name_k] = found
                else:
                    # try alias fallback: map common device entities to excel camera keys
                    # prefer user-provided aliases, then defaults
                    aliask = aliases.get(ent.lower()) if aliases else None
                    if not aliask:
                        aliask = default_aliases.get(ent_k)
                    if aliask:
                        aliask_n = _normalize_for_match(aliask)
                        if aliask_n in excel_keys:
                            latest_map[ent_k] = excel_keys[aliask_n]
                            if name_k:
                                latest_map[name_k] = excel_keys[aliask_n]

    results = []
    # if no targets were discovered, give a helpful debug message
    for t in targets:
        host = t.get('host')
        if not host:
            results.append({'host': None, 'error': 'ip-not-found', 'mac': t.get('mac'), 'serial': t.get('serial')})
            continue

        # Try embedded remote commands in order and stop at first successful output
        cmd_result = None
        for remote_cmd in REMOTE_CMDS:
            res = run_paramiko_cmd(host, args.user, args.port, args.key, None, args.use_agent, remote_cmd, timeout=args.timeout)
            if isinstance(res, dict) and res.get('error') == 'paramiko-missing':
                # fallback to ssh subprocess
                res2 = run_ssh_subprocess(host, args.user, args.port, args.key, args.use_agent, remote_cmd, timeout=args.timeout)
                if res2.get('stdout'):
                    cmd_result = {'host': host, 'mac': t.get('mac'), 'method': 'ssh-subprocess', 'command': remote_cmd, 'result': res2}
                    break
                continue

            if isinstance(res, dict) and res.get('error'):
                # try ssh subprocess as fallback
                res2 = run_ssh_subprocess(host, args.user, args.port, args.key, args.use_agent, remote_cmd, timeout=args.timeout)
                if res2.get('stdout'):
                    cmd_result = {'host': host, 'mac': t.get('mac'), 'method': 'ssh-subprocess', 'command': remote_cmd, 'result': res2}
                    break
                else:
                    continue

            # Paramiko succeeded; check output
            if res.get('stdout'):
                cmd_result = {'host': host, 'mac': t.get('mac'), 'method': 'paramiko', 'command': remote_cmd, 'result': res}
                break

        if not cmd_result:
            results.append({'host': host, 'mac': t.get('mac'), 'serial': t.get('serial'), 'error': 'no-command-output'})
        else:
            # parse TAG_VERSION from stdout if present
            stdout_text = ''
            if isinstance(cmd_result.get('result'), dict):
                stdout_text = cmd_result['result'].get('stdout', '') or ''
            elif isinstance(cmd_result.get('result'), str):
                stdout_text = cmd_result['result']

            tag = None
            m = re.search(r'TAG_VERSION\s*=\s*"?([^"\n]+)"?', stdout_text)
            if m:
                tag = m.group(1).strip()
                cmd_result.setdefault('parsed', {})['TAG_VERSION'] = tag

            if 'serial' in t:
                cmd_result['serial'] = t.get('serial')
                cmd_result['installation'] = t.get('installation')
            # After successful checkVersion, try to collect node list from the CU
            nodes_out = None
            nodes_parsed = []
            try:
                # Try paramiko first
                nres = run_paramiko_cmd(host, args.user, args.port, args.key, None, args.use_agent, NODE_LIST_CMD, timeout=max(30, args.timeout))
                if isinstance(nres, dict) and nres.get('error') == 'paramiko-missing':
                    nres2 = run_ssh_subprocess(host, args.user, args.port, args.key, args.use_agent, NODE_LIST_CMD, timeout=max(30, args.timeout))
                    if nres2.get('stdout'):
                        nodes_out = nres2.get('stdout','')
                elif isinstance(nres, dict) and nres.get('error'):
                    nres2 = run_ssh_subprocess(host, args.user, args.port, args.key, args.use_agent, NODE_LIST_CMD, timeout=max(30, args.timeout))
                    if nres2.get('stdout'):
                        nodes_out = nres2.get('stdout','')
                else:
                    if nres.get('stdout'):
                        nodes_out = nres.get('stdout','')
            except Exception:
                nodes_out = None

            if nodes_out:
                # parse lines into columns (maxsplit=6 to preserve location)
                for line in nodes_out.splitlines():
                    if not line.strip():
                        continue
                    parts = line.split(None, 6)
                    if len(parts) < 7:
                        # skip header-like or malformed lines
                        continue
                    sn, zone, cfg, fw, hw, active, location = parts
                    # Skip header lines emitted by the remote command (e.g. a leading 'sn zone ...')
                    if str(sn).strip().lower() == 'sn':
                        continue
                    nodes_parsed.append({
                        'sn': sn, 'zone': zone, 'config_label': cfg,
                        'fw': fw, 'hw': hw, 'is_active': active, 'location': location
                    })
                cmd_result['nodes_raw'] = nodes_out
                # For each node, try to attach latest_known from latest_map
                for n in nodes_parsed:
                    # build candidate keys for node (normalized)
                    node_candidates = []
                    for fld in ('config_label', 'sn', 'location'):
                        v = n.get(fld)
                        if v:
                            node_candidates.append(_normalize_for_match(v))
                    node_latest = None
                    node_matched = None
                    # exact match
                    for c in node_candidates:
                        if _normalize_for_match(c) in latest_map:
                            node_latest = latest_map[_normalize_for_match(c)]
                            node_matched = c
                            break
                    # substring
                    if not node_latest:
                        for c in node_candidates:
                            for k in latest_map.keys():
                                if not k or not c:
                                    continue
                                if k in c or c in k:
                                    node_latest = latest_map[k]
                                    node_matched = k
                                    break
                            if node_latest:
                                break
                    # fuzzy
                    if not node_latest:
                        keys = list(latest_map.keys())
                        for c in node_candidates:
                            if not c:
                                continue
                            matches = difflib.get_close_matches(c, keys, n=1, cutoff=args.match_cutoff)
                            if matches:
                                node_latest = latest_map[matches[0]]
                                node_matched = matches[0]
                                break
                    # alias substring fallback (e.g. Orion -> camera1, Aquila -> camera2)
                    if not node_latest:
                        for c in node_candidates:
                            for akey, atarget in default_aliases.items():
                                if akey and akey in c:
                                    atn = _normalize_for_match(atarget)
                                    if atn in latest_map:
                                        node_latest = latest_map[atn]
                                        node_matched = atn
                                        break
                            if node_latest:
                                break
                    n['latest_known'] = node_latest
                    if node_latest:
                        cmp = _compare_versions(n.get('fw'), node_latest)
                        if cmp is None:
                            n['is_latest'] = (str(n.get('fw')).strip() == str(node_latest).strip())
                        else:
                            n['is_latest'] = True if cmp >= 0 else False
                    else:
                        n['is_latest'] = None
                cmd_result['nodes'] = nodes_parsed

            # If we loaded releases data from Excel, attach latest info
            if 'latest_map' in locals() and latest_map:
                latest_known = None
                # For host-level check: only evaluate CUxS mapping when we have a host and TAG_VERSION
                tag = None
                if isinstance(cmd_result.get('parsed'), dict):
                    tag = cmd_result['parsed'].get('TAG_VERSION')
                if not tag and isinstance(cmd_result.get('result'), dict):
                    s = cmd_result['result'].get('stdout','')
                    mm = re.search(r'TAG_VERSION\s*=\s*"?([^"\n]+)"?', s)
                    if mm:
                        tag = mm.group(1).strip()

                # Only set host-level latest if we have a reachable host and a parsed TAG_VERSION
                if t.get('host') and tag:
                    # prefer explicit CU key mapping (cuxs)
                    cuxs_key = _normalize_for_match('cuxs')
                    if cuxs_key in latest_map:
                        latest_known = latest_map[cuxs_key]

                # Do not fallback to global_latest for hosts without TAG or for unrelated targets
                cmd_result['latest_known'] = latest_known
                cmd_result['latest_source'] = latest_source
                if latest_known:
                    # compare
                    cmp = _compare_versions(tag, latest_known) if tag else None
                    if cmp is None:
                        cmd_result['is_latest'] = (str(tag).strip() == str(latest_known).strip()) if tag else None
                    else:
                        cmd_result['is_latest'] = True if cmp >= 0 else False
                else:
                    cmd_result['is_latest'] = None

            results.append(cmd_result)

    output = {'targets': results}
    if args.export:
        try:
            with open(args.export, 'w', encoding='utf-8') as f:
                json.dump(output, f, indent=2, ensure_ascii=False)
        except Exception as e:
            print(f"[warn] Could not write export file: {e}", file=sys.stderr)

    if args.show_tag_only:
        # Print one line per target: host serial installation TAG_VERSION
        for t in results:
            host = t.get('host')
            serial = t.get('serial')
            inst = t.get('installation')
            tag = None
            if isinstance(t.get('parsed'), dict):
                tag = t['parsed'].get('TAG_VERSION')
            # fallback: try to parse from result stdout
            if not tag and isinstance(t.get('result'), dict):
                s = t['result'].get('stdout','')
                mm = re.search(r'TAG_VERSION\s*=\s*"?([^"\n]+)"?', s)
                if mm:
                    tag = mm.group(1).strip()
            print(f"{host or '-'}\t{serial or '-'}\t{inst or '-'}\t{tag or ''}")
    else:
        # If pretty requested, print aligned table with selected fields
        if args.pretty:
            cols = ["Host", "Serial", "Installation", "TAG_VERSION", "LATEST", "STATUS"]
            widths = [18, 14, 14, 12, 12, 12]
            header = f"{cols[0]:<{widths[0]}} {cols[1]:<{widths[1]}} {cols[2]:<{widths[2]}} {cols[3]:<{widths[3]}} {cols[4]:<{widths[4]}} {cols[5]:<{widths[5]}}"
            print(header)
            print('-' * (sum(widths) + 6))
            for t in results:
                host = t.get('host') or '-'
                serial = t.get('serial') or '-'
                inst = str(t.get('installation') or '-')
                tag = None
                if isinstance(t.get('parsed'), dict):
                    tag = t['parsed'].get('TAG_VERSION')
                if not tag and isinstance(t.get('result'), dict):
                    s = t['result'].get('stdout','')
                    mm = re.search(r'TAG_VERSION\s*=\s*"?([^"\n]+)"?', s)
                    if mm:
                        tag = mm.group(1).strip()
                tag = tag or '-'
                # determine latest to display: only show latest for actual hosts
                if t.get('host'):
                    latest_disp = t.get('latest_known') or global_latest or ''
                    status = '-'
                    if isinstance(t.get('is_latest'), bool):
                        status = 'up-to-date' if t.get('is_latest') else 'outdated'
                    elif t.get('latest_known') or global_latest:
                        status = 'unknown'
                else:
                    latest_disp = ''
                    status = '-'
                print(f"{host:<{widths[0]}} {serial:<{widths[1]}} {inst:<{widths[2]}} {tag:<{widths[3]}} {str(latest_disp):<{widths[4]}} {status:<{widths[5]}}")
            # After printing versions table, also print nodes table per host if present
            for t in results:
                host = t.get('host')
                nodes = t.get('nodes')
                if not nodes:
                    continue
                print('\nNodes for host: %s' % (host or '-'))
                ncols = ["sn", "zone", "config_label", "fw", "hw", "is_active", "location", "LATEST", "STATUS"]
                # slightly smaller widths to avoid terminal wrapping
                nwidths = [10, 6, 12, 9, 6, 8, 14, 10, 10]
                nheader = f"{ncols[0]:<{nwidths[0]}} {ncols[1]:<{nwidths[1]}} {ncols[2]:<{nwidths[2]}} {ncols[3]:<{nwidths[3]}} {ncols[4]:<{nwidths[4]}} {ncols[5]:<{nwidths[5]}} {ncols[6]:<{nwidths[6]}} {ncols[7]:<{nwidths[7]}} {ncols[8]:<{nwidths[8]}}"
                print(nheader)
                print('-' * (sum(nwidths) + 8))
                for n in nodes:
                    status = '-'
                    if isinstance(n.get('is_latest'), bool):
                        status = 'up-to-date' if n.get('is_latest') else 'outdated'
                    elif n.get('latest_known'):
                        status = 'unknown'
                    latest = n.get('latest_known') or '-'
                    print(f"{n.get('sn',''):<{nwidths[0]}} {n.get('zone',''):<{nwidths[1]}} {n.get('config_label',''):<{nwidths[2]}} {n.get('fw',''):<{nwidths[3]}} {n.get('hw',''):<{nwidths[4]}} {n.get('is_active',''):<{nwidths[5]}} {n.get('location',''):<{nwidths[6]}} {latest:<{nwidths[7]}} {status:<{nwidths[8]}}")
        else:
            print(json.dumps(output, indent=2))
    return 0


if __name__ == '__main__':
    try:
        sys.exit(main())
    except KeyboardInterrupt:
        print('\nInterrupted')
        sys.exit(1)
